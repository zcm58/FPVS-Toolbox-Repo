from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
import pytest
from openpyxl import load_workbook

from Tools.Publication_Report.analysis_tables import (
    _condition_pairs_by_roi,
    _individual_detectability_frames,
    _roi_response_summary,
    _planned_lateralization_contrasts,
    _semantic_color_ratio_frames,
    _z_score_report,
)
from Tools.Publication_Report.models import (
    BASE_RATE_SUMMARY_SHEET,
    COMPARISON_AGREEMENT_SHEET,
    CONDITION_ROLES_SHEET,
    CONDITION_COMPARISONS_SHEET,
    CONDITION_PAIRS_BY_ROI_SHEET,
    ELECTRODE_Z_SCORES_SHEET,
    FIGURE_MANIFEST_SHEET,
    GROUP_ELECTRODE_SIGNIFICANCE_SHEET,
    HARMONIC_SELECTION_SHEET,
    INDIVIDUAL_DETECTABILITY_SHEET,
    INDIVIDUAL_DETECTABILITY_COUNTS_SHEET,
    INDIVIDUAL_ELECTRODE_FDR_SHEET,
    INDIVIDUAL_ELECTRODE_SUMMED_Z_SHEET,
    INDIVIDUAL_ROI_SUMMED_Z_SHEET,
    NORMALITY_CHECKS_SHEET,
    OLD_VS_NEW_DETECTABILITY_COMPARISON_SHEET,
    PARAMETRIC_VS_NONPARAMETRIC_TESTS_SHEET,
    PARTICIPANT_INCLUSION_SHEET,
    PLANNED_LATERALIZATION_SHEET,
    PLANNED_ROI_COMPARISONS_HOLM_SHEET,
    ROI_HARMONIC_SUMMARY_SHEET,
    ROI_HARMONIC_VALUES_SHEET,
    ROI_DEFINITIONS_SHEET,
    ROI_RESPONSE_SUMMARY_SHEET,
    RUN_SUMMARY_SHEET,
    SEMANTIC_COLOR_RATIO_SUMMARY_SHEET,
    SEMANTIC_COLOR_RATIO_VALUES_SHEET,
    STATISTICAL_TEST_DECISIONS_SHEET,
    STATS_POSTHOC_SHEET,
    STATS_RM_ANOVA_SHEET,
    STATS_WORKFLOW_SUMMARY_SHEET,
    WARNINGS_SHEET,
    Z_SCORE_REPORT_SHEET,
    PublicationReportRequest,
    ReportOutputOptions,
    ReportRoi,
    report_rois_from_settings_pairs,
)
from Tools.Publication_Report.runner import generate_publication_report
from Tools.Publication_Report.discovery import WorkbookEntry


def _write_result_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame = pd.DataFrame({"Electrode": ["P7"], "1.2000_Hz": [0.1]})
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="BCA (uV)", index=False)


def _write_project_manifest(root: Path, *, groups: dict[str, object] | None = None) -> None:
    payload: dict[str, object] = {
        "name": "PublicationReportProject",
        "results_folder": ".",
        "subfolders": {"excel": "1 - Excel Data Files"},
    }
    if groups is not None:
        payload["groups"] = groups
    (root / "project.json").write_text(json.dumps(payload), encoding="utf-8")


def _build_project(tmp_path: Path) -> Path:
    root = tmp_path / "project"
    root.mkdir()
    excel_root = root / "1 - Excel Data Files"
    _write_project_manifest(root)
    _write_result_workbook(excel_root / "CondA" / "P01_CondA_Results.xlsx")
    _write_result_workbook(excel_root / "CondA" / "P02_CondA_Results.xlsx")
    _write_result_workbook(excel_root / "CondB" / "P01_CondB_Results.xlsx")
    _write_result_workbook(excel_root / "CondB" / "P03_CondB_Results.xlsx")
    return root


def test_generate_publication_report_writes_initial_bundle(tmp_path: Path) -> None:
    root = _build_project(tmp_path)
    request = PublicationReportRequest(
        project_root=root,
        selected_conditions=("CondA", "CondB"),
        manual_excluded_subjects=frozenset({"P02"}),
        qc_excluded_subjects=frozenset({"P03"}),
        output_options=ReportOutputOptions(
            markdown=True,
            docx=True,
            workbook=True,
            audit_json=True,
            spectra=True,
            scalp_maps=True,
            individual_figures=True,
        ),
    )

    result = generate_publication_report(request)

    assert result.output_root == root / "5 - Publication Report"
    assert result.markdown_path is not None and result.markdown_path.exists()
    assert result.docx_path is not None and result.docx_path.exists()
    assert result.workbook_path is not None and result.workbook_path.exists()
    assert result.audit_path is not None and result.audit_path.exists()
    assert result.log_path is not None and result.log_path.exists()

    markdown_text = result.markdown_path.read_text(encoding="utf-8")
    assert "# Semantic categories Publication Report" in markdown_text
    assert "CondA, CondB" in markdown_text
    assert "1 included participant(s)" in markdown_text
    assert "LOT (primary; electrodes: P7, P9, PO7, PO3, O1)" in markdown_text
    assert "ROT (primary; electrodes: P8, P10, PO8, PO4, O2)" in markdown_text
    assert "Central (supporting/exploratory; electrodes: FCz, Cz, CPz, CP1, C1, FC1)" in markdown_text

    with ZipFile(result.docx_path) as docx:
        assert "word/document.xml" in docx.namelist()

    workbook = load_workbook(result.workbook_path, read_only=True)
    assert {
        RUN_SUMMARY_SHEET,
        PARTICIPANT_INCLUSION_SHEET,
        CONDITION_ROLES_SHEET,
        ROI_DEFINITIONS_SHEET,
        FIGURE_MANIFEST_SHEET,
        WARNINGS_SHEET,
        HARMONIC_SELECTION_SHEET,
        ROI_HARMONIC_VALUES_SHEET,
        ROI_HARMONIC_SUMMARY_SHEET,
        ROI_RESPONSE_SUMMARY_SHEET,
        SEMANTIC_COLOR_RATIO_VALUES_SHEET,
        SEMANTIC_COLOR_RATIO_SUMMARY_SHEET,
        CONDITION_COMPARISONS_SHEET,
        STATS_RM_ANOVA_SHEET,
        STATS_POSTHOC_SHEET,
        STATS_WORKFLOW_SUMMARY_SHEET,
        CONDITION_PAIRS_BY_ROI_SHEET,
        COMPARISON_AGREEMENT_SHEET,
        PLANNED_LATERALIZATION_SHEET,
        NORMALITY_CHECKS_SHEET,
        PARAMETRIC_VS_NONPARAMETRIC_TESTS_SHEET,
        PLANNED_ROI_COMPARISONS_HOLM_SHEET,
        STATISTICAL_TEST_DECISIONS_SHEET,
        ELECTRODE_Z_SCORES_SHEET,
        GROUP_ELECTRODE_SIGNIFICANCE_SHEET,
        INDIVIDUAL_DETECTABILITY_SHEET,
        INDIVIDUAL_DETECTABILITY_COUNTS_SHEET,
        INDIVIDUAL_ROI_SUMMED_Z_SHEET,
        INDIVIDUAL_ELECTRODE_SUMMED_Z_SHEET,
        INDIVIDUAL_ELECTRODE_FDR_SHEET,
        OLD_VS_NEW_DETECTABILITY_COMPARISON_SHEET,
        Z_SCORE_REPORT_SHEET,
        BASE_RATE_SUMMARY_SHEET,
    }.issubset(set(workbook.sheetnames))

    participants = pd.read_excel(result.workbook_path, sheet_name=PARTICIPANT_INCLUSION_SHEET)
    inclusion_by_subject = {
        str(row.subject_id): bool(row.included)
        for row in participants.itertuples(index=False)
    }
    assert inclusion_by_subject == {"P01": True, "P02": False, "P03": False}

    audit = json.loads(result.audit_path.read_text(encoding="utf-8"))
    assert audit["selected_conditions"] == ["CondA", "CondB"]
    assert audit["manual_excluded_subjects"] == ["P02"]
    assert audit["qc_excluded_subjects"] == ["P03"]
    assert [roi["roi"] for roi in audit["rois"]] == ["LOT", "ROT", "Central"]


def _write_fullfft_detectability_workbook(
    path: Path,
    *,
    target_scale: float,
    weak_scale: float = 0.8,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    freq_cols = [f"{idx / 10:.4f}_Hz" for idx in range(51)]
    rows = []
    for electrode, scale in [("P7", target_scale), ("PO7", weak_scale), ("Cz", weak_scale)]:
        row = {"Electrode": electrode}
        for idx, column in enumerate(freq_cols):
            row[column] = 1.0 + (idx % 7) * 0.02
        row["1.2000_Hz"] = scale
        row["2.4000_Hz"] = scale
        rows.append(row)
    fullfft = pd.DataFrame(rows)
    z_score = pd.DataFrame(
        {
            "Electrode": ["P7", "PO7", "Cz"],
            "1.2000_Hz": [3.0, 0.2, 0.1],
            "2.4000_Hz": [0.4, 0.2, 0.1],
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        fullfft.to_excel(writer, sheet_name="FullFFT Amplitude (uV)", index=False)
        z_score.to_excel(writer, sheet_name="Z Score", index=False)


def test_individual_detectability_frames_use_participant_fullfft_without_grand_average(tmp_path: Path) -> None:
    p1 = tmp_path / "P01_CondA_Results.xlsx"
    p2 = tmp_path / "P02_CondA_Results.xlsx"
    _write_fullfft_detectability_workbook(p1, target_scale=4.0)
    _write_fullfft_detectability_workbook(p2, target_scale=1.1)
    workbooks = [
        WorkbookEntry(condition="CondA", subject_id="P01", path=p1),
        WorkbookEntry(condition="CondA", subject_id="P02", path=p2),
    ]

    roi, electrode, fdr, legacy = _individual_detectability_frames(
        request=PublicationReportRequest(
            project_root=tmp_path,
            rois=(ReportRoi("LOT", ("P7", "PO7"), "primary"),),
        ),
        workbooks=workbooks,
        selected_harmonics=(1.2, 2.4),
        warnings=[],
    )

    assert set(roi["harmonic_list"]) == {"1.2, 2.4"}
    assert set(electrode["harmonic_list"]) == {"1.2, 2.4"}
    p1_z = electrode.loc[(electrode["participant_id"] == "P01") & (electrode["electrode"] == "P7"), "z_sum"].iloc[0]
    p2_z = electrode.loc[(electrode["participant_id"] == "P02") & (electrode["electrode"] == "P7"), "z_sum"].iloc[0]
    assert p1_z > p2_z
    assert not fdr.empty
    assert "Legacy_Stouffer_z" in legacy.columns


def test_individual_electrode_fdr_is_scoped_within_participant_condition(tmp_path: Path) -> None:
    p1 = tmp_path / "P01_CondA_Results.xlsx"
    p2 = tmp_path / "P02_CondB_Results.xlsx"
    _write_fullfft_detectability_workbook(p1, target_scale=4.0, weak_scale=0.7)
    _write_fullfft_detectability_workbook(p2, target_scale=4.0, weak_scale=0.7)

    _roi, _electrode, fdr, _legacy = _individual_detectability_frames(
        request=PublicationReportRequest(
            project_root=tmp_path,
            rois=(ReportRoi("LOT", ("P7", "PO7"), "primary"),),
        ),
        workbooks=[
            WorkbookEntry(condition="CondA", subject_id="P01", path=p1),
            WorkbookEntry(condition="CondB", subject_id="P02", path=p2),
        ],
        selected_harmonics=(1.2, 2.4),
        warnings=[],
    )

    cond_a_q = fdr.loc[fdr["condition"] == "CondA", "p_fdr_bh"].tolist()
    cond_b_q = fdr.loc[fdr["condition"] == "CondB", "p_fdr_bh"].tolist()
    assert cond_a_q == cond_b_q
    assert len(cond_a_q) == 3


def test_generate_publication_report_rejects_missing_selected_condition(tmp_path: Path) -> None:
    root = _build_project(tmp_path)

    with pytest.raises(RuntimeError, match="Selected condition folder"):
        generate_publication_report(
            PublicationReportRequest(
                project_root=root,
                selected_conditions=("CondA", "MissingCondition"),
            )
        )


def test_generate_publication_report_rejects_multi_group_project(tmp_path: Path) -> None:
    root = _build_project(tmp_path)
    _write_project_manifest(root, groups={"Control": {}, "Patient": {}})

    with pytest.raises(RuntimeError, match="single-group workflow"):
        generate_publication_report(PublicationReportRequest(project_root=root))


def test_z_score_report_preserves_mean_column_values() -> None:
    roi_harmonic_summary = pd.DataFrame(
        [
            {
                "condition": "CondA",
                "roi": "LOT",
                "harmonic_hz": 2.4,
                "metric": "Z",
                "mean": 3.25,
                "z_one_tailed_p": 0.0006,
                "z_gt_1.64": True,
            }
        ]
    )
    base_rate_summary = pd.DataFrame(
        [
            {
                "condition": "CondA",
                "roi": "Bilateral OT",
                "base_harmonic_hz": 6.0,
                "metric": "Z",
                "mean": 12.5,
                "z_one_tailed_p": 0.0,
                "significant_z_gt_1_64": True,
            }
        ]
    )

    report = _z_score_report(
        harmonic_selection=pd.DataFrame(),
        roi_harmonic_summary=roi_harmonic_summary,
        base_rate_summary=base_rate_summary,
    )

    assert report["z_score"].tolist() == [3.25, 12.5]


def test_roi_response_summary_exports_normality_and_wilcoxon_sensitivity() -> None:
    rows = [
        {
            "condition": "CondA",
            "subject_id": f"S{index:02d}",
            "roi": "LOT",
            "roi_role": "primary",
            "harmonic_hz": 1.2,
            "metric": "BCA_uV",
            "value": value,
        }
        for index, value in enumerate([1, 2, 3, 4, 5, 100], start=1)
    ]

    summary, response = _roi_response_summary(
        pd.DataFrame(rows),
        selected_harmonics=(1.2,),
    )

    assert len(response) == 6
    row = summary.iloc[0]
    assert row["normality_p"] < 0.05
    assert row["parametric_test"] == "one_sample_t"
    assert row["nonparametric_test"] == "wilcoxon_signed_rank"
    assert row["selected_test"] == "wilcoxon_signed_rank"
    assert row["parametric_p"] == pytest.approx(row["p_value_two_tailed"])
    assert pd.notna(row["nonparametric_p"])
    assert "Shapiro-Wilk p < .05" in row["decision_reason"]


def test_condition_pairs_export_difference_normality_and_wilcoxon() -> None:
    rows = []
    for index, difference in enumerate([1, 2, 3, 4, 5, 100], start=1):
        subject = f"S{index:02d}"
        rows.append(
            {
                "condition": "CondA",
                "subject_id": subject,
                "roi": "LOT",
                "roi_role": "primary",
                "selected_harmonics_hz": "1.2",
                "summed_bca_uv": float(difference),
            }
        )
        rows.append(
            {
                "condition": "CondB",
                "subject_id": subject,
                "roi": "LOT",
                "roi_role": "primary",
                "selected_harmonics_hz": "1.2",
                "summed_bca_uv": 0.0,
            }
        )

    pairs = _condition_pairs_by_roi(pd.DataFrame(rows))

    row = pairs.iloc[0]
    assert row["mean_difference_a_minus_b"] == pytest.approx(19.1666666667)
    assert row["normality_p"] < 0.05
    assert row["parametric_test"] == "paired_t"
    assert row["nonparametric_test"] == "wilcoxon_signed_rank"
    assert row["selected_test"] == "wilcoxon_signed_rank"
    assert pd.notna(row["nonparametric_p"])


def test_semantic_color_ratio_frames_report_raw_trimmed_and_stability() -> None:
    rows = []
    for index, semantic_value in enumerate([1, 2, 3, 4, 5, 20, 7], start=1):
        subject = f"S{index:02d}"
        color_value = 0.0 if subject == "S07" else 1.0
        for condition, value in [
            ("Semantic Response", float(semantic_value)),
            ("Color Response", color_value),
        ]:
            rows.append(
                {
                    "condition": condition,
                    "subject_id": subject,
                    "roi": "Left Occipito-Temporal",
                    "roi_role": "primary",
                    "selected_harmonics_hz": "1.2, 2.4",
                    "summed_bca_uv": value,
                }
            )

    values, summary = _semantic_color_ratio_frames(pd.DataFrame(rows))

    summary_row = summary.iloc[0]
    assert summary_row["n_participants"] == 7
    assert summary_row["n_valid_ratios"] == 6
    assert summary_row["n_invalid_denominator"] == 1
    assert summary_row["min_ratio"] == pytest.approx(1.0)
    assert summary_row["max_ratio"] == pytest.approx(20.0)
    assert summary_row["mean_ratio"] == pytest.approx(35 / 6)
    assert summary_row["median_ratio"] == pytest.approx(3.5)
    assert summary_row["trimmed_n"] == 4
    assert summary_row["trimmed_min_ratio"] == pytest.approx(2.0)
    assert summary_row["trimmed_max_ratio"] == pytest.approx(5.0)
    assert summary_row["trimmed_mean_ratio"] == pytest.approx(3.5)
    assert summary_row["trimmed_median_ratio"] == pytest.approx(3.5)
    assert summary_row["percent_within_20pct_of_median"] == pytest.approx(2 / 6)
    assert summary_row["stability_note"] == "variable_ratio_across_participants"

    invalid = values.loc[values["subject_id"] == "S07"].iloc[0]
    assert bool(invalid["ratio_valid"]) is False
    assert invalid["invalid_reason"] == "zero_color_denominator"


def test_report_rois_from_settings_pairs_marks_semantic_defaults() -> None:
    rois = report_rois_from_settings_pairs(
        [
            ("Left Occipito-Temporal", ["p7", "po7"]),
            ("Right Occipito-Temporal", ["p8", "po8"]),
            ("Central", ["cz"]),
            ("Left Parietal", ["p3"]),
        ]
    )

    by_name = {roi.name: roi for roi in rois}
    assert by_name["Left Occipito-Temporal"].selected is True
    assert by_name["Left Occipito-Temporal"].role == "primary"
    assert by_name["Right Occipito-Temporal"].selected is True
    assert by_name["Right Occipito-Temporal"].role == "primary"
    assert by_name["Central"].selected is True
    assert by_name["Left Parietal"].selected is False


def test_planned_lateralization_contrasts_use_right_minus_left_and_interaction() -> None:
    rows = []
    for subject, semantic_left, semantic_right, color_left, color_right in [
        ("S01", 1.00, 1.40, 1.00, 1.02),
        ("S02", 1.10, 1.35, 0.90, 0.88),
        ("S03", 0.95, 1.30, 1.10, 1.11),
        ("S04", 1.05, 1.45, 0.95, 0.94),
    ]:
        for condition, left, right in [
            ("Semantic Response", semantic_left, semantic_right),
            ("Color Response", color_left, color_right),
        ]:
            rows.append(
                {
                    "condition": condition,
                    "subject_id": subject,
                    "roi": "Left Occipito-Temporal",
                    "roi_role": "primary",
                    "summed_bca_uv": left,
                }
            )
            rows.append(
                {
                    "condition": condition,
                    "subject_id": subject,
                    "roi": "Right Occipito-Temporal",
                    "roi_role": "primary",
                    "summed_bca_uv": right,
                }
            )

    contrasts = _planned_lateralization_contrasts(pd.DataFrame(rows))

    assert contrasts["contrast_type"].tolist() == [
        "condition_lateralization",
        "condition_lateralization",
        "lateralization_difference",
    ]
    semantic = contrasts.loc[contrasts["condition"] == "Semantic Response"].iloc[0]
    color = contrasts.loc[contrasts["condition"] == "Color Response"].iloc[0]
    interaction = contrasts.loc[
        contrasts["contrast_type"] == "lateralization_difference"
    ].iloc[0]
    assert semantic["mean_right_minus_left_uv"] == pytest.approx(0.35)
    assert semantic["direction"] == "right_greater_than_left"
    assert color["mean_right_minus_left_uv"] == pytest.approx(0.0)
    assert color["direction"] == "no_direction"
    assert "p_bh_planned_family" not in contrasts.columns
    assert pd.notna(semantic["normality_p"])
    assert pd.notna(semantic["nonparametric_p"])
    assert semantic["selected_test"] in {"paired_t", "wilcoxon_signed_rank"}
    assert semantic["p_holm_planned_family"] >= semantic["selected_p"]
    assert interaction["mean_difference_of_lateralization_uv"] == pytest.approx(0.35)
    assert interaction["condition_a"] == "Semantic Response"
    assert interaction["condition_b"] == "Color Response"
    assert pd.notna(interaction["normality_p"])
    assert pd.notna(interaction["nonparametric_p"])
    assert interaction["p_holm_planned_family"] >= interaction["selected_p"]
