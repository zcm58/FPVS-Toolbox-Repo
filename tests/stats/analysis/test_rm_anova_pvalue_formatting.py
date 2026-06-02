from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from Tools.Stats.reporting import reporting_summary
from Tools.Stats.reporting import stats_export
from Tools.Stats.reporting import stats_export_formatting as export_formatting


def test_rm_anova_excel_preserves_small_p_values(tmp_path: Path) -> None:
    df = pd.DataFrame(
        [
            {
                "Effect": "condition",
                "Num DF": 1,
                "Den DF": 19,
                "F Value": 34.2,
                "Pr > F": 3.23e-08,
                "Pr > F (GG)": 3.23e-08,
                "epsilon (GG)": 0.77,
            }
        ]
    )
    out = tmp_path / "RM-ANOVA Results.xlsx"

    stats_export.export_rm_anova_results_to_excel(df, out, lambda _msg: None)
    export_formatting.apply_rm_anova_pvalue_number_formats(out)

    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb["RM-ANOVA Table"]
    header_map = {str(cell.value): i for i, cell in enumerate(ws[1], start=1)}

    p_cell = ws.cell(row=2, column=header_map["Pr > F"])
    gg_cell = ws.cell(row=2, column=header_map["Pr > F (GG)"])

    assert float(p_cell.value) == 3.23e-08
    assert float(gg_cell.value) == 3.23e-08
    assert float(p_cell.value) != 0.0
    assert p_cell.number_format == "0.00E+00"
    assert gg_cell.number_format == "0.00E+00"
    assert ws.auto_filter.ref == ws.dimensions
    assert ws.column_dimensions["A"].width > 10

    wb.close()


def test_fmt_p_scientific_threshold() -> None:
    assert "e-" in reporting_summary.fmt_p(3.23e-08)
    assert "e" not in reporting_summary.fmt_p(0.0014).lower()
    assert reporting_summary.fmt_p(0.0) == "0"
