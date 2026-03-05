from __future__ import annotations

from datetime import datetime
import importlib.util
from pathlib import Path
import sys

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
REPORT_PATH = ROOT / "src" / "Tools" / "Stats" / "PySide6" / "lmm_reporting.py"
FORMAT_PATH = ROOT / "src" / "Tools" / "Stats" / "PySide6" / "stats_export_formatting.py"
SUMMARY_PATH = ROOT / "src" / "Tools" / "Stats" / "PySide6" / "summary_utils.py"


def _load_module(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Failed to load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


lmm_reporting = _load_module(REPORT_PATH, "lmm_reporting_under_test")
export_formatting = _load_module(FORMAT_PATH, "lmm_export_formatting_under_test")
summary_utils = _load_module(SUMMARY_PATH, "lmm_summary_utils_under_test")


def test_humanize_effect_labels() -> None:
    assert lmm_reporting.humanize_effect_label("Intercept") == "Intercept (grand mean)"
    assert "Condition: Fruit vs Veg (Sum-coded)" == lmm_reporting.humanize_effect_label("C(condition, Sum)[S.Fruit vs Veg]")
    assert "Roi: Right Occipito Temporal (Sum-coded)" == lmm_reporting.humanize_effect_label("C(roi, Sum)[S.Right Occipito Temporal]")
    label = lmm_reporting.humanize_effect_label(
        "C(condition, Sum)[S.Green Veg vs Red Veg]:C(roi, Sum)[S.Right Occipito Temporal]"
    )
    assert "Condition×Roi:" in label


def test_wald_p_underflow_fixed() -> None:
    p_10248 = lmm_reporting.compute_two_sided_wald_p(10.248)
    p_12643 = lmm_reporting.compute_two_sided_wald_p(12.643)
    assert p_10248 > 0.0
    assert p_10248 < 1e-10
    assert p_12643 > 0.0


def test_optimizer_recorded_in_attrs() -> None:
    import numpy as np
    import statsmodels.formula.api as smf

    rng = np.random.default_rng(7)
    subject_effect = {f"S{i}": float(rng.normal(0.0, 0.4)) for i in range(12)}
    rows = []
    for subject in subject_effect:
        for condition in ("A", "B"):
            for roi in ("R1", "R2"):
                rows.append(
                    {
                        "subject": subject,
                        "condition": condition,
                        "roi": roi,
                        "value": (
                            1.0
                            + subject_effect[subject]
                            + (0.5 if condition == "B" else -0.5)
                            + (0.25 if roi == "R2" else -0.25)
                            + float(rng.normal(0.0, 0.15))
                        ),
                    }
                )
    df = pd.DataFrame(rows)
    model = smf.mixedlm(
        "value ~ C(condition, Sum) * C(roi, Sum)",
        df,
        groups=df["subject"],
        re_formula="1",
    ).fit(reml=True, method="powell", maxiter=1000, full_output=True)

    lmm_df = pd.DataFrame({"Effect": ["Intercept"], "Note": [""]})
    converged, singular, optimizer, warnings = lmm_reporting.infer_lmm_diagnostics(lmm_df, model)
    lmm_reporting.attach_lmm_run_metadata(
        table=lmm_df,
        formula="value ~ C(condition, Sum) * C(roi, Sum)",
        fixed_effects=["C(condition, Sum) * C(roi, Sum)"],
        contrast_map={"condition": "Sum", "roi": "Sum"},
        method_requested="reml",
        method_used="REML",
        re_formula_requested="1",
        re_formula_used="1",
        backed_off_random_slopes=False,
        converged=converged,
        singular=singular,
        optimizer_used=optimizer,
        fit_warnings=warnings,
        rows_input=len(df),
        rows_used=len(df),
        subjects_used=df["subject"].nunique(),
        lrt_table=None,
    )
    assert lmm_df.attrs["lmm_optimizer_used"] in {"lbfgs", "powell"}


def test_lmm_attrs_populated() -> None:
    table = pd.DataFrame(
        {
            "Effect": ["Intercept", "C(condition, Sum)[S.Fruit vs Veg]"],
            "Coef.": [1.0, 0.2],
            "SE": [0.1, 0.05],
            "Z": [10.0, 4.0],
            "P>|z|": [7.0e-24, 6.3e-05],
            "Note": ["", ""],
        }
    )
    table = lmm_reporting.ensure_lmm_effect_columns(table)
    lmm_reporting.attach_lmm_run_metadata(
        table=table,
        formula="value ~ C(condition, Sum) * C(roi, Sum)",
        fixed_effects=["C(condition, Sum) * C(roi, Sum)"],
        contrast_map={"condition": "Sum", "roi": "Sum"},
        method_requested="reml",
        method_used="REML",
        re_formula_requested="1",
        re_formula_used="1",
        backed_off_random_slopes=False,
        converged=True,
        singular=False,
        optimizer_used="lbfgs",
        fit_warnings=[],
        rows_input=24,
        rows_used=24,
        subjects_used=6,
        lrt_table=None,
    )
    assert table.attrs["lmm_formula"].startswith("value ~")
    assert table.attrs["lmm_contrast_map"]["condition"] == "Sum"
    assert table.attrs["lmm_optimizer_used"] == "lbfgs"


def test_excel_p_format_scientific_for_small_p(tmp_path: Path) -> None:
    workbook_path = tmp_path / "Mixed Model Results.xlsx"
    lmm_df = pd.DataFrame(
        {
            "Effect (readable)": ["Condition: demo"],
            "Effect (raw)": ["C(condition, Sum)[S.demo]"],
            "Coef.": [1.2],
            "SE": [0.1],
            "Z": [12.0],
            "P>|z|": [3.2e-10],
            "CI Low": [1.0],
            "CI High": [1.4],
            "Note": [""],
        }
    )
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        lmm_df.to_excel(writer, sheet_name="Mixed Model", index=False)
    export_formatting.apply_lmm_number_formats_and_metadata(workbook_path, lmm_df=lmm_df)

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["Mixed Model"]
    headers = {str(c.value): i for i, c in enumerate(ws[1], start=1)}
    p_cell = ws.cell(row=2, column=headers["P>|z|"])
    assert isinstance(p_cell.value, float)
    assert p_cell.value > 0
    assert p_cell.number_format == "0.00E+00"
    assert "Metadata" in wb.sheetnames
    wb.close()


def test_lmm_text_report_written_to_stats_results_dir(tmp_path: Path) -> None:
    stats_dir = tmp_path / "3 - Statistical Analysis Results"
    table = pd.DataFrame(
        {
            "Effect (readable)": ["Intercept (grand mean)"],
            "Effect (raw)": ["Intercept"],
            "Coef.": [0.3],
            "SE": [0.1],
            "Z": [3.0],
            "P>|z|": [0.002],
        }
    )
    table.attrs["lmm_formula"] = "value ~ C(condition, Sum) * C(roi, Sum)"
    text = lmm_reporting.build_lmm_text_report(
        lmm_df=table,
        generated_local=datetime.now().astimezone(),
        project_name="Demo",
    )
    report_path = lmm_reporting.build_lmm_report_path(stats_dir, datetime.now().astimezone())
    report_path.write_text(text, encoding="utf-8")

    assert report_path.exists()
    assert report_path.parent == stats_dir


def test_gui_summary_no_placeholder() -> None:
    lmm_df = pd.DataFrame(
        {
            "Effect (readable)": ["Intercept (grand mean)", "Condition: Demo (Sum-coded)"],
            "Coef.": [0.3, 0.2],
            "P>|z|": [1.21e-24, 0.0049],
        }
    )
    frames = summary_utils.StatsSummaryFrames(mixed_model_terms=lmm_df)
    config = summary_utils.SummaryConfig(alpha=0.05)
    summary = summary_utils.build_summary_from_frames(frames=frames, config=config)
    assert "no summary is available" not in summary.lower()
    assert "Wald z-tests (normal approximation)" in summary


