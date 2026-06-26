import importlib.util

import openpyxl
import pandas as pd
import pytest

from tests import repo_root
from Tools.Plot_Generator.spectral_qc_alerts import (
    build_spectral_qc_alert_message,
    whole_participant_exclusion_candidates,
)
from Tools.Plot_Generator.spectral_qc import interpolate_fullfft_electrode_data


def _import_module():
    path = repo_root() / "src" / "Tools" / "Plot_Generator" / "plot_generator.py"
    spec = importlib.util.spec_from_file_location("plot_generator", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _write_workbook(path, *, bad_ft7=False):
    ft7_snr_1hz = 50.0 if bad_ft7 else 1.0
    ft7_snr_12hz = 50.0 if bad_ft7 else 2.0
    ft7_fft_1hz = 100.0 if bad_ft7 else 0.5
    ft7_fft_12hz = 100.0 if bad_ft7 else 0.5
    full_snr = pd.DataFrame(
        {
            "Electrode": ["Cz", "Pz", "FT7"],
            "1.0000_Hz": [1.0, 1.0, ft7_snr_1hz],
            "1.2000_Hz": [2.0, 2.0, ft7_snr_12hz],
        }
    )
    full_fft = pd.DataFrame(
        {
            "Electrode": ["Cz", "Pz", "FT7"],
            "1.0000_Hz": [0.5, 0.5, ft7_fft_1hz],
            "1.2000_Hz": [0.5, 0.5, ft7_fft_12hz],
        }
    )
    with pd.ExcelWriter(path) as writer:
        full_snr.to_excel(writer, sheet_name="FullSNR", index=False)
        full_fft.to_excel(writer, sheet_name="FullFFT Amplitude (uV)", index=False)


def test_spectral_qc_flags_off_harmonic_electrodes_without_changing_plot_values(
    tmp_path,
    monkeypatch,
):
    module = _import_module()
    project_root = tmp_path / "Project"
    excel_root = project_root / "1 - Excel Data Files"
    cond_dir = excel_root / "Cond"
    out_dir = project_root / "2 - SNR Plots"
    cond_dir.mkdir(parents=True)
    out_dir.mkdir()

    for index in range(1, 4):
        _write_workbook(
            cond_dir / f"P0{index}_Cond_Results.xlsx",
            bad_ft7=False,
        )
    _write_workbook(
        cond_dir / "P04_Cond_Results.xlsx",
        bad_ft7=True,
    )

    captured = {}
    messages = []

    def dummy_plot(self, freqs, roi_data, group_curves=None):
        captured["freqs"] = freqs
        captured["roi_data"] = roi_data

    monkeypatch.setattr(module._Worker, "_plot", dummy_plot)
    monkeypatch.setattr(
        module._Worker,
        "_emit",
        lambda self, msg, *args: messages.append(msg) if msg else None,
    )
    monkeypatch.setattr(
        module._Worker,
        "_read_analysis_float",
        lambda self, option, fallback: 1.2 if option == "oddball_freq" else 0.0,
    )

    worker = module._Worker(
        folder=str(excel_root),
        condition="Cond",
        roi_map={"All": ["Cz", "Pz"]},
        selected_roi="All",
        title="t",
        xlabel="x",
        ylabel="y",
        x_min=1.0,
        x_max=1.2,
        y_min=0.0,
        y_max=60.0,
        out_dir=str(out_dir),
        project_root=str(project_root),
        spectral_qc_enabled=True,
    )

    worker._run()

    assert captured["freqs"] == [1.0, 1.2]
    assert captured["roi_data"]["All"] == pytest.approx([1.0, 2.0])

    report = project_root / "Quality Check" / "SNR_Spectral_QC_Cond.xlsx"
    assert report.exists()
    assert str(report) in worker.qc_report_paths
    assert any("1 electrode-frequency rows flagged" in message for message in messages)

    wb = openpyxl.load_workbook(report, data_only=True)
    summary = dict(wb["Summary"].iter_rows(min_row=2, max_col=2, values_only=True))
    assert summary["Flag behavior"] == "Report-only; SNR plot aggregation values are not changed."
    rows = list(wb["Flagged Electrodes"].iter_rows(values_only=True))
    assert rows[0][:4] == ("Condition", "PID", "Electrode", "Frequency (Hz)")
    assert rows[1][0:4] == ("Cond", "P04", "FT7", 1.0)
    assert rows[1][4] == "off_harmonic_fft_snr_outlier"
    assert len(rows) == 2


def test_spectral_qc_can_be_disabled(tmp_path, monkeypatch):
    module = _import_module()
    cond_dir = tmp_path / "Cond"
    cond_dir.mkdir()
    _write_workbook(
        cond_dir / "P01_Cond_Results.xlsx",
        bad_ft7=True,
    )

    captured = {}
    monkeypatch.setattr(
        module._Worker,
        "_plot",
        lambda self, freqs, roi_data, group_curves=None: captured.update(
            {"freqs": freqs, "roi_data": roi_data}
        ),
    )
    monkeypatch.setattr(module._Worker, "_emit", lambda *args, **kwargs: None)

    worker = module._Worker(
        folder=str(tmp_path),
        condition="Cond",
        roi_map={"All": ["Cz", "Pz"]},
        selected_roi="All",
        title="t",
        xlabel="x",
        ylabel="y",
        x_min=1.0,
        x_max=1.2,
        y_min=0.0,
        y_max=60.0,
        out_dir=str(tmp_path / "plots"),
        spectral_qc_enabled=False,
    )

    worker._run()

    assert captured["roi_data"]["All"] == [1.0, 2.0]
    assert worker.qc_report_paths == []


def test_fullfft_electrode_data_uses_interpolated_plot_grid():
    df = pd.DataFrame(
        {
            "Electrode": ["CPz", "Pz"],
            "4.0167_Hz": [5.0, 100.0],
            "4.0250_Hz": [1000.0, 2.0],
        }
    )

    fft_by_electrode = interpolate_fullfft_electrode_data(
        df,
        [4.0167, 4.0250],
        ["4.0167_Hz", "4.0250_Hz"],
        [4.02],
    )

    assert fft_by_electrode["CPZ"][0] > 300.0
    assert fft_by_electrode["PZ"][0] > 50.0


def test_spectral_qc_alert_message_recommends_reprocessing_flagged_electrodes():
    widespread_flags = [
        {
            "condition": "Erotic",
            "pid": "P12",
            "electrode": f"E{index:02d}",
            "flag_count": 10,
            "min_frequency_hz": 0.61,
            "max_frequency_hz": 16.11,
            "max_fft_amplitude_uv": 100.0 + index,
            "max_snr": 60.0 + index,
        }
        for index in range(64)
    ]
    flags = widespread_flags + [
            {
                "condition": "Neutral Angry",
                "pid": "P22",
                "electrode": "P2",
                "flag_count": 4,
                "min_frequency_hz": 16.0,
                "max_frequency_hz": 16.0,
                "max_fft_amplitude_uv": 113.99,
                "max_snr": 69.48,
            },
    ]
    message = build_spectral_qc_alert_message(
        flags,
        [r"C:\Project\Quality Check\SNR_Spectral_QC_Erotic.xlsx"],
    )
    candidates = whole_participant_exclusion_candidates(flags)

    assert "Plots and processed data were not changed" in message
    assert "Whole-participant exclusion candidate(s)" in message
    assert "P12: all 64 scalp electrodes were flagged in Erotic" in message
    assert "Recommendation: exclude these participant(s), then reprocess" in message
    assert "Localized electrode candidates: 1 participant-electrode pair" in message
    assert "SNR_Spectral_QC_Erotic.xlsx" in message
    assert candidates == [
        {
            "pid": "P12",
            "conditions": ["Erotic"],
            "max_electrode_count": 64,
            "flag_count": 640,
        }
    ]
