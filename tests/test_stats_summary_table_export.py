from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from Tools.Stats.PySide6.summary_table_export import generate_summary_table_export


def test_summary_table_export_roundtrip_and_formatting(tmp_path: Path) -> None:
    participants = [f"P{i}" for i in range(1, 6)]
    conditions = ["Words", "Faces"]
    rois = ["Left Occipito Temporal", "Central", "Parietal"]

    rows = []
    for p_idx, participant in enumerate(participants, start=1):
        for c_idx, condition in enumerate(conditions, start=1):
            rows.append({"subject": participant, "condition": condition, "roi": "Left Occipito Temporal", "value": 2.0 + p_idx + c_idx})
            rows.append({"subject": participant, "condition": condition, "roi": "Central", "value": 1.0 + p_idx + c_idx})
            rows.append({"subject": participant, "condition": condition, "roi": "Parietal", "value": 0.5 + p_idx + c_idx})
    long_df = pd.DataFrame(rows)

    baseline_rows = []
    for condition in conditions:
        for roi in rois:
            baseline_rows.append(
                {
                    "condition": condition,
                    "roi": roi,
                    "p_corr": 0.01 if roi != "Parietal" else 0.2,
                    "reject": roi != "Parietal",
                }
            )
    baseline_df = pd.DataFrame(baseline_rows)

    out_path = tmp_path / "Summary Table.xlsx"
    generate_summary_table_export(
        save_path=out_path,
        long_df=long_df,
        baseline_results_df=baseline_df,
        conditions=conditions,
        rois=rois,
        log_func=lambda _msg: None,
        project_name="Demo",
        run_identifier="Phase 1",
    )

    wb = openpyxl.load_workbook(out_path)
    try:
        assert wb.sheetnames == ["ROI Summary", "Gradients", "Metadata"]

        roi_ws = wb["ROI Summary"]
        assert roi_ws.max_row - 1 == len(conditions) * len(rois)

        grad_ws = wb["Gradients"]
        assert grad_ws.max_row - 1 == len(conditions)

        for ws in [roi_ws, grad_ws, wb["Metadata"]]:
            assert ws.freeze_panes == "A2"
            assert ws.auto_filter.ref is not None
            for cell in ws[1]:
                assert cell.font.bold is True
                assert cell.alignment.horizontal == "center"
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    assert cell.alignment.horizontal == "center"
    finally:
        wb.close()
