from __future__ import annotations

import importlib.util
from pathlib import Path
import sys

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
EXPORT_MODULE_PATH = ROOT / "src" / "Tools" / "Stats" / "Legacy" / "stats_export.py"
REPORTING_MODULE_PATH = ROOT / "src" / "Tools" / "Stats" / "PySide6" / "reporting_summary.py"
FORMAT_MODULE_PATH = ROOT / "src" / "Tools" / "Stats" / "PySide6" / "stats_export_formatting.py"


def _load_module(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Failed to load module at {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


stats_export = _load_module(EXPORT_MODULE_PATH, "stats_export_under_test")
reporting_summary = _load_module(REPORTING_MODULE_PATH, "reporting_summary_under_test")
export_formatting = _load_module(FORMAT_MODULE_PATH, "stats_export_formatting_under_test")


def test_rm_anova_excel_preserves_small_p_values(tmp_path: Path) -> None:
    df = pd.DataFrame(
        [
            {
                "Effect": "condition",
                "Num DF": 1,
                "Den DF": 19,
                "F Value": 34.2,
                "Pr > F": 3.23e-08,
                "Pr > F (GG)": 3.23e-08,
                "epsilon (GG)": 0.77,
            }
        ]
    )
    out = tmp_path / "RM-ANOVA Results.xlsx"

    stats_export.export_rm_anova_results_to_excel(df, out, lambda _msg: None)
    export_formatting.apply_rm_anova_pvalue_number_formats(out)

    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb["RM-ANOVA Table"]
    header_map = {str(cell.value): i for i, cell in enumerate(ws[1], start=1)}

    p_cell = ws.cell(row=2, column=header_map["Pr > F"])
    gg_cell = ws.cell(row=2, column=header_map["Pr > F (GG)"])

    assert float(p_cell.value) == 3.23e-08
    assert float(gg_cell.value) == 3.23e-08
    assert float(p_cell.value) != 0.0
    assert p_cell.number_format == "0.00E+00"
    assert gg_cell.number_format == "0.00E+00"

    wb.close()


def test_fmt_p_scientific_threshold() -> None:
    assert "e-" in reporting_summary.fmt_p(3.23e-08)
    assert "e" not in reporting_summary.fmt_p(0.0014).lower()
    assert reporting_summary.fmt_p(0.0) == "0"
