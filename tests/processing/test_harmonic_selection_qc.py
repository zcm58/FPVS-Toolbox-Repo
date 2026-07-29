from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

from Main_App.processing import harmonic_selection_qc
from Main_App.projects import Project
from Tools.Stats.data.group_harmonic_cache import (
    clear_cached_group_harmonic_selections,
)
from Tools.Stats.io.stats_ready_export import HARMONIC_SELECTION_COLUMNS


def test_processing_harmonic_selection_qc_writes_quality_check_workbook_and_cache(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project_root = tmp_path / "Project"
    excel_root = project_root / "1 - Excel Data Files"
    condition_root = excel_root / "Faces"
    condition_root.mkdir(parents=True)
    (project_root / "project.json").write_text(
        json.dumps(
            {
                "schema_version": "2.1.0",
                "subfolders": {"excel": "1 - Excel Data Files"},
                "event_map": {"Faces": 1},
                "preprocessing": {},
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    _write_group_policy_workbook(condition_root / "S1_Faces_Results.xlsx", scale=1)
    _write_group_policy_workbook(condition_root / "S2_Faces_Results.xlsx", scale=2)
    project = SimpleNamespace(
        project_root=project_root,
        subfolders={"excel": excel_root},
        event_map={"Faces": 1},
        preprocessing={},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "load_rois_from_settings",
        lambda: {"Posterior": ["O1", "O2"], "Central": ["FZ"]},
    )
    monkeypatch.setattr(harmonic_selection_qc, "_analysis_base_frequency_hz", lambda: 6.0)
    monkeypatch.setattr(harmonic_selection_qc, "_analysis_bca_upper_limit_hz", lambda: 8.4)

    report = harmonic_selection_qc.run_processing_harmonic_selection_qc(project)

    assert report.workbook_path == project_root / "Quality Check" / "Harmonic_Selection_Summary.xlsx"
    assert report.workbook_path.exists()
    assert report.selection_metadata["detected_significant_harmonics_hz"] == pytest.approx(
        [1.2, 3.6, 7.2]
    )
    assert report.selection_metadata["selected_harmonics_hz"] == pytest.approx(
        [1.2, 2.4, 3.6, 4.8, 7.2]
    )
    workbook = load_workbook(report.workbook_path)
    assert workbook.sheetnames == ["Selection_Summary", "Harmonic_Selection"]
    harmonic_headers = [
        cell.value for cell in next(workbook["Harmonic_Selection"].iter_rows(max_row=1))
    ]
    assert harmonic_headers == HARMONIC_SELECTION_COLUMNS
    summary_values = {
        row[0].value: row[1].value
        for row in workbook["Selection_Summary"].iter_rows(min_row=2, max_col=2)
    }
    assert summary_values["Included harmonic frequencies (Hz)"] == "1.2; 2.4; 3.6; 4.8; 7.2"

    manifest = json.loads((project_root / "project.json").read_text(encoding="utf-8"))
    entries = manifest["tools"]["stats"]["group_significant_harmonics_cache"]["entries"]
    assert len(entries) == 1

    assert clear_cached_group_harmonic_selections(project_root) == 1
    repaired_report = harmonic_selection_qc.run_processing_harmonic_selection_qc(project)
    assert repaired_report.selection_metadata["selected_harmonics_hz"] == pytest.approx(
        [1.2, 2.4, 3.6, 4.8, 7.2]
    )
    repaired_manifest = json.loads(
        (project_root / "project.json").read_text(encoding="utf-8")
    )
    repaired_entries = repaired_manifest["tools"]["stats"][
        "group_significant_harmonics_cache"
    ]["entries"]
    assert len(repaired_entries) == 1

    def _unexpected_recalculation(**_kwargs):
        raise AssertionError("Downstream loading must not recalculate significant harmonics")

    monkeypatch.setattr(
        harmonic_selection_qc,
        "build_group_significant_harmonic_selection",
        _unexpected_recalculation,
    )
    loaded = harmonic_selection_qc.load_processing_harmonic_selection(project)
    assert loaded.selected_harmonics_hz == pytest.approx([1.2, 2.4, 3.6, 4.8, 7.2])
    assert loaded.selection_cache_source == "saved_processing_metadata"


def test_processing_harmonic_selection_survives_project_event_order_reload(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project_root = tmp_path / "Project"
    excel_root = project_root / "1 - Excel Data Files"
    disk_event_map = {"Faces": 1, "Objects": 2}
    live_event_map = {"Objects": 2, "Faces": 1}
    for condition in disk_event_map:
        condition_root = excel_root / condition
        condition_root.mkdir(parents=True, exist_ok=True)
        _write_group_policy_workbook(
            condition_root / f"S1_{condition}_Results.xlsx",
            scale=1,
        )
        _write_group_policy_workbook(
            condition_root / f"S2_{condition}_Results.xlsx",
            scale=2,
        )
    (project_root / "project.json").write_text(
        json.dumps(
            {
                "schema_version": "2.1.0",
                "subfolders": {"excel": "1 - Excel Data Files"},
                "event_map": disk_event_map,
                "preprocessing": {},
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    live_project = SimpleNamespace(
        project_root=project_root,
        subfolders={"excel": excel_root},
        event_map=live_event_map,
        preprocessing={},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "load_rois_from_settings",
        lambda: {"Posterior": ["O1", "O2"], "Central": ["FZ"]},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "_analysis_base_frequency_hz",
        lambda: 6.0,
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "_analysis_bca_upper_limit_hz",
        lambda: 8.4,
    )

    harmonic_selection_qc.run_processing_harmonic_selection_qc(live_project)
    reloaded_project = Project.load(project_root)
    loaded = harmonic_selection_qc.load_processing_harmonic_selection(
        reloaded_project
    )

    assert loaded.selected_harmonics_hz == pytest.approx(
        [1.2, 2.4, 3.6, 4.8, 7.2]
    )
    assert loaded.selection_cache_source == "saved_processing_metadata"


def test_processing_harmonic_selection_qc_uses_project_summation_settings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project_root = tmp_path / "Project"
    excel_root = project_root / "1 - Excel Data Files"
    condition_root = excel_root / "Faces"
    condition_root.mkdir(parents=True)
    (project_root / "project.json").write_text(
        json.dumps(
            {
                "schema_version": "2.1.0",
                "subfolders": {"excel": "1 - Excel Data Files"},
                "event_map": {"Faces": 1},
                "preprocessing": {
                    "group_significant_summation_method": "significant_only",
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    _write_group_policy_workbook(condition_root / "S1_Faces_Results.xlsx", scale=1)
    _write_group_policy_workbook(condition_root / "S2_Faces_Results.xlsx", scale=2)
    project = SimpleNamespace(
        project_root=project_root,
        subfolders={"excel": excel_root},
        event_map={"Faces": 1},
        preprocessing={"group_significant_summation_method": "significant_only"},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "load_rois_from_settings",
        lambda: {"Posterior": ["O1", "O2"], "Central": ["FZ"]},
    )
    monkeypatch.setattr(harmonic_selection_qc, "_analysis_base_frequency_hz", lambda: 6.0)
    monkeypatch.setattr(harmonic_selection_qc, "_analysis_bca_upper_limit_hz", lambda: 8.4)

    report = harmonic_selection_qc.run_processing_harmonic_selection_qc(project)

    assert report.selection_metadata["summation_method"] == "significant_only"
    assert report.selection_metadata["detected_significant_harmonics_hz"] == pytest.approx(
        [1.2, 3.6, 7.2]
    )
    assert report.selection_metadata["selected_harmonics_hz"] == pytest.approx(
        [1.2, 3.6, 7.2]
    )


def test_processing_harmonic_selection_method_upgrade_error_is_actionable(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    inputs = SimpleNamespace(
        project_root=tmp_path,
        subjects=("S1",),
        conditions=("Faces",),
        subject_data={"S1": {"Faces": str(tmp_path / "S1_Faces_Results.xlsx")}},
        base_frequency_hz=6.0,
        max_frequency_hz=8.4,
        settings=SimpleNamespace(name=harmonic_selection_qc.GROUP_SIGNIFICANT_POLICY_NAME),
        rois={"Posterior": ["O1", "O2"]},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "_processing_harmonic_selection_inputs",
        lambda _project, log_func=None: inputs,
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "build_group_harmonic_cache_request",
        lambda **_kwargs: object(),
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "lookup_cached_group_harmonic_selection",
        lambda _request: SimpleNamespace(
            hit=None,
            reason="Harmonic-selection method version changed since saved harmonics.",
        ),
    )

    with pytest.raises(RuntimeError) as exc_info:
        harmonic_selection_qc.load_processing_harmonic_selection(object())

    message = str(exc_info.value)
    assert "Settings > Recalculate Harmonics" in message
    assert "current FPVS Toolbox version" in message
    assert "EEG/FIF reprocessing is not required" in message
    assert "method version changed" in message


def test_processing_harmonic_selection_does_not_report_success_without_saved_cache(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    inputs = SimpleNamespace(
        project_root=tmp_path,
        subjects=("S1",),
        conditions=("Faces",),
        subject_data={"S1": {"Faces": str(tmp_path / "S1_Faces_Results.xlsx")}},
        base_frequency_hz=6.0,
        max_frequency_hz=8.4,
        settings=SimpleNamespace(name=harmonic_selection_qc.GROUP_SIGNIFICANT_POLICY_NAME),
        rois={"Posterior": ["O1", "O2"]},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "_processing_harmonic_selection_inputs",
        lambda _project, log_func=None: inputs,
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "build_group_significant_harmonic_selection",
        lambda **_kwargs: SimpleNamespace(to_metadata=lambda: {"selected_harmonics_hz": [1.2]}),
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "build_group_harmonic_cache_request",
        lambda **_kwargs: object(),
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "lookup_cached_group_harmonic_selection",
        lambda _request: SimpleNamespace(
            hit=None,
            reason="No saved group-significant harmonics.",
        ),
    )

    with pytest.raises(RuntimeError) as exc_info:
        harmonic_selection_qc.run_processing_harmonic_selection_qc(object())

    message = str(exc_info.value)
    assert "calculated but could not be saved" in message
    assert "downstream tools cannot load it" in message
    assert "Settings > Recalculate Harmonics" in message


def test_processing_harmonic_selection_qc_resolves_relative_excel_subfolder(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project_root = tmp_path / "Project"
    condition_root = project_root / "1 - Excel Data Files" / "Faces"
    condition_root.mkdir(parents=True)
    (project_root / "project.json").write_text(
        json.dumps(
            {
                "schema_version": "2.1.0",
                "subfolders": {"excel": "1 - Excel Data Files"},
                "event_map": {"Faces": 1},
                "preprocessing": {},
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    _write_group_policy_workbook(condition_root / "S1_Faces_Results.xlsx", scale=1)
    _write_group_policy_workbook(condition_root / "S2_Faces_Results.xlsx", scale=2)
    project = SimpleNamespace(
        project_root=project_root,
        subfolders={"excel": "1 - Excel Data Files"},
        event_map={"Faces": 1},
        preprocessing={},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "load_rois_from_settings",
        lambda: {"Posterior": ["O1", "O2"], "Central": ["FZ"]},
    )
    monkeypatch.setattr(harmonic_selection_qc, "_analysis_base_frequency_hz", lambda: 6.0)
    monkeypatch.setattr(harmonic_selection_qc, "_analysis_bca_upper_limit_hz", lambda: 8.4)

    report = harmonic_selection_qc.run_processing_harmonic_selection_qc(project)

    assert report.workbook_path == project_root / "Quality Check" / "Harmonic_Selection_Summary.xlsx"
    assert report.workbook_path.exists()


def test_processing_harmonic_inputs_omit_excluded_participant_condition(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project_root = tmp_path / "Project"
    excel_root = project_root / "1 - Excel Data Files"
    faces_root = excel_root / "Faces"
    negative_root = excel_root / "Negative Valence"
    faces_root.mkdir(parents=True)
    negative_root.mkdir(parents=True)
    p1_faces = faces_root / "P1_Faces_Results.xlsx"
    p1_negative = negative_root / "P1_Negative Valence_Results.xlsx"
    p2_negative = negative_root / "P2_Negative Valence_Results.xlsx"
    for path in (p1_faces, p1_negative, p2_negative):
        path.write_text("fixture", encoding="utf-8")
    preprocessing = {
        "manual_excluded_participant_conditions": {
            "P1": ["Negative Valence"]
        }
    }
    (project_root / "project.json").write_text(
        json.dumps(
            {
                "subfolders": {"excel": "1 - Excel Data Files"},
                "event_map": {"Faces": 1, "Negative Valence": 2},
                "preprocessing": preprocessing,
            }
        ),
        encoding="utf-8",
    )
    project = SimpleNamespace(
        project_root=project_root,
        event_map={"Faces": 1, "Negative Valence": 2},
        preprocessing=preprocessing,
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "_filter_to_completed_subjects",
        lambda **kwargs: (kwargs["subjects"], kwargs["subject_data"]),
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "filter_frequency_domain_subjects",
        lambda _root, subjects, subject_data: (subjects, subject_data, []),
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "load_rois_from_settings",
        lambda: {"Posterior": ["O1", "O2"]},
    )

    inputs = harmonic_selection_qc._processing_harmonic_selection_inputs(project)

    assert inputs.subject_data == {
        "P1": {"Faces": str(p1_faces)},
        "P2": {"Negative Valence": str(p2_negative)},
    }
    assert str(p1_negative) not in {
        path
        for participant_data in inputs.subject_data.values()
        for path in participant_data.values()
    }


def test_processing_harmonic_selection_succeeds_after_grid_outlier_exclusion(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project_root = tmp_path / "Project"
    excel_root = project_root / "1 - Excel Data Files"
    faces_root = excel_root / "Faces"
    negative_root = excel_root / "Negative Valence"
    faces_root.mkdir(parents=True)
    negative_root.mkdir(parents=True)
    _write_group_policy_workbook(
        faces_root / "S1_Faces_Results.xlsx",
        scale=1,
    )
    _write_group_policy_workbook(
        faces_root / "S2_Faces_Results.xlsx",
        scale=2,
    )
    _write_group_policy_workbook(
        negative_root / "S3_Negative Valence_Results.xlsx",
        scale=1,
        spacing_hz=0.4,
    )
    preprocessing = {
        "manual_excluded_participant_conditions": {
            "S3": ["Negative Valence"]
        }
    }
    (project_root / "project.json").write_text(
        json.dumps(
            {
                "subfolders": {"excel": "1 - Excel Data Files"},
                "event_map": {"Faces": 1, "Negative Valence": 2},
                "preprocessing": preprocessing,
            }
        ),
        encoding="utf-8",
    )
    project = SimpleNamespace(
        project_root=project_root,
        event_map={"Faces": 1, "Negative Valence": 2},
        preprocessing=preprocessing,
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "load_rois_from_settings",
        lambda: {"Posterior": ["O1", "O2"], "Central": ["FZ"]},
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "_analysis_base_frequency_hz",
        lambda: 6.0,
    )
    monkeypatch.setattr(
        harmonic_selection_qc,
        "_analysis_bca_upper_limit_hz",
        lambda: 8.4,
    )

    report = harmonic_selection_qc.run_processing_harmonic_selection_qc(project)

    assert report.workbook_path.exists()
    assert report.selection_metadata["selected_harmonics_hz"] == pytest.approx(
        [1.2, 2.4, 3.6, 4.8, 7.2]
    )


def _write_group_policy_workbook(
    path: Path,
    *,
    scale: int,
    spacing_hz: float = 0.3,
) -> None:
    frequency_values = [
        round(spacing_hz * idx, 4)
        for idx in range(0, int(round(10.2 / spacing_hz)) + 1)
    ]
    fft_values = []
    for idx, freq in enumerate(frequency_values):
        value = 20.0 if freq in {1.2, 3.6, 7.2} else (1.2 if idx % 2 == 0 else 0.8)
        fft_values.append(value)
    full_fft = pd.DataFrame(
        {
            f"{freq:.4f}_Hz": [value, value, value]
            for freq, value in zip(frequency_values, fft_values)
        },
        index=["O1", "O2", "FZ"],
    )
    full_fft.index.name = "Electrode"
    bca = pd.DataFrame(
        {
            "1.2000_Hz": [1.0 * scale, 2.0 * scale, 0.5 * scale],
            "2.4000_Hz": [100.0, 100.0, 100.0],
            "3.6000_Hz": [0.5, 0.5, 0.1],
            "4.8000_Hz": [100.0, 100.0, 100.0],
            "6.0000_Hz": [100.0, 100.0, 100.0],
            "7.2000_Hz": [1.0, 1.0, 0.1],
        },
        index=["O1", "O2", "FZ"],
    )
    bca.index.name = "Electrode"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        bca.to_excel(writer, sheet_name="BCA (uV)")
        full_fft.to_excel(writer, sheet_name="FullFFT Amplitude (uV)")
