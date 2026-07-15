from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

pytest.importorskip("PySide6")

from Main_App.gui import preprocessing_qc_workflow as workflow  # noqa: E402


class _LabelStub:
    def __init__(self) -> None:
        self.text = ""
        self.visible = True

    def setText(self, text: str) -> None:
        self.text = text

    def setVisible(self, visible: bool) -> None:
        self.visible = visible


@pytest.mark.parametrize(
    ("step", "title", "expected"),
    (
        (
            workflow._SCAN_SIGNAL_HEALTH_STEP,
            "Scan Signal Health",
            "Step 1 of 4: Scan Signal Health",
        ),
        (
            workflow._CONFIRM_REMOVED_ELECTRODES_STEP,
            "Confirm Removed Electrodes",
            "Step 2 of 4: Confirm Removed Electrodes",
        ),
        (
            workflow._CONFIRM_PARTICIPANT_EXCLUSIONS_STEP,
            "Confirm Participant Exclusions",
            "Step 3 of 4: Confirm Participant Exclusions",
        ),
        (
            workflow._REVIEW_OTHER_FLAGS_STEP,
            "Review Other Flags",
            "Step 4 of 4: Review Other Flags",
        ),
    ),
)
def test_data_quality_step_labels_are_contiguous(
    step: int,
    title: str,
    expected: str,
) -> None:
    label = _LabelStub()
    host = SimpleNamespace(processing_step_label=label)

    workflow._begin_preflight_page(
        host,
        step=step,
        title=title,
        message="Checking data quality.",
        busy=False,
        review_visible=False,
    )

    assert label.text == expected
    assert label.visible is True


def test_data_quality_preparation_hides_step_label() -> None:
    label = _LabelStub()
    host = SimpleNamespace(processing_step_label=label)

    workflow._begin_preflight_page(
        host,
        step=None,
        title="Check Raw Files",
        message="Checking BDF headers.",
        busy=True,
        review_visible=False,
    )

    assert label.visible is False


def test_group_display_requires_canonical_membership() -> None:
    assert workflow._group_display_name(None, {}) == "Single group"
    with pytest.raises(RuntimeError, match="without a group_id"):
        workflow._group_display_name(None, {"control": "Control"})
    with pytest.raises(RuntimeError, match="unknown project group_id"):
        workflow._group_display_name("missing", {"control": "Control"})


def test_live_scan_status_includes_group_label() -> None:
    host = SimpleNamespace(_preflight_qc_group_by_file={"p01.bdf": "Control"})

    assert (
        workflow._grouped_scan_progress_text(host, "Scanning P01.bdf")
        == "Scanning Control · P01.bdf"
    )


def test_live_scan_status_keeps_file_identity_with_condition_detail() -> None:
    host = SimpleNamespace(_preflight_qc_group_by_file={"p01.bdf": "Control"})

    message = "Scanning P01.bdf · Faces 2/4 · time-domain blocks"

    assert workflow._file_name_from_progress(message) == "p01.bdf"
    assert workflow._grouped_scan_progress_text(host, message) == (
        "Scanning Control · P01.bdf · Faces 2/4 · time-domain blocks"
    )


def test_preflight_worker_passes_explicit_project_and_event_scope(
    monkeypatch,
    tmp_path: Path,
) -> None:
    captured: dict[str, object] = {}

    def _scan(raw_file_infos, settings, **kwargs):  # noqa: ANN001
        captured.update(kwargs)
        captured["raw_file_infos"] = raw_file_infos
        captured["settings"] = settings
        return workflow.PreflightQcScan(results=())

    monkeypatch.setattr(workflow, "scan_preprocessing_qc", _scan)
    worker = workflow._PreflightQcWorker(
        [],
        {"event_id_map": {"Faces": 1}},
        [],
        max_workers=12,
        project_root=tmp_path,
        event_map={"Faces": 1},
    )

    worker.run()

    assert captured["project_root"] == tmp_path
    assert captured["event_map"] == {"Faces": 1}
    assert captured["max_workers"] == 12


def _hard_candidate(raw_payload: dict[str, object]) -> workflow.PreflightQcFileResult:
    return workflow.PreflightQcFileResult(
        path=Path("p34.bdf"),
        participant_id="P34",
        load_error=None,
        raw_channel_qc=raw_payload,
        raw_spectral_qc=None,
        group_id="control",
    )


def _review_candidate(
    *,
    participant_id: str = "P13",
    raw_payload: dict[str, object] | None = None,
    spectral_payload: dict[str, object] | None = None,
) -> workflow.PreflightQcFileResult:
    return workflow.PreflightQcFileResult(
        path=Path(f"{participant_id.casefold()}.bdf"),
        participant_id=participant_id,
        load_error=None,
        raw_channel_qc=raw_payload,
        raw_spectral_qc=spectral_payload,
        group_id="patient",
    )


def test_removed_electrode_review_scope_preserves_unreviewed_participants() -> None:
    existing = {"P01": ["P9"], "P02": ["Oz"]}
    review_participants = ["P03"]

    visible = workflow._filter_removed_map_for_participants(
        existing,
        review_participants,
    )
    assert visible == {}

    updated = workflow._replace_removed_map_for_participants(
        existing,
        {"P03": ["PO8"]},
        review_participants,
    )

    assert updated == {"P01": ["P9"], "P02": ["Oz"], "P03": ["PO8"]}


def test_removed_electrode_review_rows_split_auto_and_manual_sources() -> None:
    rows = workflow._removed_review_row_values(
        ["P34"],
        {"P34": ["FT7"]},
        {"P34": ["FT7", "P9"]},
        {"P34": "Control"},
        {"P34": "Low signal / flat candidate(s): FT7"},
    )

    assert rows == [
        (
            "P34",
            "Control",
            "FT7",
            "Low signal / flat candidate(s): FT7",
            "P9",
            "FT7, P9",
        )
    ]


def test_removed_electrode_review_parser_moves_auto_field_additions_to_manual() -> None:
    records, final_confirmed, warnings = workflow._removed_review_records_from_rows(
        [("P34", "FT7, P9", "", "FT7, P9")],
        {"P34": ["FT7"]},
    )

    assert final_confirmed == {"P34": ["FT7", "P9"]}
    assert records["P34"]["accepted_auto_flagged"] == ["FT7"]
    assert records["P34"]["manual_additions"] == ["P9"]
    assert records["P34"]["manual_only_missed_by_auto"] == ["P9"]
    assert "P34: moved to Manual additions: P9" in warnings


def test_removed_electrode_review_parser_tracks_rejected_auto_flags() -> None:
    records, final_confirmed, warnings = workflow._removed_review_records_from_rows(
        [("P36", "FT7", "", "FT7")],
        {"P36": ["FT7", "P9"]},
    )

    assert final_confirmed == {"P36": ["FT7"]}
    assert records["P36"]["accepted_auto_flagged"] == ["FT7"]
    assert records["P36"]["rejected_auto_flagged"] == ["P9"]
    assert records["P36"]["manual_additions"] == []
    assert warnings == []


def test_removed_electrode_review_parser_ignores_reason_column() -> None:
    records, final_confirmed, warnings = workflow._removed_review_records_from_rows(
        [("P36", "FT7", "High-amplitude candidate(s): P9", "", "FT7")],
        {"P36": ["FT7", "P9"]},
    )

    assert final_confirmed == {"P36": ["FT7"]}
    assert records["P36"]["rejected_auto_flagged"] == ["P9"]
    assert warnings == []


def test_removed_electrode_review_reasons_include_electrode_candidate_classes() -> None:
    scan = workflow.PreflightQcScan(
        results=(
            _review_candidate(
                raw_payload={
                    "channels_to_interpolate": ["FT7"],
                    "high_amplitude_channels": ["P9"],
                    "rare_burst_channels": ["P10"],
                }
            ),
        )
    )

    assert workflow._removed_review_reason_map(scan) == {
        "P13": (
            "Low signal / flat candidate(s): FT7; "
            "High-amplitude candidate(s): P9; "
            "Rare-burst candidate(s): P10"
        )
    }


def test_remaining_review_rows_exclude_removed_electrode_candidate_classes() -> None:
    scan = workflow.PreflightQcScan(
        results=(
            _review_candidate(
                participant_id="P13",
                raw_payload={
                    "high_amplitude_channels": ["C5", "T7"],
                    "rare_burst_channels": ["P9"],
                },
            ),
            _review_candidate(
                participant_id="P17",
                raw_payload={
                    "high_amplitude_channels": ["Iz"],
                    "spatial_outlier_channels": ["AF3", "F7"],
                },
            ),
        )
    )

    assert workflow._remaining_review_rows(
        scan,
        set(),
        {"patient": "Patient"},
    ) == [
        (
            "P17",
            "Patient",
            "p17.bdf",
            "spatially inconsistent channel(s): AF3, F7",
        )
    ]


def test_hard_participant_exclusion_reason_is_condensed_with_details() -> None:
    result = _hard_candidate(
        {
            "excluded": True,
            "message": (
                "p34.bdf excluded by raw channel-health QC: participant-level raw "
                "amplitude baseline was excessively noisy."
            ),
            "triggered_rules": ["raw_amplitude_baseline_failure"],
            "raw_baseline_median_std_uv": 32193.4,
            "raw_baseline_median_p2p_99_uv": 260426.5,
            "bad_channels": ["CPz", "FT7"],
            "thresholds": {
                "baseline_exclusion_median_std_uv": 10000.0,
                "baseline_exclusion_median_p2p_99_uv": 100000.0,
            },
        }
    )

    assert workflow._hard_candidate_flag(result) == "Hard raw QC"
    assert workflow._hard_candidate_reason(result) == "Extremely noisy baseline"
    assert workflow._hard_candidate_row_values(
        [result],
        {"control": "Control"},
    ) == [
        ("P34", "Control", "Hard raw QC", "Extremely noisy baseline", "")
    ]
    assert "far outside the expected range" in workflow._hard_candidate_plain_explanation(result)

    details = workflow._hard_candidate_detail_text(result, {"control": "Control"})
    assert "Group: Control" in details
    assert "Median STD: 32193.4 uV (hard exclusion >= 10000.0 uV)" in details
    assert "Median P2P99: 260426.5 uV (hard exclusion >= 100000.0 uV)" in details
    assert "raw_amplitude_baseline_failure" in details
    assert "Original raw QC message" in details


def test_review_flags_workbook_preserves_group_membership(tmp_path: Path) -> None:
    host = SimpleNamespace(currentProject=SimpleNamespace(project_root=tmp_path))

    path = workflow._write_preflight_review_flags(
        host,
        [("P17", "Patient", "p17.bdf", "spatially inconsistent channel(s): AF3")],
    )

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(workbook["Review Flags"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows == [
        ("PID", "Group", "Source File", "Flagged Item"),
        ("P17", "Patient", "p17.bdf", "spatially inconsistent channel(s): AF3"),
    ]
