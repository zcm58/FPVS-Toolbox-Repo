from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from Main_App.processing.processing_controller import RawFileInfo
from Main_App.processing.processing_ledger import (
    classify_processing_inputs,
    record_processing_results,
)
from Main_App.processing.qc_summary_export import (
    QC_SUMMARY_FILENAME,
    QC_SUMMARY_HEADERS,
    QUALITY_CHECK_FOLDER,
    build_processing_qc_rows,
    export_processing_qc_summary,
)
from Main_App.projects.project import Project


def _project_with_raws(tmp_path: Path):
    project = Project.load(tmp_path / "project")
    raw_dir = tmp_path / "raw"
    raw_dir.mkdir()
    raw_p01 = raw_dir / "P01.bdf"
    raw_p02 = raw_dir / "P02.bdf"
    raw_p01.write_bytes(b"raw p01")
    raw_p02.write_bytes(b"raw p02")
    project.input_folder = raw_dir
    project.event_map = {"Condition A": 1}
    project.save()
    return project, [
        RawFileInfo(raw_p01.resolve(), "P01"),
        RawFileInfo(raw_p02.resolve(), "P02"),
    ]


def _settings() -> dict[str, object]:
    return {
        "high_pass": 0.1,
        "low_pass": 50.0,
        "downsample": 256,
        "epoch_start": -1.0,
        "epoch_end": 125.0,
        "base_freq": 6.0,
        "oddball_freq": 1.2,
        "bca_upper_limit": 14.4,
    }


def _write_expected_output_for_first_participant(plan) -> None:
    for output_path in plan.states[0].expected_outputs:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("ok", encoding="utf-8")


def test_processing_qc_summary_rows_and_formatting(tmp_path: Path) -> None:
    project, infos = _project_with_raws(tmp_path)
    plan = classify_processing_inputs(project, infos, _settings(), project.event_map)
    _write_expected_output_for_first_participant(plan)
    results = [
        {
            "status": "ok",
            "file": str(infos[0].path),
            "audit": {
                "n_rejected": 4,
                "raw_qc_bad_channels": ["P9"],
                "raw_qc_manual_removed_channels": ["FT7"],
                "raw_qc_low_variance_channels": ["P9"],
                "raw_qc_high_amplitude_channels": ["FT8"],
                "raw_qc_spatial_outlier_channels": ["FT7"],
                "raw_qc_warning_rules": ["possible_bad_channel_cluster"],
                "kurtosis_bad_channels": ["P1", "P3"],
                "interpolated_channels": ["FT7", "P9", "P1", "P3"],
            },
        },
        {
            "status": "excluded",
            "file": str(infos[1].path),
            "reason": "recording_not_started",
            "message": "Header-only BDF.",
            "raw_channel_qc": {
                "bad_channels": [],
                "n_bad_channels": 0,
            },
        },
    ]

    record_processing_results(
        project,
        plan,
        results,
        run_mode="Batch",
        user_choice="incremental",
        cancelled=False,
    )

    rows = build_processing_qc_rows(project, plan, results)
    assert rows == [
        {
            "PID": "P01",
            "Manually Removed Electrodes": "FT7",
            "Auto-Detected Removed Electrodes (Low SD)": "P9",
            "Flagged Removed-Electrode Candidates (High Amplitude)": "FT8",
            "Flagged Removed-Electrode Candidates (Spatial Consistency)": "FT7",
            "Kurtosis-Rejected Electrodes": "P1, P3",
            "Electrodes Interpolated": "FT7, P9, P1, P3",
            "Total Number of Electrodes removed/rejected": 4,
            "Raw QC Warnings": "possible_bad_channel_cluster",
            "Missing Conditions": "None",
            "Included in Final Set": "Included",
            "Exclusion Reason": "",
        },
        {
            "PID": "P02",
            "Manually Removed Electrodes": "None",
            "Auto-Detected Removed Electrodes (Low SD)": "None",
            "Flagged Removed-Electrode Candidates (High Amplitude)": "None",
            "Flagged Removed-Electrode Candidates (Spatial Consistency)": "None",
            "Kurtosis-Rejected Electrodes": "None",
            "Electrodes Interpolated": "None",
            "Total Number of Electrodes removed/rejected": 0,
            "Raw QC Warnings": "None",
            "Missing Conditions": "None",
            "Included in Final Set": "Excluded",
            "Exclusion Reason": "Header-only BDF.",
        },
    ]

    output = export_processing_qc_summary(project, plan, results)
    assert output == (project.project_root / QUALITY_CHECK_FOLDER / QC_SUMMARY_FILENAME)
    assert output.exists()

    workbook = load_workbook(output)
    worksheet = workbook.active
    assert [cell.value for cell in worksheet[1]] == list(QC_SUMMARY_HEADERS)
    assert worksheet.auto_filter.ref == worksheet.dimensions
    assert worksheet.freeze_panes == "A2"
    assert all(cell.font.bold for cell in worksheet[1])
    for row in worksheet.iter_rows(min_row=1, max_row=3, max_col=len(QC_SUMMARY_HEADERS)):
        for cell in row:
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
    assert worksheet.column_dimensions["C"].width >= len(
        "Auto-Detected Removed Electrodes (Low SD)"
    )


def test_processing_qc_summary_uses_ledger_for_skipped_completed_participant(tmp_path: Path) -> None:
    project, infos = _project_with_raws(tmp_path)
    plan = classify_processing_inputs(project, infos[:1], _settings(), project.event_map)
    _write_expected_output_for_first_participant(plan)
    record_processing_results(
        project,
        plan,
        [
            {
                "status": "ok",
                "file": str(infos[0].path),
                "audit": {
                    "n_rejected": 3,
                    "raw_qc_bad_channels": ["P9"],
                    "raw_qc_manual_removed_channels": ["FT7"],
                    "kurtosis_bad_channels": ["Oz"],
                    "interpolated_channels": ["FT7", "P9", "Oz"],
                },
            }
        ],
        run_mode="Batch",
        user_choice="incremental",
        cancelled=False,
    )

    # Simulate a later incremental run where P01 was already completed and skipped.
    rows = build_processing_qc_rows(project, plan, [])
    assert rows[0]["Included in Final Set"] == "Included"
    assert rows[0]["Exclusion Reason"] == ""
    assert rows[0]["Missing Conditions"] == "None"
    assert rows[0]["Manually Removed Electrodes"] == "FT7"
    assert rows[0]["Auto-Detected Removed Electrodes (Low SD)"] == "P9"
    assert rows[0]["Flagged Removed-Electrode Candidates (High Amplitude)"] == "None"
    assert rows[0]["Flagged Removed-Electrode Candidates (Spatial Consistency)"] == "None"
    assert rows[0]["Kurtosis-Rejected Electrodes"] == "Oz"
    assert rows[0]["Electrodes Interpolated"] == "FT7, P9, Oz"
    assert rows[0]["Total Number of Electrodes removed/rejected"] == 3

    ledger = json.loads(
        (project.project_root / ".fpvs_processing" / "processing_ledger.json").read_text(
            encoding="utf-8"
        )
    )
    assert ledger["entries"]["P01"]["raw_qc_bad_channels"] == ["P9"]
    assert ledger["entries"]["P01"]["raw_qc_manual_removed_channels"] == ["FT7"]
    assert ledger["entries"]["P01"]["interpolated_channels"] == ["FT7", "P9", "Oz"]


def test_processing_qc_summary_flags_partial_condition_participant(tmp_path: Path) -> None:
    project, infos = _project_with_raws(tmp_path)
    project.event_map = {"Condition A": 1, "Condition B": 2}
    project.save()
    plan = classify_processing_inputs(project, infos[:1], _settings(), project.event_map)
    present_output = plan.states[0].expected_outputs[0]
    present_output.parent.mkdir(parents=True, exist_ok=True)
    present_output.write_text("ok", encoding="utf-8")
    results = [
        {
            "status": "ok",
            "file": str(infos[0].path),
            "audit": {
                "n_rejected": 3,
                "raw_qc_bad_channels": ["P9"],
                "raw_qc_manual_removed_channels": ["FT7"],
                "kurtosis_bad_channels": ["P8"],
                "interpolated_channels": ["FT7", "P9", "P8"],
            },
        }
    ]

    record_processing_results(
        project,
        plan,
        results,
        run_mode="Batch",
        user_choice="incremental",
        cancelled=False,
    )

    rows = build_processing_qc_rows(project, plan, results)

    assert rows[0]["Missing Conditions"] == "Condition B"
    assert rows[0]["Included in Final Set"] == "Included (partial conditions)"
    assert rows[0]["Exclusion Reason"] == ""
    assert rows[0]["Manually Removed Electrodes"] == "FT7"
    assert rows[0]["Auto-Detected Removed Electrodes (Low SD)"] == "P9"
    assert rows[0]["Flagged Removed-Electrode Candidates (High Amplitude)"] == "None"
    assert rows[0]["Flagged Removed-Electrode Candidates (Spatial Consistency)"] == "None"
    assert rows[0]["Kurtosis-Rejected Electrodes"] == "P8"


def test_processing_qc_summary_uses_matching_cache_for_legacy_failed_entry(tmp_path: Path) -> None:
    project, infos = _project_with_raws(tmp_path)
    plan = classify_processing_inputs(project, infos[:1], _settings(), project.event_map)
    raw_stat = infos[0].path.stat()
    ledger_path = project.project_root / ".fpvs_processing" / "processing_ledger.json"
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    ledger_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "entries": {
                    "P01": {
                        "participant_id": "P01",
                        "raw_file": str(infos[0].path),
                        "raw_size": raw_stat.st_size,
                        "raw_mtime_ns": raw_stat.st_mtime_ns,
                        "status": "failed",
                        "raw_qc_bad_channels": [],
                        "kurtosis_bad_channels": [],
                        "interpolated_channels": [],
                        "n_rejected": 0,
                    }
                },
            }
        ),
        encoding="utf-8",
    )
    cache_dir = project.project_root / ".fpvs_cache" / "preprocessed"
    cache_dir.mkdir(parents=True, exist_ok=True)
    (cache_dir / "P01_fake.json").write_text(
        json.dumps(
            {
                "payload": {
                    "source_path": str(infos[0].path),
                    "source_size": raw_stat.st_size,
                    "source_mtime_ns": raw_stat.st_mtime_ns,
                },
                    "raw_qc_bad_channels": ["P9"],
                    "raw_qc_manual_removed_channels": ["FT7"],
                    "kurtosis_bad_channels": ["P8"],
                    "interpolated_channels": ["FT7", "P9", "P8"],
                    "n_rejected": 3,
            }
        ),
        encoding="utf-8",
    )

    rows = build_processing_qc_rows(project, plan, [])

    assert rows[0]["Included in Final Set"] == "Excluded"
    assert rows[0]["Exclusion Reason"] == ""
    assert rows[0]["Missing Conditions"] == "None"
    assert rows[0]["Manually Removed Electrodes"] == "FT7"
    assert rows[0]["Auto-Detected Removed Electrodes (Low SD)"] == "P9"
    assert rows[0]["Flagged Removed-Electrode Candidates (High Amplitude)"] == "None"
    assert rows[0]["Flagged Removed-Electrode Candidates (Spatial Consistency)"] == "None"
    assert rows[0]["Kurtosis-Rejected Electrodes"] == "P8"
    assert rows[0]["Electrodes Interpolated"] == "FT7, P9, P8"
    assert rows[0]["Total Number of Electrodes removed/rejected"] == 3


def test_processing_qc_summary_treats_legacy_missing_condition_as_included(
    tmp_path: Path,
) -> None:
    project, infos = _project_with_raws(tmp_path)
    project.event_map = {"Condition A": 1, "Condition B": 2}
    project.save()
    plan = classify_processing_inputs(project, infos[:1], _settings(), project.event_map)
    present_output = plan.states[0].expected_outputs[0]
    present_output.parent.mkdir(parents=True, exist_ok=True)
    present_output.write_text("ok", encoding="utf-8")
    raw_stat = infos[0].path.stat()
    ledger_path = project.project_root / ".fpvs_processing" / "processing_ledger.json"
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    ledger_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "entries": {
                    "P01": {
                        "participant_id": "P01",
                        "raw_file": str(infos[0].path),
                        "raw_size": raw_stat.st_size,
                        "raw_mtime_ns": raw_stat.st_mtime_ns,
                        "status": "failed",
                        "raw_qc_bad_channels": ["P9"],
                        "kurtosis_bad_channels": ["P8"],
                        "interpolated_channels": ["P9", "P8"],
                        "n_rejected": 2,
                    }
                },
            }
        ),
        encoding="utf-8",
    )

    rows = build_processing_qc_rows(project, plan, [])

    assert rows[0]["Missing Conditions"] == "Condition B"
    assert rows[0]["Included in Final Set"] == "Included (partial conditions)"
