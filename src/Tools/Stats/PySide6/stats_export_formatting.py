"""Formatting helpers for Stats Excel exports.

This module applies output-only formatting adjustments after legacy exports.
"""

from __future__ import annotations

import logging
import math
from pathlib import Path

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

RM_ANOVA_P_COLUMNS: tuple[str, ...] = ("Pr > F", "Pr > F (GG)", "Pr > F (HF)")
SCIENTIFIC_P_THRESHOLD = 0.001
SCIENTIFIC_FMT = "0.00E+00"
DEFAULT_P_FMT = "0.0000"


def log_rm_anova_p_minima(anova_df: pd.DataFrame) -> None:
    """Emit diagnostic stats confirming whether tiny p-values are real zeros or display artifacts."""

    def _series_metrics(column: str) -> tuple[float | None, bool]:
        if column not in anova_df.columns:
            return None, False
        numeric = pd.to_numeric(anova_df[column], errors="coerce")
        exact_zero = bool((numeric == 0.0).fillna(False).any())
        positives = numeric[(numeric > 0.0) & numeric.notna()]
        if positives.empty:
            return None, exact_zero
        return float(positives.min()), exact_zero

    min_p_unc, any_exact_zero_unc = _series_metrics("Pr > F")
    min_p_gg, any_exact_zero_gg = _series_metrics("Pr > F (GG)")
    logger.info(
        "rm_anova_p_min",
        extra={
            "min_p_unc": min_p_unc,
            "min_p_gg": min_p_gg,
            "any_exact_zero_unc": any_exact_zero_unc,
            "any_exact_zero_gg": any_exact_zero_gg,
        },
    )


def apply_rm_anova_pvalue_number_formats(
    workbook_path: str | Path,
    *,
    sheet_name: str = "RM-ANOVA Table",
) -> None:
    """Apply scientific formatting for tiny RM-ANOVA p-values in Excel output."""

    path = Path(workbook_path)
    workbook = openpyxl.load_workbook(path)
    try:
        if sheet_name not in workbook.sheetnames:
            logger.info(
                "rm_anova_sheet_not_found_for_formatting",
                extra={"sheet_name": sheet_name, "path": str(path)},
            )
            return
        worksheet = workbook[sheet_name]
        headers = {
            str(cell.value).strip(): idx
            for idx, cell in enumerate(worksheet[1], start=1)
            if isinstance(cell.value, str)
        }
        target_indices = [headers[col] for col in RM_ANOVA_P_COLUMNS if col in headers]
        if not target_indices:
            return
        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in target_indices:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                if isinstance(value, bool) or value is None:
                    continue
                try:
                    numeric = float(value)
                except Exception:
                    continue
                if not math.isfinite(numeric):
                    continue
                if numeric != 0.0 and abs(numeric) < SCIENTIFIC_P_THRESHOLD:
                    cell.number_format = SCIENTIFIC_FMT
                else:
                    cell.number_format = DEFAULT_P_FMT
    finally:
        workbook.save(path)
        workbook.close()



LMM_P_COLUMNS: tuple[str, ...] = ("P>|z|", "P>|t|", "p (chi2)")


def apply_lmm_number_formats_and_metadata(
    workbook_path: str | Path,
    *,
    sheet_name: str = "Mixed Model",
    lmm_df: pd.DataFrame | None = None,
) -> None:
    """Apply p-value display rules and add optional LRT/metadata sheets for LMM exports."""

    path = Path(workbook_path)
    workbook = openpyxl.load_workbook(path)
    try:
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            headers = {
                str(cell.value).strip(): idx
                for idx, cell in enumerate(worksheet[1], start=1)
                if isinstance(cell.value, str)
            }
            for row_idx in range(2, worksheet.max_row + 1):
                for col_name in ("P>|z|", "P>|t|"):
                    if col_name not in headers:
                        continue
                    cell = worksheet.cell(row=row_idx, column=headers[col_name])
                    value = cell.value
                    if isinstance(value, bool) or value is None:
                        continue
                    try:
                        numeric = float(value)
                    except Exception:
                        continue
                    if not math.isfinite(numeric):
                        continue
                    cell.number_format = SCIENTIFIC_FMT if (numeric != 0.0 and abs(numeric) < SCIENTIFIC_P_THRESHOLD) else DEFAULT_P_FMT

        attrs = lmm_df.attrs if isinstance(lmm_df, pd.DataFrame) else {}
        lrt_table = attrs.get("lrt_table") if isinstance(attrs.get("lrt_table"), pd.DataFrame) else None
        if isinstance(lrt_table, pd.DataFrame):
            if "LRT" in workbook.sheetnames:
                del workbook["LRT"]
            ws_lrt = workbook.create_sheet("LRT")
            ws_lrt.append(list(lrt_table.columns))
            for row in lrt_table.itertuples(index=False):
                ws_lrt.append(list(row))
            for row in ws_lrt.iter_rows(min_row=2, max_row=ws_lrt.max_row):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws_lrt[1], start=1) if isinstance(cell.value, str)}
            if "p (chi2)" in headers:
                col_idx = headers["p (chi2)"]
                for row_idx in range(2, ws_lrt.max_row + 1):
                    cell = ws_lrt.cell(row=row_idx, column=col_idx)
                    try:
                        numeric = float(cell.value)
                    except Exception:
                        continue
                    if math.isfinite(numeric) and numeric != 0.0 and abs(numeric) < SCIENTIFIC_P_THRESHOLD:
                        cell.number_format = SCIENTIFIC_FMT
                    else:
                        cell.number_format = DEFAULT_P_FMT
            for col in ws_lrt.columns:
                width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
                ws_lrt.column_dimensions[col[0].column_letter].width = min(width, 80)

        if "Metadata" in workbook.sheetnames:
            del workbook["Metadata"]
        ws_meta = workbook.create_sheet("Metadata")
        ws_meta.append(["Field", "Value"])
        metadata_rows = [
            ("Formula", attrs.get("lmm_formula", "")),
            ("Processed terms", ", ".join(attrs.get("lmm_processed_terms", [])) if isinstance(attrs.get("lmm_processed_terms"), list) else ""),
            ("Contrast map", "; ".join(f"{k}: {v}" for k, v in (attrs.get("lmm_contrast_map", {}) or {}).items())),
            ("Method used", attrs.get("lmm_method_used", "")),
            ("Optimizer", attrs.get("lmm_optimizer_used", "")),
            ("Converged", attrs.get("lmm_converged", "")),
            ("Singular", attrs.get("lmm_singular", "")),
            ("re_formula requested", attrs.get("lmm_re_formula_requested", "")),
            ("re_formula used", attrs.get("lmm_re_formula_used", "")),
            ("Backed off random slopes", attrs.get("lmm_backed_off_random_slopes", "")),
            ("Rows input", attrs.get("lmm_rows_input", "")),
            ("Rows used", attrs.get("lmm_rows_used", "")),
            ("Rows dropped", attrs.get("lmm_rows_dropped", "")),
            ("Subjects used", attrs.get("lmm_subjects_used", "")),
            ("Warnings", "; ".join(attrs.get("lmm_fit_warnings", [])) if isinstance(attrs.get("lmm_fit_warnings"), list) else ""),
            ("Inference", "Wald z-tests (normal approximation)"),
        ]
        for key, value in metadata_rows:
            ws_meta.append([key, value])
        for col in ws_meta.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
            ws_meta.column_dimensions[col[0].column_letter].width = min(width, 120)
    finally:
        workbook.save(path)
        workbook.close()
