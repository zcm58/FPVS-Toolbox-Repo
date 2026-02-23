from __future__ import annotations

from datetime import datetime
from pathlib import Path
import random

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from scipy.stats import ttest_1samp
from statsmodels.stats.multitest import multipletests

from Tools.Stats.Legacy.stats_export import _auto_format_and_write_excel

P_COL_HEADER = "Baseline vs 0 p (BH-FDR)"
REJECT_COL_HEADER = "Baseline vs 0 reject"


def _identify_lot_roi(roi_labels: list[str]) -> str:
    normalized = {roi: roi.strip().lower() for roi in roi_labels}
    candidates = [
        roi
        for roi, norm in normalized.items()
        if norm == "lot"
        or "left occipito temporal" in norm
        or "left-occipito-temporal" in norm
        or "left occipito-temporal" in norm
        or "occipito temporal" in norm
    ]
    unique = sorted(set(candidates))
    if len(unique) != 1:
        raise ValueError("Cannot compute LOT−Central gradient: LOT ROI not found/unambiguous.")
    return unique[0]


def _identify_central_roi(roi_labels: list[str]) -> str:
    candidates = [roi for roi in roi_labels if roi.strip().lower() == "central"]
    if len(candidates) != 1:
        raise ValueError("Cannot compute LOT−Central gradient: Central ROI not found/unambiguous.")
    return candidates[0]


def _enforce_sheet_formatting(workbook_path: Path, *, decimal_cols: dict[str, list[str]], p_cols: dict[str, list[str]]) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    try:
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            header_font = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = center
            headers = {str(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}
            for col_name in decimal_cols.get(ws.title, []):
                if col_name in headers:
                    col_idx = headers[col_name]
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).number_format = "0.000000"
            for col_name in p_cols.get(ws.title, []):
                if col_name in headers:
                    col_idx = headers[col_name]
                    for row_idx in range(2, ws.max_row + 1):
                        val = ws.cell(row=row_idx, column=col_idx).value
                        if isinstance(val, (float, int)):
                            ws.cell(row=row_idx, column=col_idx).number_format = "0.00E+00" if (val != 0 and abs(float(val)) < 0.001) else "0.000000"

            for col_cells in ws.columns:
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 80)
    finally:
        wb.save(workbook_path)
        wb.close()


def generate_summary_table_export(
    *,
    save_path: str | Path,
    long_df: pd.DataFrame,
    baseline_results_df: pd.DataFrame,
    conditions: list[str],
    rois: list[str],
    log_func,
    project_name: str | None = None,
    run_identifier: str | None = None,
    is_between_group: bool = False,
) -> Path:
    if is_between_group:
        raise NotImplementedError("Summary Table export currently supports single-group runs only.")

    required_cols = ["subject", "condition", "roi", "value"]
    missing = [col for col in required_cols if col not in long_df.columns]
    if missing:
        raise ValueError(f"Summary Table export missing required long-format columns: {missing}")

    lot_roi = _identify_lot_roi(rois)
    central_roi = _identify_central_roi(rois)

    baseline_required = ["condition", "roi", "p_corr", "reject"]
    baseline_missing = [col for col in baseline_required if col not in baseline_results_df.columns]
    if baseline_missing:
        raise ValueError(f"Summary Table export missing baseline-vs-zero columns: {baseline_missing}")

    log_func("Generating Summary Table.xlsx…")
    summary_rows = []
    baseline_lookup: dict[tuple[str, str], tuple[float, bool]] = {}
    for row in baseline_results_df.itertuples(index=False):
        baseline_lookup[(str(row.condition), str(row.roi))] = (float(row.p_corr), bool(row.reject))

    for condition in conditions:
        for roi in rois:
            mask = (long_df["condition"].astype(str) == str(condition)) & (long_df["roi"].astype(str) == str(roi))
            values = long_df.loc[mask, "value"].astype(float)
            key = (str(condition), str(roi))
            if key not in baseline_lookup:
                raise ValueError(f"Missing baseline-vs-zero BH-FDR result for Condition={condition}, ROI={roi}.")
            p_corr, reject = baseline_lookup[key]
            summary_rows.append(
                {
                    "Condition": condition,
                    "ROI": roi,
                    "N": int(values.shape[0]),
                    "Mean (Summed BCA)": float(values.mean()),
                    "SD": float(values.std(ddof=1)),
                    P_COL_HEADER: float(p_corr),
                    REJECT_COL_HEADER: bool(reject),
                }
            )

    roi_summary_df = pd.DataFrame(summary_rows)

    gradient_rows = []
    gradient_pvals = []
    paired_cache: dict[str, pd.Series] = {}
    for condition in conditions:
        cond_df = long_df.loc[long_df["condition"].astype(str) == str(condition), ["subject", "roi", "value"]].copy()
        lot_series = cond_df.loc[cond_df["roi"].astype(str) == lot_roi, ["subject", "value"]].set_index("subject")["value"].astype(float)
        central_series = cond_df.loc[cond_df["roi"].astype(str) == central_roi, ["subject", "value"]].set_index("subject")["value"].astype(float)
        common = sorted(set(lot_series.index).intersection(set(central_series.index)))
        gradients = (lot_series.loc[common] - central_series.loc[common]).astype(float)
        sd = float(gradients.std(ddof=1))
        if sd == 0:
            raise ValueError(f"Cannot compute dz for condition '{condition}': gradient SD is zero.")
        p_raw = float(ttest_1samp(gradients.to_numpy(), popmean=0.0, alternative="two-sided").pvalue)
        gradient_pvals.append(p_raw)
        paired_cache[str(condition)] = gradients
        gradient_rows.append(
            {
                "Condition": condition,
                "Gradient definition": "LOT − Central",
                "N": int(gradients.shape[0]),
                "Mean gradient": float(gradients.mean()),
                "SD gradient": sd,
                "_p_raw": p_raw,
                "_dz": float(gradients.mean() / sd),
            }
        )

    reject, p_corr, _, _ = multipletests(np.asarray(gradient_pvals, dtype=float), alpha=0.05, method="fdr_bh")
    for idx, row in enumerate(gradient_rows):
        row["LOT vs Central p (BH-FDR)"] = float(p_corr[idx])
        row["LOT vs Central reject"] = bool(reject[idx])
        row["Effect size (dz)"] = row.pop("_dz")
        row.pop("_p_raw")

    gradients_df = pd.DataFrame(gradient_rows)

    metadata_rows = [
        {"Field": "DV column name", "Value": "value"},
        {"Field": "DV description", "Value": "Summed BCA"},
        {"Field": "Included conditions", "Value": ", ".join(str(c) for c in conditions)},
        {"Field": "Included ROIs", "Value": ", ".join(str(r) for r in rois)},
        {"Field": "Baseline vs zero correction", "Value": "Benjamini–Hochberg (BH-FDR)"},
        {"Field": "Gradient correction", "Value": "Benjamini–Hochberg (BH-FDR) across conditions"},
        {"Field": "Timestamp", "Value": datetime.now().isoformat(timespec="seconds")},
    ]
    if project_name:
        metadata_rows.append({"Field": "Project name", "Value": project_name})
    if run_identifier:
        metadata_rows.append({"Field": "Run identifier", "Value": run_identifier})
    metadata_df = pd.DataFrame(metadata_rows)

    save_path = Path(save_path)
    with pd.ExcelWriter(save_path, engine="xlsxwriter") as writer:
        _auto_format_and_write_excel(writer, roi_summary_df, "ROI Summary", log_func)
        _auto_format_and_write_excel(writer, gradients_df, "Gradients", log_func)
        _auto_format_and_write_excel(writer, metadata_df, "Metadata", log_func)

    _enforce_sheet_formatting(
        save_path,
        decimal_cols={
            "ROI Summary": ["Mean (Summed BCA)", "SD"],
            "Gradients": ["Mean gradient", "SD gradient", "Effect size (dz)"],
        },
        p_cols={
            "ROI Summary": [P_COL_HEADER],
            "Gradients": ["LOT vs Central p (BH-FDR)"],
        },
    )

    _validate_summary_table(
        save_path=save_path,
        roi_summary_df=roi_summary_df,
        gradients_df=gradients_df,
        long_df=long_df,
        baseline_lookup=baseline_lookup,
        conditions=conditions,
        rois=rois,
        lot_roi=lot_roi,
        central_roi=central_roi,
    )

    log_func(f"ROI Summary rows: {len(roi_summary_df)}")
    log_func(f"Gradient rows: {len(gradients_df)}")
    log_func(f"Saved Summary Table.xlsx: {save_path}")
    return save_path


def _validate_summary_table(*, save_path: Path, roi_summary_df: pd.DataFrame, gradients_df: pd.DataFrame, long_df: pd.DataFrame, baseline_lookup: dict[tuple[str, str], tuple[float, bool]], conditions: list[str], rois: list[str], lot_roi: str, central_roi: str) -> None:
    wb = openpyxl.load_workbook(save_path)
    try:
        expected = {"ROI Summary", "Gradients", "Metadata"}
        if set(wb.sheetnames) != expected:
            raise RuntimeError(f"Summary Table workbook sheets mismatch: found={wb.sheetnames}")

        ws = wb["ROI Summary"]
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        if len(rows) != len(conditions) * len(rois):
            raise RuntimeError("ROI Summary row count mismatch.")
        row_conditions = {str(r[0]) for r in rows}
        row_rois = {str(r[1]) for r in rows}
        if row_conditions != set(map(str, conditions)):
            raise RuntimeError("ROI Summary conditions mismatch.")
        if row_rois != set(map(str, rois)):
            raise RuntimeError("ROI Summary ROIs mismatch.")

        rng = random.Random(17)
        sample = rng.sample([(c, r) for c in conditions for r in rois], k=min(3, len(conditions) * len(rois)))
        workbook_lookup = {(str(r[0]), str(r[1])): r for r in rows}
        for condition, roi in sample:
            row = workbook_lookup[(str(condition), str(roi))]
            vals = long_df.loc[(long_df["condition"].astype(str) == str(condition)) & (long_df["roi"].astype(str) == str(roi)), "value"].astype(float)
            if abs(float(row[3]) - float(vals.mean())) > 1e-9:
                raise RuntimeError("ROI Summary mean validation failed.")
            if abs(float(row[4]) - float(vals.std(ddof=1))) > 1e-9:
                raise RuntimeError("ROI Summary SD validation failed.")
            if int(row[2]) != int(vals.shape[0]):
                raise RuntimeError("ROI Summary N validation failed.")
            p_corr, reject = baseline_lookup[(str(condition), str(roi))]
            if abs(float(row[5]) - float(p_corr)) > 1e-12:
                raise RuntimeError("Baseline-vs-zero corrected p validation failed.")
            if bool(row[6]) != bool(reject):
                raise RuntimeError("Baseline-vs-zero reject validation failed.")

        grad_ws = wb["Gradients"]
        grad_rows = list(grad_ws.iter_rows(min_row=2, max_row=grad_ws.max_row, values_only=True))
        grad_lookup = {str(r[0]): r for r in grad_rows}
        for condition in conditions:
            row = grad_lookup[str(condition)]
            cond_df = long_df.loc[long_df["condition"].astype(str) == str(condition), ["subject", "roi", "value"]].copy()
            lot = cond_df.loc[cond_df["roi"].astype(str) == lot_roi, ["subject", "value"]].set_index("subject")["value"].astype(float)
            central = cond_df.loc[cond_df["roi"].astype(str) == central_roi, ["subject", "value"]].set_index("subject")["value"].astype(float)
            common = sorted(set(lot.index).intersection(set(central.index)))
            gradients = lot.loc[common] - central.loc[common]
            mean = float(gradients.mean())
            sd = float(gradients.std(ddof=1))
            dz = mean / sd
            if abs(float(row[3]) - mean) > 1e-9:
                raise RuntimeError("Gradient mean validation failed.")
            if abs(float(row[4]) - sd) > 1e-9:
                raise RuntimeError("Gradient SD validation failed.")
            if abs(float(row[7]) - dz) > 1e-9:
                raise RuntimeError("Gradient dz validation failed.")

        for sheet_name in expected:
            ws = wb[sheet_name]
            if ws.freeze_panes != "A2":
                raise RuntimeError(f"Formatting validation failed: freeze panes not set on {sheet_name}.")
            if ws.auto_filter.ref is None:
                raise RuntimeError(f"Formatting validation failed: autofilter missing on {sheet_name}.")
            for cell in ws[1]:
                if not bool(cell.font.bold):
                    raise RuntimeError(f"Formatting validation failed: non-bold header on {sheet_name}.")
                if cell.alignment.horizontal != "center":
                    raise RuntimeError(f"Formatting validation failed: header alignment on {sheet_name}.")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.alignment.horizontal != "center":
                        raise RuntimeError(f"Formatting validation failed: data alignment on {sheet_name}.")
    finally:
        wb.close()
