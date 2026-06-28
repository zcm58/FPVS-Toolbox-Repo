"""Stats-ready Summed BCA workbook export helpers."""

from __future__ import annotations

import logging
import math
import re
from collections import Counter
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from pathlib import Path
from time import perf_counter

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from Tools.Stats.analysis.dv_policies import prepare_summed_bca_data
from Tools.Stats.analysis.dv_policy_settings import (
    FIXED_PREDEFINED_POLICY_ID,
    GROUP_SIGNIFICANT_POLICY_ID,
)

logger = logging.getLogger("Tools.Stats")
STATS_READY_WORKBOOK_NAME = "Stats_Ready_Summed_BCA.xlsx"
SINGLE_GROUP_LABEL = "single_group"

LONG_FORMAT_SHEET = "Long_Format"
WIDE_FORMAT_SHEET = "Wide_Format"
SELECTION_SUMMARY_SHEET = "Selection_Summary"
SELECTION_SUMMARY_COLUMNS = [
    "Summary Item",
    "Value",
]
HARMONIC_SELECTION_SHEET = "Harmonic_Selection"
HARMONIC_SELECTION_COLUMNS = [
    "requested_harmonic_hz",
    "z_score",
    "target_amplitude_uv",
    "noise_mean_uv",
    "noise_std_uv",
    "selected",
    "included_in_summation",
    "excluded_base_rate",
    "exclusion_reason",
    "warning",
]


@dataclass(frozen=True)
class StatsReadyExport:
    """Prepared workbook frames and lightweight export metadata."""

    frames: dict[str, pd.DataFrame]
    workbook_path: Path | None
    row_count: int


def _safe_text(value: object, *, default: str = "") -> str:
    text = str(value).strip() if value is not None else ""
    return text or default


def _numeric_or_nan(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    return number if math.isfinite(number) else math.nan


def _parse_frequency(value: object) -> float | None:
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip()
    if text.endswith("_Hz"):
        text = text[:-3]
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _sequence_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, Mapping):
        return ";".join(f"{key}:{item}" for key, item in value.items())
    try:
        return ";".join(str(item) for item in value)
    except TypeError:
        return str(value)


def _format_number(value: object) -> object:
    number = _numeric_or_nan(value)
    if math.isfinite(number):
        rounded = float(f"{number:.12g}")
        return int(rounded) if rounded.is_integer() else rounded
    return value


def _format_sequence_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        return "; ".join(str(_format_number(item)) for item in value)
    except TypeError:
        return str(_format_number(value))


def _yes_no(value: bool | None) -> str:
    if value is None:
        return ""
    return "Yes" if value else "No"


def _common_rossion_meta(dv_metadata: Mapping[str, object]) -> Mapping[str, object]:
    rossion_meta = dv_metadata.get("rossion_method")
    return rossion_meta if isinstance(rossion_meta, Mapping) else {}


def _common_fixed_predefined_meta(dv_metadata: Mapping[str, object]) -> Mapping[str, object]:
    fixed_meta = dv_metadata.get("fixed_predefined_harmonics")
    return fixed_meta if isinstance(fixed_meta, Mapping) else {}


def _common_group_significant_meta(dv_metadata: Mapping[str, object]) -> Mapping[str, object]:
    group_meta = dv_metadata.get("group_significant_harmonics")
    return group_meta if isinstance(group_meta, Mapping) else {}


def _resolve_group_labels(
    subjects: list[str],
    group_map: Mapping[str, object] | None,
) -> dict[str, str]:
    if not group_map:
        return {subject: SINGLE_GROUP_LABEL for subject in subjects}

    labels: dict[str, str | None] = {}
    has_known_group = False
    for subject in subjects:
        group_value = group_map.get(subject)
        if group_value is None:
            group_value = group_map.get(subject.upper())
        group = _safe_text(group_value)
        if group:
            has_known_group = True
            labels[subject] = group
        else:
            labels[subject] = None

    if not has_known_group:
        return {subject: SINGLE_GROUP_LABEL for subject in subjects}

    missing = [subject for subject, group in labels.items() if not group]
    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(
            "Group labels are missing for exported subject(s): "
            f"{joined}. Fix project participant metadata before exporting stats-ready data."
        )
    return {subject: str(group) for subject, group in labels.items()}


def _safe_identifier(value: object) -> str:
    text = re.sub(r"[^0-9A-Za-z_]+", "_", _safe_text(value, default="value"))
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = "value"
    if text[0].isdigit():
        text = f"v_{text}"
    return text


def _wide_column_names(conditions: list[str], rois: list[str]) -> dict[tuple[str, str], str]:
    raw_names = {
        (condition, roi): f"{_safe_identifier(condition)}__{_safe_identifier(roi)}"
        for condition in conditions
        for roi in rois
    }
    counts: Counter[str] = Counter()
    out: dict[tuple[str, str], str] = {}
    for key, raw in raw_names.items():
        counts[raw] += 1
        out[key] = raw if counts[raw] == 1 else f"{raw}_{counts[raw]}"
    return out


def _build_long_frame(
    *,
    subjects: list[str],
    conditions: list[str],
    rois: list[str],
    subject_data: Mapping[str, Mapping[str, str]],
    summed_bca: Mapping[str, Mapping[str, Mapping[str, object]]],
    provenance_map: Mapping[tuple[str, str, str], Mapping[str, object]],
    dv_metadata: Mapping[str, object],
    dv_policy: Mapping[str, object] | None,
    group_map: Mapping[str, object] | None,
) -> pd.DataFrame:
    _ = dv_metadata, dv_policy, provenance_map
    group_labels = _resolve_group_labels(subjects, group_map)

    rows: list[dict[str, object]] = []
    for subject in subjects:
        for condition in conditions:
            condition_data = summed_bca.get(subject, {}).get(condition, {})
            for roi in rois:
                rows.append(
                    {
                        "subject_id": subject,
                        "group_id": group_labels[subject],
                        "condition": condition,
                        "roi": roi,
                        "summed_bca_uv": _numeric_or_nan(condition_data.get(roi)),
                    }
                )

    columns = [
        "subject_id",
        "group_id",
        "condition",
        "roi",
        "summed_bca_uv",
    ]
    frame = pd.DataFrame(rows, columns=columns)
    if frame.empty:
        raise RuntimeError("Stats-ready export produced no subject x condition x ROI rows.")
    if not frame["summed_bca_uv"].notna().any():
        raise RuntimeError(
            "Stats-ready export produced no finite Summed BCA values. "
            "Check source workbooks, selected ROIs, and harmonic policy settings."
        )
    return frame


def _build_jasp_wide_frame(
    long_df: pd.DataFrame,
    *,
    conditions: list[str],
    rois: list[str],
) -> pd.DataFrame:
    column_map = _wide_column_names(conditions, rois)
    id_columns = ["subject_id", "group_id"]
    subjects_df = long_df.loc[:, id_columns].drop_duplicates().reset_index(drop=True)

    wide = subjects_df.copy()
    for condition in conditions:
        for roi in rois:
            column = column_map[(condition, roi)]
            values = long_df[
                (long_df["condition"] == condition) & (long_df["roi"] == roi)
            ].loc[:, id_columns + ["summed_bca_uv"]]
            values = values.rename(columns={"summed_bca_uv": column})
            wide = wide.merge(values, on=id_columns, how="left")
    return wide


def build_stats_ready_frames(
    *,
    subjects: list[str],
    conditions: list[str],
    subject_data: Mapping[str, Mapping[str, str]],
    rois: Mapping[str, list[str]],
    summed_bca: Mapping[str, Mapping[str, Mapping[str, object]]],
    provenance_map: Mapping[tuple[str, str, str], Mapping[str, object]],
    dv_metadata: Mapping[str, object],
    dv_policy: Mapping[str, object] | None = None,
    group_map: Mapping[str, object] | None = None,
) -> dict[str, pd.DataFrame]:
    """Build all stats-ready workbook sheets from canonical Summed BCA data."""

    subject_list = [str(subject) for subject in subjects]
    condition_list = [str(condition) for condition in conditions]
    roi_names = [str(roi) for roi in rois.keys()]
    if not subject_list:
        raise RuntimeError("Stats-ready export requires at least one subject.")
    if len(condition_list) < 2:
        raise RuntimeError("Stats-ready export requires at least two selected conditions.")
    if not roi_names:
        raise RuntimeError("Stats-ready export requires at least one ROI.")

    long_df = _build_long_frame(
        subjects=subject_list,
        conditions=condition_list,
        rois=roi_names,
        subject_data=subject_data,
        summed_bca=summed_bca,
        provenance_map=provenance_map,
        dv_metadata=dv_metadata,
        dv_policy=dv_policy,
        group_map=group_map,
    )
    jasp_wide_df = _build_jasp_wide_frame(
        long_df,
        conditions=condition_list,
        rois=roi_names,
    )
    return {
        LONG_FORMAT_SHEET: long_df,
        WIDE_FORMAT_SHEET: jasp_wide_df,
        SELECTION_SUMMARY_SHEET: _build_selection_summary_frame(dv_metadata),
        HARMONIC_SELECTION_SHEET: _build_harmonic_selection_frame(dv_metadata),
    }


def _build_selection_summary_frame(dv_metadata: Mapping[str, object]) -> pd.DataFrame:
    fixed_meta = _common_fixed_predefined_meta(dv_metadata)
    group_meta = _common_group_significant_meta(dv_metadata)
    if group_meta:
        rows = _group_significant_summary_rows(group_meta)
    elif fixed_meta:
        rows = _fixed_predefined_summary_rows(fixed_meta)
    else:
        rossion_meta = _common_rossion_meta(dv_metadata)
        rows = _legacy_rossion_summary_rows(rossion_meta) if rossion_meta else []
    return pd.DataFrame(rows, columns=SELECTION_SUMMARY_COLUMNS)


def _summary_row(key: str, value: object) -> dict[str, object]:
    return {"Summary Item": key, "Value": _summary_value(value)}


def _summary_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set, Mapping)):
        return _sequence_cell(value)
    return value


def _highest_harmonic_hz(values: object) -> float:
    if not isinstance(values, (list, tuple, set)):
        return math.nan
    numeric = [_numeric_or_nan(value) for value in values]
    finite = [value for value in numeric if math.isfinite(value)]
    return max(finite) if finite else math.nan


def _harmonic_index(frequency_hz: object, oddball_hz: object) -> int | float:
    frequency = _numeric_or_nan(frequency_hz)
    oddball = _numeric_or_nan(oddball_hz)
    if (
        not (math.isfinite(frequency) and math.isfinite(oddball))
        or frequency <= 0
        or oddball <= 0
    ):
        return math.nan
    return int(round(frequency / oddball))


def _selected_harmonic_indices(values: object, oddball_hz: object) -> list[int]:
    if not isinstance(values, (list, tuple, set)):
        return []
    indices = {
        int(index)
        for value in values
        if math.isfinite(index := _harmonic_index(value, oddball_hz))
    }
    return sorted(index for index in indices if index > 0)


def _missing_harmonic_indices(selected_indices: list[int], highest_index: object) -> list[int]:
    highest = _numeric_or_nan(highest_index)
    if not math.isfinite(highest) or highest <= 0:
        return []
    selected_set = set(selected_indices)
    return [index for index in range(1, int(round(highest)) + 1) if index not in selected_set]


def _group_significant_summary_rows(group_meta: Mapping[str, object]) -> list[dict[str, object]]:
    included = group_meta.get("selected_harmonics_hz", []) or group_meta.get(
        "included_harmonics_hz",
        [],
    )
    detected = group_meta.get("detected_significant_harmonics_hz", []) or included
    oddball_hz = group_meta.get("oddball_frequency_hz")
    highest_hz = group_meta.get("highest_significant_harmonic_hz")
    if highest_hz in (None, ""):
        highest_hz = _highest_harmonic_hz(detected)
    highest_index = group_meta.get("highest_significant_harmonic_index")
    if highest_index in (None, ""):
        highest_index = _harmonic_index(highest_hz, oddball_hz)
    detected_indices = _selected_harmonic_indices(detected, oddball_hz)
    included_indices = _selected_harmonic_indices(included, oddball_hz)
    missing_detected_indices = _missing_harmonic_indices(detected_indices, highest_index)
    missing_included_indices = _missing_harmonic_indices(included_indices, highest_index)
    included_nonsignificant = sorted(set(included_indices).difference(detected_indices))
    harmonic_one_label = f"Harmonic 1 ({_format_number(oddball_hz)} Hz) significant?"
    rows = [
        _summary_row(
            "Harmonic policy",
            group_meta.get("harmonic_policy_label")
            or group_meta.get("harmonic_policy", GROUP_SIGNIFICANT_POLICY_ID),
        ),
        _summary_row("Selection scope", group_meta.get("selection_scope", "")),
        _summary_row("Base frequency (Hz)", _format_number(group_meta.get("base_frequency_hz"))),
        _summary_row("Oddball frequency (Hz)", _format_number(oddball_hz)),
        _summary_row("Z threshold", _format_number(group_meta.get("z_threshold"))),
        _summary_row("Electrode selection scope", group_meta.get("electrode_scope", "")),
        _summary_row("Summation method", group_meta.get("summation_method", "")),
        _summary_row("Highest significant harmonic (Hz)", _format_number(highest_hz)),
        _summary_row("Highest significant harmonic index", _format_number(highest_index)),
        _summary_row(harmonic_one_label, _yes_no(1 in detected_indices)),
        _summary_row(
            "All harmonics 1 through highest index significant?",
            _yes_no(not missing_detected_indices and bool(detected_indices)),
        ),
        _summary_row(
            "Non-significant harmonic indices within 1..highest",
            _format_sequence_cell(missing_detected_indices)
            if missing_detected_indices
            else "None",
        ),
        _summary_row(
            "All non-base harmonics 1 through highest included?",
            _yes_no(not missing_included_indices and bool(included_indices)),
        ),
        _summary_row(
            "Non-significant harmonic indices included in summation",
            _format_sequence_cell(included_nonsignificant)
            if included_nonsignificant
            else "None",
        ),
        _summary_row("Significant harmonic indices", _format_sequence_cell(detected_indices)),
        _summary_row("Significant harmonic frequencies (Hz)", _format_sequence_cell(detected)),
        _summary_row("Included harmonic indices", _format_sequence_cell(included_indices)),
        _summary_row("Included harmonic frequencies (Hz)", _format_sequence_cell(included)),
        _summary_row("Selection source", group_meta.get("selection_cache_source", "")),
        _summary_row("Selection saved at", group_meta.get("selection_cache_saved_at", "")),
    ]
    return rows


def _fixed_predefined_summary_rows(fixed_meta: Mapping[str, object]) -> list[dict[str, object]]:
    selected = fixed_meta.get("fixed_harmonic_included_frequencies_hz", [])
    highest_hz = fixed_meta.get("highest_selected_harmonic_hz")
    if highest_hz in (None, ""):
        highest_hz = _highest_harmonic_hz(selected)
    highest_index = fixed_meta.get("highest_selected_harmonic_index")
    if highest_index in (None, ""):
        highest_index = _harmonic_index(highest_hz, fixed_meta.get("oddball_frequency_hz"))
    return [
        _summary_row(
            "Harmonic policy",
            fixed_meta.get("harmonic_policy_label")
            or fixed_meta.get("harmonic_policy", FIXED_PREDEFINED_POLICY_ID),
        ),
        _summary_row("Selected harmonics (Hz)", _format_sequence_cell(selected)),
        _summary_row("Highest selected harmonic (Hz)", _format_number(highest_hz)),
        _summary_row("Highest selected harmonic index", _format_number(highest_index)),
        _summary_row(
            "SNR used for statistics?",
            _yes_no(bool(fixed_meta.get("snr_used_for_statistics"))),
        ),
        _summary_row(
            "Applied uniformly across participants?",
            _yes_no(bool(fixed_meta.get("applied_uniformly_across_participants"))),
        ),
        _summary_row(
            "Applied uniformly across conditions?",
            _yes_no(bool(fixed_meta.get("applied_uniformly_across_conditions"))),
        ),
        _summary_row(
            "Applied uniformly across ROIs?",
            _yes_no(bool(fixed_meta.get("applied_uniformly_across_rois"))),
        ),
    ]


def _legacy_rossion_summary_rows(rossion_meta: Mapping[str, object]) -> list[dict[str, object]]:
    return [
        _summary_row("Harmonic policy", "legacy_rossion_method"),
        _summary_row("Selection scope", rossion_meta.get("selection_scope", "")),
        _summary_row(
            "Selected harmonics (Hz)",
            _format_sequence_cell(rossion_meta.get("common_harmonics_hz", "")),
        ),
    ]


def _build_harmonic_selection_frame(dv_metadata: Mapping[str, object]) -> pd.DataFrame:
    fixed_meta = _common_fixed_predefined_meta(dv_metadata)
    if fixed_meta:
        rows = _fixed_predefined_selection_rows(fixed_meta)
        return _harmonic_selection_frame_from_rows(rows)
    group_meta = _common_group_significant_meta(dv_metadata)
    if group_meta:
        rows = _group_significant_selection_rows(group_meta)
        return _harmonic_selection_frame_from_rows(rows)
    rossion_meta = _common_rossion_meta(dv_metadata)
    if not rossion_meta:
        return pd.DataFrame()
    rows: list[dict[str, object]] = []
    _append_selection_rows(
        rows,
        selection_scope=_safe_text(rossion_meta.get("selection_scope")),
        z_map=rossion_meta.get("selection_z_by_harmonic"),
        selected_harmonics=rossion_meta.get("common_harmonics_hz"),
        excluded_harmonics=rossion_meta.get("excluded_base_harmonics_hz"),
        sensitivity=False,
    )
    sensitivity_meta = rossion_meta.get("sensitivity_union_roi_electrodes")
    if isinstance(sensitivity_meta, Mapping):
        _append_selection_rows(
            rows,
            selection_scope=_safe_text(sensitivity_meta.get("selection_scope")),
            z_map=sensitivity_meta.get("selection_z_by_harmonic"),
            selected_harmonics=sensitivity_meta.get("harmonics_hz"),
            excluded_harmonics=sensitivity_meta.get("excluded_base_harmonics_hz"),
            sensitivity=True,
        )
    return _harmonic_selection_frame_from_rows(rows)


def _harmonic_selection_frame_from_rows(rows: list[dict[str, object]]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=HARMONIC_SELECTION_COLUMNS)


def _fixed_predefined_selection_rows(fixed_meta: Mapping[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row_data in fixed_meta.get("selection_rows", []) or []:
        if not isinstance(row_data, Mapping):
            continue
        requested = _parse_frequency(row_data.get("requested_frequency_hz"))
        rows.append(
            {
                "requested_harmonic_hz": requested,
                "z_score": math.nan,
                "target_amplitude_uv": math.nan,
                "noise_mean_uv": math.nan,
                "noise_std_uv": math.nan,
                "selected": bool(row_data.get("included")),
                "included_in_summation": bool(row_data.get("included")),
                "excluded_base_rate": row_data.get("exclusion_reason") == "base_rate_overlap",
                "exclusion_reason": row_data.get("exclusion_reason", ""),
                "warning": row_data.get("warning", ""),
            }
        )
    return rows


def _group_significant_selection_rows(group_meta: Mapping[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row_data in group_meta.get("selection_rows", []) or []:
        if not isinstance(row_data, Mapping):
            continue
        rows.append(
            {
                "requested_harmonic_hz": row_data.get("target_frequency_hz"),
                "z_score": _numeric_or_nan(row_data.get("z_score")),
                "target_amplitude_uv": _numeric_or_nan(row_data.get("target_amplitude_uv")),
                "noise_mean_uv": _numeric_or_nan(row_data.get("noise_mean_uv")),
                "noise_std_uv": _numeric_or_nan(row_data.get("noise_std_uv")),
                "selected": bool(row_data.get("selected")),
                "included_in_summation": bool(
                    row_data.get("included_in_summation", row_data.get("selected"))
                ),
                "excluded_base_rate": bool(row_data.get("excluded_base_rate")),
                "exclusion_reason": row_data.get("exclusion_reason", ""),
                "warning": row_data.get("warning", ""),
            }
        )
    return rows


def _append_selection_rows(
    rows: list[dict[str, object]],
    *,
    selection_scope: str,
    z_map: object,
    selected_harmonics: object,
    excluded_harmonics: object,
    sensitivity: bool,
) -> None:
    _ = selection_scope
    selected = {
        float(freq)
        for freq in (selected_harmonics if isinstance(selected_harmonics, (list, tuple, set)) else [])
    }
    excluded = {
        float(freq)
        for freq in (excluded_harmonics if isinstance(excluded_harmonics, (list, tuple, set)) else [])
    }
    seen_harmonics: set[float] = set()
    if isinstance(z_map, Mapping):
        for raw_freq, raw_z in z_map.items():
            freq = _parse_frequency(raw_freq)
            if freq is None:
                continue
            seen_harmonics.add(freq)
            rows.append(
                {
                    "requested_harmonic_hz": freq,
                    "z_score": _numeric_or_nan(raw_z),
                    "target_amplitude_uv": math.nan,
                    "noise_mean_uv": math.nan,
                    "noise_std_uv": math.nan,
                    "selected": freq in selected,
                    "included_in_summation": freq in selected,
                    "excluded_base_rate": freq in excluded,
                    "exclusion_reason": "base_rate_overlap" if freq in excluded else "",
                    "warning": "Sensitivity selection row." if sensitivity else "",
                }
            )
    for freq in sorted(excluded.difference(seen_harmonics)):
        rows.append(
            {
                "requested_harmonic_hz": freq,
                "z_score": math.nan,
                "target_amplitude_uv": math.nan,
                "noise_mean_uv": math.nan,
                "noise_std_uv": math.nan,
                "selected": False,
                "included_in_summation": False,
                "excluded_base_rate": True,
                "exclusion_reason": "base_rate_overlap",
                "warning": "Sensitivity selection row." if sensitivity else "",
            }
        )


def write_stats_ready_workbook(
    save_path: str | Path,
    frames: Mapping[str, pd.DataFrame],
) -> Path:
    """Write the stats-ready workbook and format the harmonic summary sheet."""

    target = Path(save_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        for worksheet in workbook.worksheets:
            _format_table_sheet(worksheet)
        if SELECTION_SUMMARY_SHEET in workbook.sheetnames:
            _format_selection_summary_sheet(workbook[SELECTION_SUMMARY_SHEET])
    return target


def _format_table_sheet(worksheet: object) -> None:
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 14),
            72,
        )


def _format_selection_summary_sheet(worksheet: object) -> None:
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.freeze_panes = "A2"


def prepare_stats_ready_export(
    *,
    subjects: list[str],
    conditions: list[str],
    subject_data: Mapping[str, Mapping[str, str]],
    base_freq: float,
    rois: Mapping[str, list[str]],
    dv_policy: Mapping[str, object] | None,
    group_map: Mapping[str, object] | None,
    log_func: Callable[[str], None],
    save_path: str | Path | None = None,
    max_freq: float | None = None,
    selection_conditions: list[str] | None = None,
    project_root: str | None = None,
) -> StatsReadyExport:
    """Prepare and optionally write the external-statistics Summed BCA export."""

    started_at = perf_counter()
    log_func(
        "Stats-ready export: preparing Summed BCA data "
        f"for {len(subjects)} participants x {len(conditions)} selected conditions."
    )
    logger.debug(
        "stats_ready_prepare_summed_bca_start",
        extra={"subjects": len(subjects), "conditions": len(conditions)},
    )
    provenance_map: dict[tuple[str, str, str], dict[str, object]] = {}
    dv_metadata: dict[str, object] = {}
    summed_bca = prepare_summed_bca_data(
        subjects=list(subjects),
        conditions=list(conditions),
        subject_data=dict(subject_data),
        base_freq=float(base_freq),
        log_func=log_func,
        rois=dict(rois),
        provenance_map=provenance_map,
        dv_policy=dict(dv_policy or {}),
        dv_metadata=dv_metadata,
        max_freq=max_freq,
        selection_conditions=selection_conditions,
        project_root=project_root,
    )
    if not summed_bca:
        raise RuntimeError("Stats-ready export could not prepare Summed BCA data.")

    log_func(
        "Stats-ready export: Summed BCA data prepared "
        f"in {perf_counter() - started_at:.1f}s; building workbook frames."
    )
    logger.debug(
        "stats_ready_prepare_summed_bca_done",
        extra={"elapsed_s": perf_counter() - started_at},
    )
    frames_started = perf_counter()
    frames = build_stats_ready_frames(
        subjects=list(subjects),
        conditions=list(conditions),
        subject_data=subject_data,
        rois=rois,
        summed_bca=summed_bca,
        provenance_map=provenance_map,
        dv_metadata=dv_metadata,
        dv_policy=dv_policy,
        group_map=group_map,
    )
    log_func(
        "Stats-ready export: workbook frames built "
        f"in {perf_counter() - frames_started:.1f}s."
    )
    logger.debug(
        "stats_ready_frames_built",
        extra={"elapsed_s": perf_counter() - frames_started, "sheets": len(frames)},
    )
    workbook_path = None
    if save_path:
        write_started = perf_counter()
        log_func(f"Stats-ready export: writing workbook to {save_path}.")
        logger.debug("stats_ready_workbook_write_start", extra={"path": str(save_path)})
        workbook_path = write_stats_ready_workbook(save_path, frames)
        log_func(
            "Stats-ready export: workbook write finished "
            f"in {perf_counter() - write_started:.1f}s."
        )
        logger.debug(
            "stats_ready_workbook_write_done",
            extra={"elapsed_s": perf_counter() - write_started, "path": str(save_path)},
        )
    log_func(
        "Stats-ready export: finished "
        f"in {perf_counter() - started_at:.1f}s."
    )
    logger.debug("stats_ready_export_done", extra={"elapsed_s": perf_counter() - started_at})
    return StatsReadyExport(
        frames=frames,
        workbook_path=workbook_path,
        row_count=len(frames[LONG_FORMAT_SHEET]),
    )
