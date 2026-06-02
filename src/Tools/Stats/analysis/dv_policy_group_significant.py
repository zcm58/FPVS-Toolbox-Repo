"""Group-level significant-harmonic Summed BCA DV policy helpers."""
from __future__ import annotations

import logging
import threading
from dataclasses import dataclass, replace
from pathlib import Path
from time import perf_counter
from typing import Callable, Dict, List, Optional, Sequence

import numpy as np
from openpyxl import load_workbook
import pandas as pd

from Tools.Stats.analysis.dv_policy_settings import (
    DVPolicySettings,
    GROUP_SIGNIFICANT_POLICY_ID,
    GROUP_SIGNIFICANT_POLICY_LABEL,
    GROUP_SIGNIFICANT_POLICY_NAME,
    LOCKED_ODDBALL_FREQUENCY_HZ,
)
from Tools.Stats.analysis.stats_analysis import _current_rois_map
from Tools.Stats.data.group_harmonic_cache import (
    GroupHarmonicCacheRequest,
    build_group_harmonic_cache_request,
    lookup_cached_group_harmonic_selection,
    save_cached_group_harmonic_selection,
)
from Tools.Stats.io.excel_io import safe_read_excel

logger = logging.getLogger("Tools.Stats")

FULL_FFT_AMPLITUDE_SHEET_NAME = "FullFFT Amplitude (uV)"
GROUP_SIGNIFICANT_BASE_TOLERANCE_HZ = 0.01
GROUP_SIGNIFICANT_MATCHING_TOLERANCE_HZ = 0.01
GROUP_SIGNIFICANT_NOISE_WINDOW_BINS = 10
GROUP_SIGNIFICANT_FULLFFT_PROGRESS_INTERVAL = 5
GROUP_SIGNIFICANT_BCA_PROGRESS_INTERVAL = 10
GROUP_SIGNIFICANT_SELECTION_CACHE_MAX_ENTRIES = 8
_GROUP_SELECTION_CACHE_LOCK = threading.Lock()
_GROUP_SELECTION_CACHE: dict[
    "GroupSignificantSelectionCacheKey",
    "GroupSignificantHarmonicSelection",
] = {}


@dataclass(frozen=True)
class GroupSignificantHarmonicRow:
    harmonic_index: int
    target_frequency_hz: float
    matched_frequency_hz: float | None
    matched_column: str | None
    matched_bin_index: int | None
    z_score: float | None
    selected: bool
    excluded_base_rate: bool
    exclusion_reason: str
    warning: str
    target_amplitude_uv: float | None = None
    noise_mean_uv: float | None = None
    noise_std_uv: float | None = None
    noise_bin_indices: tuple[int, ...] = ()
    noise_frequencies_hz: tuple[float, ...] = ()
    noise_amplitudes_uv: tuple[float, ...] = ()
    noise_used_bin_indices: tuple[int, ...] = ()
    noise_used_frequencies_hz: tuple[float, ...] = ()
    noise_used_amplitudes_uv: tuple[float, ...] = ()


@dataclass(frozen=True)
class GroupSignificantNoiseStats:
    mean_uv: float
    std_uv: float
    candidate_bin_indices: tuple[int, ...]
    used_bin_indices: tuple[int, ...]


@dataclass(frozen=True)
class GroupSignificantHarmonicSelection:
    harmonic_domain_hz: list[float]
    selected_harmonics_hz: list[float]
    selected_columns: list[str]
    selected_bin_indices: list[int]
    z_by_harmonic: dict[float, float]
    excluded_base_harmonics_hz: list[float]
    oddball_frequency_hz: float
    base_frequency_hz: float
    z_threshold: float
    electrode_scope: str
    selection_scope: str
    selection_conditions: list[str]
    selection_subjects: list[str]
    selection_spectra_count: int
    selection_electrode_count: int
    frequency_resolution_hz: float | None
    base_overlap_tolerance_hz: float
    matching_tolerance_hz: float
    noise_window_bins: int
    rows: list[GroupSignificantHarmonicRow]
    selection_cache_source: str = "computed_this_run"
    selection_cache_saved_at: str | None = None
    selection_cache_key: str | None = None

    def to_metadata(self) -> dict[str, object]:
        highest_meta = _highest_selected_harmonic_metadata(
            self.selected_harmonics_hz,
            oddball_hz=self.oddball_frequency_hz,
            prefix="highest_significant_harmonic",
        )
        return {
            "harmonic_policy": GROUP_SIGNIFICANT_POLICY_ID,
            "harmonic_policy_label": GROUP_SIGNIFICANT_POLICY_LABEL,
            "dependent_variable": "summed_bca",
            "selection_source_sheet": FULL_FFT_AMPLITUDE_SHEET_NAME,
            "selection_amplitude_summary": "grand_average_raw_amplitude_spectrum",
            "selection_scope": self.selection_scope,
            "selection_conditions": list(self.selection_conditions),
            "selection_subjects": list(self.selection_subjects),
            "selection_spectra_count": int(self.selection_spectra_count),
            "selection_electrode_count": int(self.selection_electrode_count),
            "electrode_scope": self.electrode_scope,
            "z_threshold": float(self.z_threshold),
            "z_score_source": "computed_from_grand_averaged_amplitude_spectrum",
            "noise_window_bins": int(self.noise_window_bins),
            "base_frequency_hz": float(self.base_frequency_hz),
            "oddball_frequency_hz": float(self.oddball_frequency_hz),
            "base_overlap_exclusion_enabled": True,
            "base_overlap_tolerance_hz": float(self.base_overlap_tolerance_hz),
            "matching_tolerance_hz": float(self.matching_tolerance_hz),
            "frequency_resolution_hz": self.frequency_resolution_hz,
            "harmonic_domain_hz": list(self.harmonic_domain_hz),
            "common_harmonics_hz": list(self.selected_harmonics_hz),
            "selected_harmonics_hz": list(self.selected_harmonics_hz),
            **highest_meta,
            "selected_columns": list(self.selected_columns),
            "selected_bin_indices": list(self.selected_bin_indices),
            "selection_z_by_harmonic": dict(self.z_by_harmonic),
            "excluded_base_harmonics_hz": list(self.excluded_base_harmonics_hz),
            "applied_uniformly_across_participants": True,
            "applied_uniformly_across_conditions": True,
            "applied_uniformly_across_rois": True,
            "snr_used_for_statistics": False,
            "bca_negative_values_retained": True,
            "bca_near_zero_values_retained": True,
            "selection_rows": [
                {
                    "harmonic_index": row.harmonic_index,
                    "target_frequency_hz": row.target_frequency_hz,
                    "matched_frequency_hz": row.matched_frequency_hz,
                    "matched_column": row.matched_column,
                    "matched_bin_index": row.matched_bin_index,
                    "z_score": row.z_score,
                    "selected": row.selected,
                    "excluded_base_rate": row.excluded_base_rate,
                    "exclusion_reason": row.exclusion_reason,
                    "warning": row.warning,
                    "target_amplitude_uv": row.target_amplitude_uv,
                    "noise_mean_uv": row.noise_mean_uv,
                    "noise_std_uv": row.noise_std_uv,
                    "noise_bin_indices": list(row.noise_bin_indices),
                    "noise_frequencies_hz": list(row.noise_frequencies_hz),
                    "noise_amplitudes_uv": list(row.noise_amplitudes_uv),
                    "noise_used_bin_indices": list(row.noise_used_bin_indices),
                    "noise_used_frequencies_hz": list(row.noise_used_frequencies_hz),
                    "noise_used_amplitudes_uv": list(row.noise_used_amplitudes_uv),
                }
                for row in self.rows
            ],
            "selection_cache_source": self.selection_cache_source,
            "selection_cache_saved_at": self.selection_cache_saved_at,
            "selection_cache_key": self.selection_cache_key,
            "methods_summary": _methods_summary(self),
        }


@dataclass(frozen=True)
class RequiredFullFftColumns:
    usecols: list[str]
    frequency_columns: list[tuple[float, str, int]]
    candidate_indices: list[int]
    excluded_base_indices: list[int]
    required_indices: list[int]


@dataclass(frozen=True)
class WorkbookSignature:
    subject: str
    condition: str
    path: str
    size_bytes: int | None
    mtime_ns: int | None


@dataclass(frozen=True)
class GroupSignificantSelectionCacheKey:
    subjects: tuple[str, ...]
    conditions: tuple[str, ...]
    workbooks: tuple[WorkbookSignature, ...]
    rois: tuple[tuple[str, tuple[str, ...]], ...]
    base_frequency_hz: float
    oddball_frequency_hz: float | None
    max_freq_hz: float | None
    z_threshold: float
    electrode_scope: str
    project_processing_signature_hash: str | None = None


def clear_group_significant_selection_cache() -> None:
    with _GROUP_SELECTION_CACHE_LOCK:
        _GROUP_SELECTION_CACHE.clear()


def preflight_group_significant_full_fft_columns(
    *,
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
    base_frequency_hz: float,
    log_func: Callable[[str], None],
    max_freq: float | None = None,
) -> int:
    """Validate exact FullFFT harmonic columns before expensive Stats reads."""
    required = _plan_required_full_fft_columns(
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        base_frequency_hz=base_frequency_hz,
        max_freq=max_freq,
        log_func=log_func,
    )
    return _preflight_required_full_fft_columns(
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        required=required,
        log_func=log_func,
    )


def _get_cached_group_significant_selection(
    cache_key: GroupSignificantSelectionCacheKey,
) -> GroupSignificantHarmonicSelection | None:
    with _GROUP_SELECTION_CACHE_LOCK:
        return _GROUP_SELECTION_CACHE.get(cache_key)


def _store_group_significant_selection(
    cache_key: GroupSignificantSelectionCacheKey,
    selection: GroupSignificantHarmonicSelection,
) -> None:
    with _GROUP_SELECTION_CACHE_LOCK:
        if (
            cache_key not in _GROUP_SELECTION_CACHE
            and len(_GROUP_SELECTION_CACHE) >= GROUP_SIGNIFICANT_SELECTION_CACHE_MAX_ENTRIES
        ):
            oldest_key = next(iter(_GROUP_SELECTION_CACHE))
            _GROUP_SELECTION_CACHE.pop(oldest_key, None)
        _GROUP_SELECTION_CACHE[cache_key] = selection


def _group_significant_selection_cache_key(
    *,
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
    rois: Dict[str, List[str]],
    base_frequency_hz: float,
    max_freq: float | None,
    settings: DVPolicySettings,
    project_processing_signature_hash: str | None = None,
) -> GroupSignificantSelectionCacheKey:
    subject_key = tuple(str(subject) for subject in subjects)
    condition_key = tuple(str(condition) for condition in conditions)
    workbook_signatures = tuple(
        _workbook_signature(
            subject=subject,
            condition=condition,
            file_path=subject_data.get(subject, {}).get(condition),
        )
        for subject in subject_key
        for condition in condition_key
    )
    rois_key = tuple(
        (str(roi_name), tuple(str(channel).upper().strip() for channel in channels or ()))
        for roi_name, channels in sorted((rois or {}).items())
    )
    return GroupSignificantSelectionCacheKey(
        subjects=subject_key,
        conditions=condition_key,
        workbooks=workbook_signatures,
        rois=rois_key,
        base_frequency_hz=float(base_frequency_hz),
        oddball_frequency_hz=LOCKED_ODDBALL_FREQUENCY_HZ,
        max_freq_hz=float(max_freq) if max_freq is not None else None,
        z_threshold=float(settings.group_significant_z_threshold),
        electrode_scope=str(settings.group_significant_electrode_scope),
        project_processing_signature_hash=project_processing_signature_hash,
    )


def _log_project_cache_warnings(
    cache_request: GroupHarmonicCacheRequest | None,
    log_func: Callable[[str], None],
) -> None:
    if cache_request is None:
        return
    for warning in cache_request.ledger_warnings:
        log_func(f"Warning: {warning}")


def _load_project_cached_selection(
    cache_request: GroupHarmonicCacheRequest | None,
    log_func: Callable[[str], None],
) -> GroupSignificantHarmonicSelection | None:
    lookup = lookup_cached_group_harmonic_selection(cache_request)
    if lookup.hit is None:
        reason = lookup.reason
        if reason and reason != "No saved group-significant harmonics.":
            log_func(f"Project harmonic cache miss: {reason}")
        return None
    try:
        selection = group_significant_selection_from_metadata(
            lookup.hit.selection_metadata,
        )
    except Exception as exc:  # noqa: BLE001
        log_func(
            "Project harmonic cache entry could not be read; recalculating "
            f"group-level significant harmonics. Error: {exc}"
        )
        logger.warning("stats_group_harmonics_project_cache_invalid", exc_info=True)
        return None
    return replace(
        selection,
        selection_cache_source="saved_project_metadata",
        selection_cache_saved_at=lookup.hit.saved_at,
        selection_cache_key=lookup.hit.cache_key,
    )


def _save_project_cached_selection(
    cache_request: GroupHarmonicCacheRequest | None,
    selection: GroupSignificantHarmonicSelection,
    log_func: Callable[[str], None],
) -> GroupSignificantHarmonicSelection:
    if cache_request is None:
        return selection
    try:
        saved_at = save_cached_group_harmonic_selection(
            cache_request,
            selection.to_metadata(),
        )
    except Exception as exc:  # noqa: BLE001
        log_func(
            "Warning: could not save selected significant harmonics to project metadata; "
            f"future exports may need to recalculate. Error: {exc}"
        )
        logger.warning("stats_group_harmonics_project_cache_save_failed", exc_info=True)
        return selection
    if not saved_at:
        return selection
    log_func(
        "Saved group-level significant harmonics to project metadata for future exports."
    )
    return replace(
        selection,
        selection_cache_source="computed_this_run_saved_project_metadata",
        selection_cache_saved_at=saved_at,
        selection_cache_key=cache_request.cache_key,
    )


def group_significant_selection_from_metadata(
    metadata: Dict[str, object],
) -> GroupSignificantHarmonicSelection:
    """Rehydrate a saved group-significant selection from manifest metadata."""

    if not isinstance(metadata, dict):
        raise TypeError("selection metadata must be a dictionary")
    selected_harmonics = _metadata_float_list(metadata.get("selected_harmonics_hz"))
    if not selected_harmonics:
        raise ValueError("saved selection contains no selected harmonics")
    oddball = _metadata_float(
        metadata.get("oddball_frequency_hz"),
        default=LOCKED_ODDBALL_FREQUENCY_HZ,
    )
    rows = [
        _group_significant_row_from_metadata(row_data)
        for row_data in (metadata.get("selection_rows") or [])
        if isinstance(row_data, dict)
    ]
    selected_columns = [
        str(column)
        for column in _metadata_sequence(metadata.get("selected_columns"))
        if str(column).strip()
    ]
    if not selected_columns:
        selected_columns = [f"{freq:.4f}_Hz" for freq in selected_harmonics]
    return GroupSignificantHarmonicSelection(
        harmonic_domain_hz=_metadata_float_list(metadata.get("harmonic_domain_hz")),
        selected_harmonics_hz=selected_harmonics,
        selected_columns=selected_columns,
        selected_bin_indices=_metadata_int_list(metadata.get("selected_bin_indices")),
        z_by_harmonic=_metadata_float_map(metadata.get("selection_z_by_harmonic")),
        excluded_base_harmonics_hz=_metadata_float_list(
            metadata.get("excluded_base_harmonics_hz")
        ),
        oddball_frequency_hz=oddball,
        base_frequency_hz=_metadata_float(metadata.get("base_frequency_hz"), default=np.nan),
        z_threshold=_metadata_float(metadata.get("z_threshold"), default=np.nan),
        electrode_scope=str(metadata.get("electrode_scope") or ""),
        selection_scope=str(metadata.get("selection_scope") or ""),
        selection_conditions=[
            str(value) for value in _metadata_sequence(metadata.get("selection_conditions"))
        ],
        selection_subjects=[
            str(value) for value in _metadata_sequence(metadata.get("selection_subjects"))
        ],
        selection_spectra_count=_metadata_int(metadata.get("selection_spectra_count"), default=0),
        selection_electrode_count=_metadata_int(
            metadata.get("selection_electrode_count"),
            default=0,
        ),
        frequency_resolution_hz=_metadata_optional_float(
            metadata.get("frequency_resolution_hz")
        ),
        base_overlap_tolerance_hz=_metadata_float(
            metadata.get("base_overlap_tolerance_hz"),
            default=GROUP_SIGNIFICANT_BASE_TOLERANCE_HZ,
        ),
        matching_tolerance_hz=_metadata_float(
            metadata.get("matching_tolerance_hz"),
            default=GROUP_SIGNIFICANT_MATCHING_TOLERANCE_HZ,
        ),
        noise_window_bins=_metadata_int(
            metadata.get("noise_window_bins"),
            default=GROUP_SIGNIFICANT_NOISE_WINDOW_BINS,
        ),
        rows=rows,
        selection_cache_source=str(metadata.get("selection_cache_source") or "saved_project_metadata"),
        selection_cache_saved_at=(
            str(metadata.get("selection_cache_saved_at"))
            if metadata.get("selection_cache_saved_at") not in (None, "")
            else None
        ),
        selection_cache_key=(
            str(metadata.get("selection_cache_key"))
            if metadata.get("selection_cache_key") not in (None, "")
            else None
        ),
    )


def _group_significant_row_from_metadata(row_data: dict[str, object]) -> GroupSignificantHarmonicRow:
    return GroupSignificantHarmonicRow(
        harmonic_index=_metadata_int(row_data.get("harmonic_index"), default=0),
        target_frequency_hz=_metadata_float(row_data.get("target_frequency_hz"), default=np.nan),
        matched_frequency_hz=_metadata_optional_float(row_data.get("matched_frequency_hz")),
        matched_column=(
            str(row_data.get("matched_column"))
            if row_data.get("matched_column") not in (None, "")
            else None
        ),
        matched_bin_index=_metadata_optional_int(row_data.get("matched_bin_index")),
        z_score=_metadata_optional_float(row_data.get("z_score")),
        selected=bool(row_data.get("selected")),
        excluded_base_rate=bool(row_data.get("excluded_base_rate")),
        exclusion_reason=str(row_data.get("exclusion_reason") or ""),
        warning=str(row_data.get("warning") or ""),
        target_amplitude_uv=_metadata_optional_float(row_data.get("target_amplitude_uv")),
        noise_mean_uv=_metadata_optional_float(row_data.get("noise_mean_uv")),
        noise_std_uv=_metadata_optional_float(row_data.get("noise_std_uv")),
        noise_bin_indices=tuple(_metadata_int_list(row_data.get("noise_bin_indices"))),
        noise_frequencies_hz=tuple(_metadata_float_list(row_data.get("noise_frequencies_hz"))),
        noise_amplitudes_uv=tuple(_metadata_float_list(row_data.get("noise_amplitudes_uv"))),
        noise_used_bin_indices=tuple(
            _metadata_int_list(row_data.get("noise_used_bin_indices"))
        ),
        noise_used_frequencies_hz=tuple(
            _metadata_float_list(row_data.get("noise_used_frequencies_hz"))
        ),
        noise_used_amplitudes_uv=tuple(
            _metadata_float_list(row_data.get("noise_used_amplitudes_uv"))
        ),
    )


def _metadata_sequence(value: object) -> list[object]:
    if isinstance(value, (list, tuple, set)):
        return list(value)
    return []


def _metadata_float(value: object, *, default: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return float(default)
    return float(number) if np.isfinite(number) else float(default)


def _metadata_optional_float(value: object) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return float(number) if np.isfinite(number) else None


def _metadata_int(value: object, *, default: int) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return int(default)
    return int(number)


def _metadata_optional_int(value: object) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _metadata_float_list(value: object) -> list[float]:
    out: list[float] = []
    for item in _metadata_sequence(value):
        number = _metadata_optional_float(item)
        if number is not None:
            out.append(float(number))
    return out


def _metadata_int_list(value: object) -> list[int]:
    out: list[int] = []
    for item in _metadata_sequence(value):
        number = _metadata_optional_int(item)
        if number is not None:
            out.append(int(number))
    return out


def _metadata_float_map(value: object) -> dict[float, float]:
    if not isinstance(value, dict):
        return {}
    out: dict[float, float] = {}
    for raw_key, raw_value in value.items():
        key = _metadata_optional_float(raw_key)
        map_value = _metadata_optional_float(raw_value)
        if key is not None and map_value is not None:
            out[float(key)] = float(map_value)
    return out


def _workbook_signature(
    *,
    subject: str,
    condition: str,
    file_path: str | None,
) -> WorkbookSignature:
    if not file_path:
        return WorkbookSignature(
            subject=str(subject),
            condition=str(condition),
            path="",
            size_bytes=None,
            mtime_ns=None,
        )
    path = Path(file_path)
    try:
        resolved = str(path.resolve(strict=False))
    except OSError:
        resolved = str(path)
    try:
        stat = path.stat()
    except OSError:
        return WorkbookSignature(
            subject=str(subject),
            condition=str(condition),
            path=resolved,
            size_bytes=None,
            mtime_ns=None,
        )
    return WorkbookSignature(
        subject=str(subject),
        condition=str(condition),
        path=resolved,
        size_bytes=int(stat.st_size),
        mtime_ns=int(stat.st_mtime_ns),
    )


def build_group_significant_harmonic_selection(
    *,
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
    base_frequency_hz: float,
    rois: Dict[str, List[str]],
    log_func: Callable[[str], None],
    settings: DVPolicySettings,
    max_freq: float | None = None,
    project_root: str | Path | None = None,
) -> GroupSignificantHarmonicSelection:
    started = perf_counter()
    cache_request = build_group_harmonic_cache_request(
        project_root=project_root,
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        base_frequency_hz=base_frequency_hz,
        max_freq_hz=max_freq,
        settings=settings,
    )
    _log_project_cache_warnings(cache_request, log_func)
    cache_key = _group_significant_selection_cache_key(
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        rois=rois,
        base_frequency_hz=base_frequency_hz,
        max_freq=max_freq,
        settings=settings,
        project_processing_signature_hash=(
            cache_request.project_processing_signature_hash
            if cache_request is not None
            else None
        ),
    )
    cached = _get_cached_group_significant_selection(cache_key)
    if cached is not None:
        elapsed = perf_counter() - started
        log_func(
            "[PERF] Group harmonic selection cache hit: "
            f"reusing {len(cached.selected_harmonics_hz)} selected harmonics "
            f"for {len(subjects) * len(conditions)} planned workbooks "
            f"in {elapsed:.2f}s."
        )
        logger.info(
            "stats_group_harmonics_selection_cache_hit",
            extra={
                "elapsed_s": elapsed,
                "subjects": len(subjects),
                "conditions": len(conditions),
                "selected_harmonics": cached.selected_harmonics_hz,
            },
        )
        return cached
    project_cached = _load_project_cached_selection(cache_request, log_func)
    if project_cached is not None:
        elapsed = perf_counter() - started
        log_func(
            "[PERF] Project harmonic cache hit: "
            f"using {len(project_cached.selected_harmonics_hz)} saved harmonics "
            f"for {len(subjects) * len(conditions)} planned workbooks "
            f"in {elapsed:.2f}s."
        )
        logger.info(
            "stats_group_harmonics_project_cache_hit",
            extra={
                "elapsed_s": elapsed,
                "subjects": len(subjects),
                "conditions": len(conditions),
                "selected_harmonics": project_cached.selected_harmonics_hz,
            },
        )
        _store_group_significant_selection(cache_key, project_cached)
        return project_cached
    log_func(
        "[PERF] Group harmonic selection cache miss: "
        "building selection from FullFFT amplitude spectra."
    )
    logger.info(
        "stats_group_harmonics_selection_cache_miss",
        extra={"subjects": len(subjects), "conditions": len(conditions)},
    )
    base = float(base_frequency_hz)
    oddball = _resolve_group_oddball_frequency(
        base_frequency_hz=base,
    )
    required = _plan_required_full_fft_columns(
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        base_frequency_hz=base,
        max_freq=max_freq,
        log_func=log_func,
    )
    planned_workbook_count = _preflight_required_full_fft_columns(
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        required=required,
        log_func=log_func,
    )
    log_func(
        "[PERF] Group harmonic selection column plan: "
        f"{len(required.usecols) - 1} FullFFT frequency columns needed "
        f"from {len(required.frequency_columns)} available columns "
        f"across {planned_workbook_count} workbooks."
    )
    logger.info(
        "stats_group_harmonics_column_plan",
        extra={
            "needed_frequency_columns": len(required.usecols) - 1,
            "available_frequency_columns": len(required.frequency_columns),
            "candidate_indices": len(required.candidate_indices),
        },
    )
    grand_average, columns, bin_indices, spectra_count, electrode_count = _build_grand_average_amplitude(
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        rois=rois,
        electrode_scope=settings.group_significant_electrode_scope,
        log_func=log_func,
        frequency_columns=required.frequency_columns,
        required_indices=required.required_indices,
    )
    if grand_average.empty:
        raise RuntimeError("Group-level harmonic selection found no usable amplitude spectra.")

    frequency_resolution = _frequency_resolution(
        [freq for freq, _column, _idx in required.frequency_columns]
    )
    amplitude_by_bin = {
        int(bin_idx): float(value)
        for bin_idx, value in zip(bin_indices, grand_average.to_numpy(dtype=float))
        if np.isfinite(value)
    }
    column_by_bin = {
        int(bin_idx): str(column)
        for _freq, column, bin_idx in required.frequency_columns
        if int(bin_idx) in set(bin_indices)
    }
    freq_by_bin = {
        int(bin_idx): float(freq)
        for freq, _column, bin_idx in required.frequency_columns
        if int(bin_idx) in set(bin_indices)
    }
    rows: list[GroupSignificantHarmonicRow] = []
    harmonic_domain: list[float] = []
    selected_freqs: list[float] = []
    selected_columns: list[str] = []
    selected_indices: list[int] = []
    z_by_harmonic: dict[float, float] = {}
    excluded_base: list[float] = []
    seen_indices: set[int] = set()

    for matched_idx in required.candidate_indices:
        matched_freq = freq_by_bin.get(int(matched_idx))
        if matched_freq is None:
            continue
        harmonic_index = int(round(matched_freq / oddball))
        target_freq = float(harmonic_index * oddball)
        diff = abs(float(matched_freq) - target_freq)
        if diff > GROUP_SIGNIFICANT_MATCHING_TOLERANCE_HZ:
            rows.append(
                GroupSignificantHarmonicRow(
                    harmonic_index=harmonic_index,
                    target_frequency_hz=target_freq,
                    matched_frequency_hz=matched_freq,
                    matched_column=None,
                    matched_bin_index=matched_idx,
                    z_score=None,
                    selected=False,
                    excluded_base_rate=False,
                    exclusion_reason="no_full_fft_bin_within_tolerance",
                    warning=(
                        f"Nearest full-spectrum bin differs by {diff:g} Hz, "
                        f"above tolerance {GROUP_SIGNIFICANT_MATCHING_TOLERANCE_HZ:g} Hz."
                    ),
                )
            )
            continue
        if matched_idx in seen_indices:
            continue
        seen_indices.add(matched_idx)
        matched_column = column_by_bin.get(int(matched_idx), f"{matched_freq:.4f}_Hz")
        is_base_overlap = _is_base_overlap(
            matched_freq,
            base,
            GROUP_SIGNIFICANT_BASE_TOLERANCE_HZ,
        )
        if is_base_overlap:
            excluded_base.append(matched_freq)
            rows.append(
                GroupSignificantHarmonicRow(
                    harmonic_index=harmonic_index,
                    target_frequency_hz=target_freq,
                    matched_frequency_hz=matched_freq,
                    matched_column=matched_column,
                    matched_bin_index=matched_idx,
                    z_score=None,
                    selected=False,
                    excluded_base_rate=True,
                    exclusion_reason="base_rate_overlap",
                    warning="Base-rate overlap excluded from oddball summation.",
                )
            )
            continue

        noise_stats = _compute_noise_stats_for_planned_bin(
            amplitude_by_bin,
            matched_idx,
            window_size=GROUP_SIGNIFICANT_NOISE_WINDOW_BINS,
            min_bins=4,
        )
        target_amp = amplitude_by_bin.get(int(matched_idx), np.nan)
        noise_mean = noise_stats.mean_uv
        noise_std = noise_stats.std_uv
        z_score = (target_amp - noise_mean) / noise_std if noise_std > 1e-12 else np.nan
        z_value = float(z_score) if np.isfinite(z_score) else np.nan
        noise_bin_indices = noise_stats.candidate_bin_indices
        noise_used_bin_indices = noise_stats.used_bin_indices
        noise_frequencies = tuple(
            float(freq_by_bin[idx])
            for idx in noise_bin_indices
            if idx in freq_by_bin
        )
        noise_amplitudes = tuple(
            float(amplitude_by_bin[idx])
            for idx in noise_bin_indices
            if idx in amplitude_by_bin
        )
        noise_used_frequencies = tuple(
            float(freq_by_bin[idx])
            for idx in noise_used_bin_indices
            if idx in freq_by_bin
        )
        noise_used_amplitudes = tuple(
            float(amplitude_by_bin[idx])
            for idx in noise_used_bin_indices
            if idx in amplitude_by_bin
        )
        selected_frequency = float(target_freq)
        harmonic_domain.append(selected_frequency)
        z_by_harmonic[selected_frequency] = z_value
        selected = bool(np.isfinite(z_value) and z_value > settings.group_significant_z_threshold)
        if selected:
            selected_freqs.append(selected_frequency)
            selected_columns.append(f"{selected_frequency:.4f}_Hz")
            selected_indices.append(matched_idx)
        rows.append(
            GroupSignificantHarmonicRow(
                harmonic_index=harmonic_index,
                target_frequency_hz=target_freq,
                matched_frequency_hz=matched_freq,
                matched_column=matched_column,
                matched_bin_index=matched_idx,
                z_score=z_value if np.isfinite(z_value) else None,
                selected=selected,
                excluded_base_rate=False,
                exclusion_reason="" if selected else "z_below_threshold",
                warning="" if selected else "Z-score did not exceed threshold.",
                target_amplitude_uv=float(target_amp) if np.isfinite(target_amp) else None,
                noise_mean_uv=float(noise_mean) if np.isfinite(noise_mean) else None,
                noise_std_uv=float(noise_std) if np.isfinite(noise_std) else None,
                noise_bin_indices=noise_bin_indices,
                noise_frequencies_hz=noise_frequencies,
                noise_amplitudes_uv=noise_amplitudes,
                noise_used_bin_indices=noise_used_bin_indices,
                noise_used_frequencies_hz=noise_used_frequencies,
                noise_used_amplitudes_uv=noise_used_amplitudes,
            )
        )

    if not selected_freqs:
        candidate_summary = _format_candidate_z_summary(rows)
        log_func(
            "[PERF] Group harmonic selection found no significant harmonics. "
            f"Threshold z>{settings.group_significant_z_threshold:g}; "
            f"tested candidates: {candidate_summary}."
        )
        _log_candidate_diagnostics(rows, log_func)
        logger.warning(
            "stats_group_harmonics_no_selection",
            extra={
                "z_threshold": settings.group_significant_z_threshold,
                "candidate_summary": candidate_summary,
                "candidate_rows": [
                    {
                        "harmonic_index": row.harmonic_index,
                        "target_frequency_hz": row.target_frequency_hz,
                        "matched_frequency_hz": row.matched_frequency_hz,
                        "z_score": row.z_score,
                        "excluded_base_rate": row.excluded_base_rate,
                        "exclusion_reason": row.exclusion_reason,
                    }
                    for row in rows
                ],
            },
        )
        raise RuntimeError(
            "Group-level significant harmonic selection found no oddball harmonics "
            f"above z>{settings.group_significant_z_threshold:g}. "
            f"Tested candidates: {candidate_summary}. "
            "Use the fixed/predefined policy or inspect the regenerated full-spectrum workbooks."
        )
    elapsed = perf_counter() - started
    log_func(
        "[PERF] Group harmonic selection finished: "
        f"{spectra_count} spectra, {electrode_count} electrodes/spectrum max, "
        f"{len(selected_freqs)} selected harmonics in {elapsed:.2f}s."
    )
    logger.info(
        "stats_group_harmonics_selection_done",
        extra={
            "elapsed_s": elapsed,
            "spectra_count": spectra_count,
            "selected_harmonics": selected_freqs,
        },
    )

    selection = GroupSignificantHarmonicSelection(
        harmonic_domain_hz=harmonic_domain,
        selected_harmonics_hz=selected_freqs,
        selected_columns=selected_columns,
        selected_bin_indices=selected_indices,
        z_by_harmonic=z_by_harmonic,
        excluded_base_harmonics_hz=excluded_base,
        oddball_frequency_hz=float(oddball),
        base_frequency_hz=base,
        z_threshold=float(settings.group_significant_z_threshold),
        electrode_scope=str(settings.group_significant_electrode_scope),
        selection_scope="group_level_all_scalp_electrodes_all_selected_conditions",
        selection_conditions=list(conditions),
        selection_subjects=list(subjects),
        selection_spectra_count=int(spectra_count),
        selection_electrode_count=int(electrode_count),
        frequency_resolution_hz=frequency_resolution,
        base_overlap_tolerance_hz=GROUP_SIGNIFICANT_BASE_TOLERANCE_HZ,
        matching_tolerance_hz=GROUP_SIGNIFICANT_MATCHING_TOLERANCE_HZ,
        noise_window_bins=GROUP_SIGNIFICANT_NOISE_WINDOW_BINS,
        rows=rows,
    )
    selection = _save_project_cached_selection(cache_request, selection, log_func)
    _store_group_significant_selection(cache_key, selection)
    return selection


def _prepare_group_significant_bca_data(
    *,
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
    base_freq: float,
    log_func: Callable[[str], None],
    rois: Optional[Dict[str, List[str]]] = None,
    provenance_map: Optional[dict[tuple[str, str, str], dict[str, object]]] = None,
    settings: DVPolicySettings,
    dv_metadata: Optional[dict[str, object]] = None,
    max_freq: float | None = None,
    project_root: str | Path | None = None,
) -> Optional[Dict[str, Dict[str, Dict[str, float]]]]:
    if not subjects or not subject_data:
        log_func("No subject data. Scan folder first.")
        return None

    rois_map = rois if rois is not None else _current_rois_map()
    if not rois_map:
        log_func("No ROIs defined or available.")
        return None

    started = perf_counter()
    selection = build_group_significant_harmonic_selection(
        subjects=subjects,
        conditions=conditions,
        subject_data=subject_data,
        base_frequency_hz=base_freq,
        rois=rois_map,
        log_func=log_func,
        settings=settings,
        max_freq=max_freq,
        project_root=project_root,
    )
    log_func(
        "Group-level significant harmonics selected: "
        + ", ".join(f"{freq:g} Hz" for freq in selection.selected_harmonics_hz)
    )
    highest_meta = _highest_selected_harmonic_metadata(
        selection.selected_harmonics_hz,
        oddball_hz=selection.oddball_frequency_hz,
        prefix="highest_significant_harmonic",
    )
    highest_hz = highest_meta.get("highest_significant_harmonic_hz")
    highest_index = highest_meta.get("highest_significant_harmonic_index")
    if highest_hz is not None and np.isfinite(float(highest_hz)):
        if highest_index is not None:
            log_func(
                "Highest significant oddball harmonic: "
                f"{float(highest_hz):g} Hz (index {int(highest_index)})."
            )
        else:
            log_func(f"Highest significant oddball harmonic: {float(highest_hz):g} Hz.")

    log_func(
        "[PERF] Group harmonic selection phase complete in "
        f"{perf_counter() - started:.2f}s."
    )

    bca_started = perf_counter()
    all_subject_data: Dict[str, Dict[str, Dict[str, float]]] = {}
    bca_tasks = [
        (pid, cond_name, subject_data.get(pid, {}).get(cond_name))
        for pid in subjects
        for cond_name in conditions
    ]
    total_bca_tasks = len(bca_tasks)
    log_func(
        "[PERF] Group policy BCA aggregation started: "
        f"{total_bca_tasks} workbook reads across {len(rois_map)} ROIs."
    )
    for task_index, (pid, cond_name, file_path) in enumerate(bca_tasks, start=1):
        all_subject_data.setdefault(pid, {})
        all_subject_data[pid].setdefault(cond_name, {})
        read_started = perf_counter()
        roi_values, roi_provenance = _aggregate_bca_for_all_rois(
            file_path=file_path,
            rois=rois_map,
            log_func=log_func,
            harmonic_freqs=list(selection.selected_harmonics_hz),
            provenance_enabled=provenance_map is not None,
        )
        read_elapsed = perf_counter() - read_started
        for roi_name in rois_map.keys():
            all_subject_data[pid][cond_name][roi_name] = roi_values.get(roi_name, np.nan)
            if provenance_map is not None:
                provenance = roi_provenance.get(
                    roi_name,
                    {
                        "source_file": file_path,
                        "sheet": "BCA (uV)",
                        "row_label": None,
                        "col_label": list(selection.selected_columns),
                        "raw_cell": None,
                        "harmonic_policy": GROUP_SIGNIFICANT_POLICY_ID,
                    },
                )
                provenance["harmonic_policy"] = GROUP_SIGNIFICANT_POLICY_ID
                provenance_map[(pid, cond_name, roi_name)] = provenance
        if _should_log_progress(
            task_index,
            total_bca_tasks,
            GROUP_SIGNIFICANT_BCA_PROGRESS_INTERVAL,
        ):
            elapsed = perf_counter() - bca_started
            log_func(
                "[PERF] Group policy BCA aggregation progress: "
                f"{task_index}/{total_bca_tasks} workbooks "
                f"(participant={pid}, condition={cond_name}, "
                f"last_read={read_elapsed:.2f}s, elapsed={elapsed:.2f}s)."
            )
            logger.info(
                "stats_group_harmonics_bca_progress",
                extra={
                    "index": task_index,
                    "total": total_bca_tasks,
                    "participant": pid,
                    "condition": cond_name,
                    "last_read_s": read_elapsed,
                    "elapsed_s": elapsed,
                },
            )
    log_func(
        "[PERF] Group policy BCA aggregation finished: "
        f"{len(subjects) * len(conditions)} workbook reads for "
        f"{len(rois_map)} ROIs in {perf_counter() - bca_started:.2f}s."
    )

    if dv_metadata is not None:
        dv_metadata.update(
            settings.to_metadata(base_freq=base_freq, selected_conditions=conditions)
        )
        dv_metadata["policy_name"] = GROUP_SIGNIFICANT_POLICY_NAME
        dv_metadata["group_significant_harmonics"] = selection.to_metadata()

    total = 0
    finite = 0
    for _pid, conds in all_subject_data.items():
        for _cond, rois_dict in conds.items():
            for _roi, val in rois_dict.items():
                total += 1
                if val is not None and np.isfinite(val):
                    finite += 1
    log_func(f"[DEBUG] Summed BCA finite cells: {finite}/{total}")
    log_func(f"Summed BCA data prep complete in {perf_counter() - started:.2f}s.")
    return all_subject_data


def _build_grand_average_amplitude(
    *,
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
    rois: Dict[str, List[str]],
    electrode_scope: str,
    log_func: Callable[[str], None],
    frequency_columns: list[tuple[float, str, int]],
    required_indices: list[int],
) -> tuple[pd.Series, list[str], list[int], int, int]:
    started = perf_counter()
    spectra: list[pd.Series] = []
    columns: list[str] = []
    bin_indices: list[int] = []
    electrode_count = 0
    read_elapsed = 0.0
    max_frequency_columns_read = 0
    fft_tasks = [
        (pid, cond_name, subject_data.get(pid, {}).get(cond_name))
        for pid in subjects
        for cond_name in conditions
    ]
    total_fft_tasks = len(fft_tasks)
    log_func(
        "[PERF] FullFFT grand-average read started: "
        f"{total_fft_tasks} workbook reads; "
        f"{len(required_indices)} planned frequency columns per reference grid."
    )
    for task_index, (pid, cond_name, file_path) in enumerate(fft_tasks, start=1):
        if not file_path or not Path(file_path).exists():
            log_func(f"Missing file for {pid} {cond_name}: {file_path}")
            continue
        read_started = perf_counter()
        series, file_columns, n_electrodes = _load_mean_amplitude_series(
            file_path,
            rois=rois,
            electrode_scope=electrode_scope,
            reference_frequency_columns=frequency_columns,
            required_indices=required_indices,
        )
        file_read_elapsed = perf_counter() - read_started
        read_elapsed += file_read_elapsed
        if series.empty:
            log_func(f"No usable full-spectrum amplitude data for {pid} {cond_name}.")
            continue
        spectra.append(series)
        max_frequency_columns_read = max(max_frequency_columns_read, len(file_columns))
        if not columns:
            columns = file_columns
            bin_lookup = {column: int(idx) for _freq, column, idx in frequency_columns}
            bin_indices = [bin_lookup[column] for column in file_columns if column in bin_lookup]
        electrode_count = max(electrode_count, int(n_electrodes))
        if _should_log_progress(
            task_index,
            total_fft_tasks,
            GROUP_SIGNIFICANT_FULLFFT_PROGRESS_INTERVAL,
        ):
            elapsed = perf_counter() - started
            log_func(
                "[PERF] FullFFT grand-average read progress: "
                f"{task_index}/{total_fft_tasks} workbooks "
                f"(participant={pid}, condition={cond_name}, "
                f"columns={len(file_columns)}, electrodes={n_electrodes}, "
                f"spectra={len(spectra)}, last_read={file_read_elapsed:.2f}s, "
                f"elapsed={elapsed:.2f}s)."
            )
            logger.info(
                "stats_group_harmonics_fullfft_progress",
                extra={
                    "index": task_index,
                    "total": total_fft_tasks,
                    "participant": pid,
                    "condition": cond_name,
                    "columns": len(file_columns),
                    "electrodes": n_electrodes,
                    "spectra_count": len(spectra),
                    "last_read_s": file_read_elapsed,
                    "elapsed_s": elapsed,
                },
            )

    if not spectra:
        raise RuntimeError(
            "Group-level significant harmonic selection requires workbooks with a "
            f"'{FULL_FFT_AMPLITUDE_SHEET_NAME}' sheet for included participants and conditions."
        )
    frame = pd.concat(spectra, axis=1)
    grand_average = frame.mean(axis=1, skipna=True).sort_index()
    columns = [f"{float(freq):.4f}_Hz" for freq in grand_average.index]
    bin_lookup = {column: idx for _freq, column, idx in frequency_columns}
    bin_indices = [int(bin_lookup[column]) for column in columns if column in bin_lookup]
    elapsed = perf_counter() - started
    log_func(
        "[PERF] FullFFT grand-average read: "
        f"{len(spectra)} workbooks x up to {max_frequency_columns_read} frequency columns "
        f"in {elapsed:.2f}s (read phase {read_elapsed:.2f}s)."
    )
    logger.info(
        "stats_group_harmonics_fullfft_read_done",
        extra={
            "elapsed_s": elapsed,
            "read_elapsed_s": read_elapsed,
            "spectra_count": len(spectra),
            "frequency_columns": max_frequency_columns_read,
        },
    )
    return grand_average, columns, bin_indices, len(spectra), electrode_count


def _plan_required_full_fft_columns(
    *,
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
    base_frequency_hz: float,
    max_freq: float | None,
    log_func: Callable[[str], None],
) -> RequiredFullFftColumns:
    started = perf_counter()
    base = float(base_frequency_hz)
    oddball = _resolve_group_oddball_frequency(
        base_frequency_hz=base,
    )
    header_columns = _find_first_full_fft_columns(subjects, conditions, subject_data)
    frequency_columns = _parse_frequency_columns(header_columns)
    if not frequency_columns:
        raise RuntimeError(
            "Group-level significant harmonic selection found no frequency columns "
            f"in '{FULL_FFT_AMPLITUDE_SHEET_NAME}'."
        )

    freq_axis = np.asarray([freq for freq, _column, _idx in frequency_columns], dtype=float)
    max_limit = float(max_freq) if max_freq is not None else float(np.nanmax(freq_axis))
    highest_k = int(np.floor(max_limit / oddball))
    candidate_indices: list[int] = []
    excluded_base_indices: list[int] = []
    required_indices: set[int] = set()
    all_indices = {int(idx) for _freq, _column, idx in frequency_columns}

    for harmonic_index in range(1, highest_k + 1):
        target_freq = float(harmonic_index * oddball)
        exact_match = _find_exact_frequency_column(frequency_columns, target_freq)
        if exact_match is None:
            continue
        matched_freq, _column, matched_idx = exact_match
        candidate_indices.append(int(matched_idx))
        required_indices.add(int(matched_idx))
        if _is_base_overlap(matched_freq, base, GROUP_SIGNIFICANT_BASE_TOLERANCE_HZ):
            excluded_base_indices.append(int(matched_idx))
            continue
        for noise_idx in _noise_indices_for_bin(
            int(matched_idx),
            available_indices=all_indices,
            window_size=GROUP_SIGNIFICANT_NOISE_WINDOW_BINS,
        ):
            required_indices.add(int(noise_idx))

    if not candidate_indices:
        raise RuntimeError(
            "Group-level significant harmonic selection requires exact nominal "
            "oddball harmonic columns in the FullFFT sheet. Regenerate workbooks "
            "with FFT crop/on-bin output; fixed-epoch fallback workbooks cannot "
            "be used for this selection method."
        )

    column_by_idx = {int(idx): str(column) for _freq, column, idx in frequency_columns}
    ordered_columns = [
        column_by_idx[idx]
        for idx in sorted(required_indices)
        if idx in column_by_idx
    ]
    usecols = ["Electrode", *ordered_columns]
    log_func(
        "[PERF] FullFFT required-column plan built in "
        f"{perf_counter() - started:.2f}s: "
        f"{len(candidate_indices)} candidate oddball bins, "
        f"{len(ordered_columns)} frequency columns to read."
    )
    return RequiredFullFftColumns(
        usecols=usecols,
        frequency_columns=frequency_columns,
        candidate_indices=candidate_indices,
        excluded_base_indices=excluded_base_indices,
        required_indices=sorted(required_indices),
    )


def _resolve_group_oddball_frequency(
    *,
    base_frequency_hz: float,
) -> float:
    base = float(base_frequency_hz)
    if not np.isfinite(base) or base <= 0:
        raise RuntimeError(
            "Group-level significant harmonic selection requires a positive finite "
            f"base frequency. Received base_frequency_hz={base_frequency_hz!r}."
        )
    oddball = float(LOCKED_ODDBALL_FREQUENCY_HZ)
    if oddball >= base:
        raise RuntimeError(
            "Group-level significant harmonic selection requires the oddball "
            "frequency to be lower than the base frequency. The Stats oddball "
            "frequency is locked at 1.2 Hz; the BCA upper limit only controls "
            "where the 1.2 Hz harmonic list stops. "
            f"Received base_frequency_hz={base:g} and oddball_frequency_hz={oddball:g}. "
            "Check Settings > FPVS base frequency."
        )
    if _is_base_overlap(oddball, base, GROUP_SIGNIFICANT_BASE_TOLERANCE_HZ):
        raise RuntimeError(
            "Group-level significant harmonic selection requires an oddball "
            "frequency that is not itself a base-rate overlap. The Stats oddball "
            "frequency is locked at 1.2 Hz; the BCA upper limit only controls "
            "where the 1.2 Hz harmonic list stops. "
            f"Received base_frequency_hz={base:g} and oddball_frequency_hz={oddball:g}. "
            "Check Settings > FPVS base frequency."
        )
    return float(oddball)


def _preflight_required_full_fft_columns(
    *,
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
    required: RequiredFullFftColumns,
    log_func: Callable[[str], None],
) -> int:
    candidate_columns = _columns_for_required_indices(
        required.frequency_columns,
        required.candidate_indices,
    )
    required_columns = _columns_for_required_indices(
        required.frequency_columns,
        required.required_indices,
    )
    planned_workbooks = 0

    for pid in subjects:
        for cond_name in conditions:
            file_path = subject_data.get(pid, {}).get(cond_name)
            if not file_path or not Path(file_path).exists():
                continue
            planned_workbooks += 1
            header_columns = _read_full_fft_header(file_path)
            header_names = {
                str(column)
                for column in header_columns
                if isinstance(column, str)
            }

            missing_candidates = [
                column
                for column in candidate_columns
                if column not in header_names
            ]
            if missing_candidates:
                raise RuntimeError(
                    "Group-level significant harmonic selection requires exact "
                    "nominal oddball harmonic columns in every included "
                    "FullFFT sheet before reading amplitude or BCA data. "
                    f"Missing candidate columns in {file_path}: "
                    f"{missing_candidates[:8]}"
                )

            missing_required = [
                column
                for column in required_columns
                if column not in header_names
            ]
            if missing_required:
                raise RuntimeError(
                    "Group-level significant harmonic selection requires matching "
                    "FullFFT candidate and neighboring-noise columns in every "
                    "included workbook before reading amplitude or BCA data. "
                    f"Missing columns in {file_path}: {missing_required[:8]}"
                )

    if planned_workbooks <= 0:
        raise RuntimeError(
            "Group-level significant harmonic selection requires workbooks with a "
            f"'{FULL_FFT_AMPLITUDE_SHEET_NAME}' sheet for included participants and conditions."
        )

    log_func(
        "[PERF] FullFFT exact-column preflight passed: "
        f"{planned_workbooks} workbooks, "
        f"{len(candidate_columns)} exact oddball harmonic columns, "
        f"{len(required_columns)} total candidate/noise columns."
    )
    return planned_workbooks


def _columns_for_required_indices(
    frequency_columns: Sequence[tuple[float, str, int]],
    required_indices: Sequence[int],
) -> list[str]:
    requested = {int(idx) for idx in required_indices}
    return [
        str(column)
        for _freq, column, idx in frequency_columns
        if int(idx) in requested
    ]


def _find_first_full_fft_columns(
    subjects: List[str],
    conditions: List[str],
    subject_data: Dict[str, Dict[str, str]],
) -> list[object]:
    for pid in subjects:
        for cond_name in conditions:
            file_path = subject_data.get(pid, {}).get(cond_name)
            if not file_path or not Path(file_path).exists():
                continue
            try:
                workbook = load_workbook(
                    file_path,
                    read_only=True,
                    data_only=True,
                )
                try:
                    worksheet = workbook[FULL_FFT_AMPLITUDE_SHEET_NAME]
                    return [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
                finally:
                    workbook.close()
            except KeyError as exc:
                raise RuntimeError(
                    "Group-level significant harmonic selection requires regenerated "
                    f"workbooks with a '{FULL_FFT_AMPLITUDE_SHEET_NAME}' sheet: {file_path}"
                ) from exc
    return []


def _load_mean_amplitude_series(
    file_path: str,
    *,
    rois: Dict[str, List[str]],
    electrode_scope: str,
    reference_frequency_columns: list[tuple[float, str, int]],
    required_indices: list[int],
) -> tuple[pd.Series, list[str], int]:
    workbook = None
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook[FULL_FFT_AMPLITUDE_SHEET_NAME]
        header_columns = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    except KeyError as exc:
        if workbook is not None:
            workbook.close()
        raise RuntimeError(
            "Group-level significant harmonic selection requires regenerated "
            f"workbooks with a '{FULL_FFT_AMPLITUDE_SHEET_NAME}' sheet: {file_path}"
        ) from exc
    except StopIteration:
        if workbook is not None:
            workbook.close()
        return pd.Series(dtype=float), [], 0

    try:
        usecols, local_to_reference = _plan_workbook_full_fft_usecols_from_header(
            header_columns,
            reference_frequency_columns=reference_frequency_columns,
            required_indices=required_indices,
        )
        required_columns = [
            str(column)
            for _freq, column, idx in reference_frequency_columns
            if int(idx) in set(required_indices)
        ]
        matched_reference_columns = {
            str(reference_column)
            for reference_items in local_to_reference.values()
            for _reference_freq, reference_column in reference_items
        }
        missing_columns = [
            column
            for column in required_columns
            if column not in matched_reference_columns
        ]
        if missing_columns:
            raise RuntimeError(
                "Group-level significant harmonic selection requires matching "
                f"FullFFT columns in every included workbook. Missing columns in {file_path}: "
                f"{missing_columns[:8]}"
            )

        ordered_local_columns = [column for column in usecols[1:] if column in header_columns]
        if not ordered_local_columns:
            raise RuntimeError(
                "Group-level significant harmonic selection requires matching "
                f"FullFFT columns in every included workbook: {file_path}"
            )

        wanted_electrodes = _wanted_electrodes_for_scope(
            rois=rois,
            electrode_scope=electrode_scope,
        )
        position_by_column = {
            str(column): index
            for index, column in enumerate(header_columns)
            if isinstance(column, str)
        }
        selected_positions = {
            column: position_by_column[column]
            for column in ordered_local_columns
            if column in position_by_column
        }
        local_values = {column: [] for column in selected_positions}
        electrode_count = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            electrode = str(row[0]).upper().strip() if row[0] is not None else ""
            if not electrode:
                continue
            if wanted_electrodes is not None and electrode not in wanted_electrodes:
                continue
            electrode_count += 1
            for column, position in selected_positions.items():
                value = row[position] if position < len(row) else np.nan
                local_values[column].append(_numeric_or_nan(value))
    finally:
        if workbook is not None:
            workbook.close()

    values: dict[float, float] = {}
    reference_columns: list[str] = []
    for local_column in ordered_local_columns:
        column_values = np.asarray(local_values.get(local_column, []), dtype=float)
        finite_values = column_values[np.isfinite(column_values)]
        value = float(finite_values.mean()) if finite_values.size else np.nan
        for reference_freq, reference_column in local_to_reference.get(local_column, []):
            values[reference_freq] = value
            reference_columns.append(reference_column)

    series = pd.Series(values, dtype=float)
    return pd.to_numeric(series, errors="coerce"), reference_columns, electrode_count


def _plan_workbook_full_fft_usecols_from_header(
    header_columns: Sequence[object],
    *,
    reference_frequency_columns: list[tuple[float, str, int]],
    required_indices: list[int],
) -> tuple[list[str], dict[str, list[tuple[float, str]]]]:
    local_frequency_columns = _parse_frequency_columns(header_columns)
    if not local_frequency_columns:
        return ["Electrode"], {}

    reference_by_idx = {
        int(idx): (float(freq), str(column))
        for freq, column, idx in reference_frequency_columns
        if int(idx) in set(required_indices)
    }
    local_by_column = {str(column): float(freq) for freq, column, _idx in local_frequency_columns}
    local_to_reference: dict[str, list[tuple[float, str]]] = {}
    for required_idx in sorted(set(required_indices)):
        reference = reference_by_idx.get(int(required_idx))
        if reference is None:
            continue
        reference_freq, reference_column = reference
        local_freq = local_by_column.get(reference_column)
        if local_freq is None:
            continue
        if abs(float(local_freq) - reference_freq) > GROUP_SIGNIFICANT_MATCHING_TOLERANCE_HZ:
            continue
        local_to_reference.setdefault(str(reference_column), []).append(
            (reference_freq, reference_column)
        )

    return ["Electrode", *local_to_reference.keys()], local_to_reference


def _read_full_fft_header(file_path: str | Path) -> list[object]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[FULL_FFT_AMPLITUDE_SHEET_NAME]
        return [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    except KeyError as exc:
        raise RuntimeError(
            "Group-level significant harmonic selection requires regenerated "
            f"workbooks with a '{FULL_FFT_AMPLITUDE_SHEET_NAME}' sheet: {file_path}"
        ) from exc
    finally:
        workbook.close()


def _aggregate_bca_for_all_rois(
    *,
    file_path: str | None,
    rois: Dict[str, List[str]],
    log_func: Callable[[str], None],
    harmonic_freqs: List[float],
    provenance_enabled: bool,
) -> tuple[dict[str, float], dict[str, dict[str, object]]]:
    values = {roi_name: np.nan for roi_name in rois.keys()}
    provenance: dict[str, dict[str, object]] = {}
    if not file_path or not Path(file_path).exists():
        log_func(f"Missing file: {file_path}")
        return values, provenance

    started = perf_counter()
    try:
        df_bca = safe_read_excel(
            file_path,
            sheet_name="BCA (uV)",
            index_col="Electrode",
            use_cache=False,
        )
    except Exception as exc:  # noqa: BLE001
        log_func(f"Error reading BCA sheet for {file_path}: {exc}")
        return values, provenance

    read_elapsed = perf_counter() - started
    df_bca.index = df_bca.index.astype(str).str.upper().str.strip()
    cols_to_sum = [f"{float(freq_val):.4f}_Hz" for freq_val in harmonic_freqs]
    missing_bca_columns = [column for column in cols_to_sum if column not in df_bca.columns]
    if missing_bca_columns:
        raise RuntimeError(
            "Group-level significant harmonic summation requires exact selected "
            f"BCA harmonic columns in every included workbook. Missing columns in {file_path}: "
            f"{missing_bca_columns[:8]}"
        )

    numeric_bca = (
        df_bca[cols_to_sum]
        .apply(pd.to_numeric, errors="coerce")
        .replace([np.inf, -np.inf], np.nan)
    )
    for roi_name, roi_channels in rois.items():
        roi_chans = [
            str(ch).strip().upper()
            for ch in (roi_channels or [])
            if str(ch).strip().upper() in numeric_bca.index
        ]
        if not roi_chans:
            log_func(f"No overlapping BCA data for ROI {roi_name} in {file_path}.")
            if provenance_enabled:
                provenance[roi_name] = _empty_provenance(file_path, row_label=[], col_label=cols_to_sum)
            continue
        df_roi = numeric_bca.loc[roi_chans].dropna(how="all")
        if df_roi.empty:
            log_func(f"No data for ROI {roi_name} in {file_path}.")
            if provenance_enabled:
                provenance[roi_name] = _empty_provenance(file_path, row_label=roi_chans, col_label=cols_to_sum)
            continue
        bca_vals = df_roi.sum(axis=1, min_count=1)
        bca_vals = pd.to_numeric(bca_vals, errors="coerce").replace([np.inf, -np.inf], np.nan)
        if bca_vals.notna().any():
            out = float(bca_vals.mean(skipna=True))
            values[roi_name] = out if np.isfinite(out) else np.nan
        if provenance_enabled:
            provenance[roi_name] = {
                "source_file": file_path,
                "sheet": "BCA (uV)",
                "row_label": roi_chans,
                "col_label": cols_to_sum,
                "raw_cell": df_roi.to_dict(orient="index"),
                "harmonic_policy": GROUP_SIGNIFICANT_POLICY_ID,
            }
    logger.info(
        "stats_group_harmonics_bca_workbook_done",
        extra={
            "elapsed_s": perf_counter() - started,
            "read_elapsed_s": read_elapsed,
            "path": str(file_path),
            "roi_count": len(rois),
            "harmonic_count": len(cols_to_sum),
        },
    )
    return values, provenance


def _finite_harmonic_values(values: Sequence[object]) -> list[float]:
    finite_values: list[float] = []
    if values is None:
        return finite_values
    for value in values:
        try:
            number = float(value)
        except (TypeError, ValueError):
            continue
        if np.isfinite(number):
            finite_values.append(number)
    return finite_values


def _harmonic_index_for_frequency(frequency_hz: float, oddball_hz: float) -> int | None:
    try:
        frequency = float(frequency_hz)
        oddball = float(oddball_hz)
    except (TypeError, ValueError):
        return None
    if not (np.isfinite(frequency) and np.isfinite(oddball)) or frequency <= 0 or oddball <= 0:
        return None
    harmonic_index = int(round(frequency / oddball))
    if harmonic_index <= 0:
        return None
    if abs(float(harmonic_index * oddball) - frequency) > GROUP_SIGNIFICANT_MATCHING_TOLERANCE_HZ:
        return None
    return harmonic_index


def _highest_selected_harmonic_metadata(
    values: Sequence[object],
    *,
    oddball_hz: float,
    prefix: str,
) -> dict[str, object]:
    finite_values = _finite_harmonic_values(values)
    if not finite_values:
        return {
            f"{prefix}_hz": np.nan,
            f"{prefix}_index": None,
        }
    highest_hz = float(max(finite_values))
    return {
        f"{prefix}_hz": highest_hz,
        f"{prefix}_index": _harmonic_index_for_frequency(highest_hz, oddball_hz),
    }


def _empty_provenance(
    file_path: str | None,
    *,
    row_label: list[str],
    col_label: list[str],
) -> dict[str, object]:
    return {
        "source_file": file_path,
        "sheet": "BCA (uV)",
        "row_label": row_label,
        "col_label": col_label,
        "raw_cell": None,
        "harmonic_policy": GROUP_SIGNIFICANT_POLICY_ID,
    }


def _noise_indices_for_bin(
    target_idx: int,
    *,
    available_indices: set[int],
    window_size: int,
) -> list[int]:
    low = max(0, int(target_idx) - int(window_size))
    high = int(target_idx) + int(window_size)
    excluded = {int(target_idx) - 1, int(target_idx), int(target_idx) + 1}
    return [
        idx
        for idx in range(low, high + 1)
        if idx in available_indices and idx not in excluded
    ]


def _compute_noise_stats_for_planned_bin(
    amplitude_by_bin: dict[int, float],
    target_idx: int,
    *,
    window_size: int,
    min_bins: int,
) -> GroupSignificantNoiseStats:
    indices = _noise_indices_for_bin(
        int(target_idx),
        available_indices=set(amplitude_by_bin.keys()),
        window_size=window_size,
    )
    if len(indices) < min_bins:
        return GroupSignificantNoiseStats(
            mean_uv=0.0,
            std_uv=0.0,
            candidate_bin_indices=tuple(indices),
            used_bin_indices=(),
        )
    finite_pairs = [
        (idx, amplitude_by_bin[idx])
        for idx in indices
        if np.isfinite(amplitude_by_bin.get(idx, np.nan))
    ]
    if len(finite_pairs) < min_bins:
        return GroupSignificantNoiseStats(
            mean_uv=0.0,
            std_uv=0.0,
            candidate_bin_indices=tuple(idx for idx, _value in finite_pairs),
            used_bin_indices=(),
        )
    used_pairs = list(finite_pairs)
    if len(used_pairs) > 2:
        noise_vals = np.asarray([value for _idx, value in used_pairs], dtype=float)
        max_idx = int(noise_vals.argmax())
        min_idx = int(noise_vals.argmin())
        used_pairs = [
            pair
            for pair_index, pair in enumerate(used_pairs)
            if pair_index not in {max_idx, min_idx}
        ]
    if not used_pairs:
        return GroupSignificantNoiseStats(
            mean_uv=0.0,
            std_uv=0.0,
            candidate_bin_indices=tuple(idx for idx, _value in finite_pairs),
            used_bin_indices=(),
        )
    used_values = np.asarray([value for _idx, value in used_pairs], dtype=float)
    return GroupSignificantNoiseStats(
        mean_uv=float(used_values.mean()),
        std_uv=float(used_values.std(ddof=0)),
        candidate_bin_indices=tuple(idx for idx, _value in finite_pairs),
        used_bin_indices=tuple(idx for idx, _value in used_pairs),
    )


def _electrodes_for_scope(
    df_fft: pd.DataFrame,
    *,
    rois: Dict[str, List[str]],
    electrode_scope: str,
) -> list[str]:
    wanted = _wanted_electrodes_for_scope(rois=rois, electrode_scope=electrode_scope)
    if wanted is not None:
        return [idx for idx in df_fft.index.astype(str) if idx in wanted]
    _ = rois
    return list(df_fft.index.astype(str))


def _wanted_electrodes_for_scope(
    *,
    rois: Dict[str, List[str]],
    electrode_scope: str,
) -> set[str] | None:
    if electrode_scope != "union_roi_electrodes":
        return None
    return {
        str(ch).strip().upper()
        for channels in (rois or {}).values()
        for ch in (channels or [])
        if str(ch).strip()
    }


def _parse_frequency_columns(columns: Sequence[object]) -> list[tuple[float, str, int]]:
    out: list[tuple[float, str, int]] = []
    freq_idx = 0
    for col_name in columns:
        if not isinstance(col_name, str) or not col_name.endswith("_Hz"):
            continue
        try:
            out.append((float(col_name[:-3]), col_name, freq_idx))
            freq_idx += 1
        except ValueError:
            continue
    return sorted(out, key=lambda item: item[0])


def _find_exact_frequency_column(
    frequency_columns: Sequence[tuple[float, str, int]],
    target_freq: float,
) -> tuple[float, str, int] | None:
    for freq, column, idx in frequency_columns:
        if abs(float(freq) - float(target_freq)) <= 1e-9:
            return float(freq), str(column), int(idx)
    return None


def _should_log_progress(index: int, total: int, interval: int) -> bool:
    if total <= 0:
        return False
    if index in {1, total}:
        return True
    return interval > 0 and index % interval == 0


def _numeric_or_nan(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return np.nan
    return number if np.isfinite(number) else np.nan


def _frequency_resolution(freqs: Sequence[float]) -> float | None:
    unique = sorted(set(float(freq) for freq in freqs))
    if len(unique) < 2:
        return None
    diffs = np.diff(np.asarray(unique, dtype=float))
    diffs = diffs[np.isfinite(diffs) & (diffs > 0)]
    if diffs.size == 0:
        return None
    return float(np.median(diffs))


def _format_candidate_z_summary(rows: Sequence[GroupSignificantHarmonicRow]) -> str:
    parts: list[str] = []
    for row in rows:
        if row.excluded_base_rate:
            parts.append(f"{row.target_frequency_hz:.4f} Hz excluded base overlap")
            continue
        if row.z_score is None:
            reason = row.exclusion_reason or "not tested"
            parts.append(f"{row.target_frequency_hz:.4f} Hz {reason}")
            continue
        parts.append(f"{row.target_frequency_hz:.4f} Hz z={row.z_score:.3f}")
    return "; ".join(parts) if parts else "none"


def _log_candidate_diagnostics(
    rows: Sequence[GroupSignificantHarmonicRow],
    log_func: Callable[[str], None],
) -> None:
    for row in rows:
        if row.excluded_base_rate:
            log_func(
                "[DEBUG] Group harmonic candidate "
                f"{row.target_frequency_hz:.4f} Hz skipped: base-rate overlap "
                f"(column={row.matched_column}, bin={row.matched_bin_index})."
            )
            continue
        if row.target_amplitude_uv is None:
            log_func(
                "[DEBUG] Group harmonic candidate "
                f"{row.target_frequency_hz:.4f} Hz not tested: "
                f"{row.exclusion_reason or row.warning or 'missing amplitude'} "
                f"(column={row.matched_column}, bin={row.matched_bin_index})."
            )
            continue
        log_func(
            "[DEBUG] Group harmonic candidate "
            f"{row.target_frequency_hz:.4f} Hz "
            f"(column={row.matched_column}, bin={row.matched_bin_index}, "
            f"matched={_format_optional_float(row.matched_frequency_hz)} Hz): "
            f"grand_avg_amp_uv={row.target_amplitude_uv:.10g}, "
            f"noise_mean_uv={_format_optional_float(row.noise_mean_uv)}, "
            f"noise_std_uv={_format_optional_float(row.noise_std_uv)}, "
            f"z={_format_optional_float(row.z_score)}, "
            f"noise_bins={_format_noise_pairs(row.noise_frequencies_hz, row.noise_amplitudes_uv)}, "
            f"noise_used_bins={_format_noise_pairs(row.noise_used_frequencies_hz, row.noise_used_amplitudes_uv)}."
        )


def _format_optional_float(value: float | None) -> str:
    if value is None or not np.isfinite(value):
        return "nan"
    return f"{float(value):.10g}"


def _format_noise_pairs(freqs: Sequence[float], amplitudes: Sequence[float]) -> str:
    if not freqs or not amplitudes:
        return "[]"
    parts = [
        f"{float(freq):.4f}:{float(amplitude):.10g}"
        for freq, amplitude in zip(freqs, amplitudes)
    ]
    return "[" + ", ".join(parts) + "]"


def _is_base_overlap(freq: float, base: float, tolerance_hz: float) -> bool:
    if base <= 0:
        return False
    multiple = round(float(freq) / float(base))
    if multiple <= 0:
        return False
    return abs(float(freq) - multiple * float(base)) < float(tolerance_hz)


def _methods_summary(selection: GroupSignificantHarmonicSelection) -> str:
    harmonics = ", ".join(f"{freq:g}" for freq in selection.selected_harmonics_hz)
    excluded = ", ".join(f"{freq:g}" for freq in selection.excluded_base_harmonics_hz)
    return (
        "Baseline-corrected amplitudes were summed across a common group-level "
        "set of significant oddball harmonics selected from the grand-averaged "
        f"raw amplitude spectrum ({harmonics} Hz). Candidate oddball harmonics "
        f"were tested against neighboring-bin noise with z>{selection.z_threshold:g}; "
        "base-rate overlaps were excluded"
        + (f" ({excluded} Hz)." if excluded else ".")
        + " The same selected harmonic list was applied to every participant, "
        "condition, and ROI. SNR values were not used as the primary dependent variable."
    )
