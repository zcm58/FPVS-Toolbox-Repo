"""Quality Check workbook export for SNR plot spectral QC."""

from __future__ import annotations

import math
from pathlib import Path
import re
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from Main_App.processing.qc_summary_export import QUALITY_CHECK_FOLDER
from Tools.Plot_Generator.spectral_qc import SpectralQcResult, SpectralQcThresholds

SPECTRAL_QC_REPORT_PREFIX = "SNR_Spectral_QC"
_ILLEGAL_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*]+')


def _safe_fragment(value: str) -> str:
    text = str(value or "").strip() or "Condition"
    text = _ILLEGAL_FILENAME_CHARS.sub("_", text)
    text = re.sub(r"\s+", "_", text).strip(" ._")
    return text or "Condition"


def resolve_quality_check_dir(
    *,
    project_root: str | None,
    input_folder: str,
    out_dir: str,
) -> Path:
    """Return the project Quality Check folder, with standalone fallbacks."""

    if project_root:
        return Path(project_root).resolve() / QUALITY_CHECK_FOLDER

    for raw_path in (input_folder, out_dir):
        path = Path(raw_path)
        if path.name in {"1 - Excel Data Files", "2 - SNR Plots"}:
            return path.resolve().parent / QUALITY_CHECK_FOLDER
    return Path(out_dir).resolve() / QUALITY_CHECK_FOLDER


def _write_rows(ws, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 54)


def write_spectral_qc_report(
    *,
    result: SpectralQcResult,
    condition: str,
    quality_check_dir: Path,
    thresholds: SpectralQcThresholds,
    oddball_freq: float,
    base_freq: float,
) -> SpectralQcResult:
    """Write a tidy spectral QC workbook and return the result with path."""

    quality_check_dir.mkdir(parents=True, exist_ok=True)
    report_path = quality_check_dir / f"{SPECTRAL_QC_REPORT_PREFIX}_{_safe_fragment(condition)}.xlsx"
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary_rows = [
        ("Condition", condition),
        ("Checked electrode-frequency cells", result.checked_cells),
        ("Flagged electrode-frequency cells", result.flagged_cells),
        ("Data mutation", "None; source Excel workbooks and raw data were not modified."),
        ("Flag behavior", "Report-only; SNR plot aggregation values are not changed."),
        ("Reason", "Off-harmonic raw FFT amplitude outlier with visible SNR impact."),
        ("Oddball frequency (Hz)", oddball_freq),
        ("Base frequency (Hz)", base_freq),
        ("SNR threshold", thresholds.snr_threshold),
        ("Minimum FFT amplitude (uV)", thresholds.min_fft_uv),
        ("Fold-over-median threshold", thresholds.fold_threshold),
        ("Robust z threshold", thresholds.robust_z_threshold),
        ("Expected-frequency tolerance (Hz)", thresholds.expected_frequency_tolerance_hz),
    ]
    _write_rows(summary, ("Field", "Value"), summary_rows)

    flags = wb.create_sheet("Flagged Electrodes")
    headers = (
        "Condition",
        "PID",
        "Electrode",
        "Frequency (Hz)",
        "Reason",
        "SNR value",
        "FFT amplitude (uV)",
        "Group median FFT (uV)",
        "Robust z",
        "Fold over median",
        "Source workbook",
    )
    flag_rows = [
        (
            record.condition,
            record.pid,
            record.electrode,
            round(record.frequency_hz, 4),
            record.reason,
            round(record.snr_value, 6),
            round(record.fft_amplitude_uv, 6),
            round(record.group_median_fft_uv, 6),
            round(record.robust_z, 6) if math.isfinite(record.robust_z) else "inf",
            round(record.fold_over_median, 6)
            if math.isfinite(record.fold_over_median)
            else "inf",
            record.source_workbook,
        )
        for record in result.records
    ]
    _write_rows(flags, headers, flag_rows)

    wb.save(report_path)
    return SpectralQcResult(
        records=result.records,
        report_path=report_path,
        checked_cells=result.checked_cells,
        flagged_cells=result.flagged_cells,
    )
