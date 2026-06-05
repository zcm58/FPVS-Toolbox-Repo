"""Workbook and audit export helpers for publication reports."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Mapping

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from Tools.Publication_Report.models import (
    CONDITION_ROLES_SHEET,
    FIGURE_MANIFEST_SHEET,
    PARTICIPANT_INCLUSION_SHEET,
    ROI_DEFINITIONS_SHEET,
    RUN_SUMMARY_SHEET,
    WARNINGS_SHEET,
)


def write_report_workbook(path: Path, frames: Mapping[str, pd.DataFrame]) -> Path:
    """Write publication report source-data workbook."""

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        for worksheet in writer.book.worksheets:
            _format_sheet(worksheet)
    return target


def write_audit_json(path: Path, payload: Mapping[str, object]) -> Path:
    """Write a stable machine-readable audit file."""

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    return target


def build_initial_frames(
    *,
    run_summary: pd.DataFrame,
    participant_inclusion: pd.DataFrame,
    condition_roles: pd.DataFrame,
    roi_definitions: pd.DataFrame,
    figure_manifest: pd.DataFrame,
    warnings: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """Return the initial workbook sheet set."""

    return {
        RUN_SUMMARY_SHEET: run_summary,
        PARTICIPANT_INCLUSION_SHEET: participant_inclusion,
        CONDITION_ROLES_SHEET: condition_roles,
        ROI_DEFINITIONS_SHEET: roi_definitions,
        FIGURE_MANIFEST_SHEET: figure_manifest,
        WARNINGS_SHEET: warnings,
    }


def _format_sheet(worksheet: object) -> None:
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
    for column_index, cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12),
            70,
        )
