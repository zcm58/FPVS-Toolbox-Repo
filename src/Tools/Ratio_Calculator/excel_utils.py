from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter

from .constants import EXCEL_COL_PADDING_CHARS, EXCEL_MAX_COL_WIDTH, EXCEL_MIN_COL_WIDTH


def apply_excel_qol(writer: pd.ExcelWriter) -> None:
    wb = writer.book
    for _, ws in writer.sheets.items():
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)

            width = max(
                EXCEL_MIN_COL_WIDTH,
                min(EXCEL_MAX_COL_WIDTH, max_len + EXCEL_COL_PADDING_CHARS),
            )
            ws.column_dimensions[col_letter].width = width

    _ = wb
