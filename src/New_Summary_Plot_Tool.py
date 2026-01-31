#!/usr/bin/env python3
"""
ROI SUM/Z/BCA Aggregator + Ratio Reporter for FPVS Excel Outputs (single-file pipeline)

PURPOSE:
Compute participant-level ROI summed metrics for Z, SNR, and BCA across FIXED oddball harmonics
(regardless of significance), compute per-ROI Semantic/Color ratios for all three metrics, generate
professional summary tables, and generate plots with manual-exclusion denotation.

PIPELINE STEPS:
1. Batch read Excel outputs from two condition folders (A and B).
2. Pair participants present in BOTH folders.
3. Define FIXED oddball harmonic list:
   - Oddball base: 1.2 Hz; include all harmonics up to and including 16.8 Hz
   - Exclude {6, 12, 18, 24} Hz (base-rate frequency and harmonics), even if above limit
4. For each participant x ROI x condition:
   - Compute ROI-mean at each included harmonic for Z, SNR, and BCA
   - SUM across harmonics -> sum_Z, sum_SNR, sum_BCA
5. Manual exclusion only:
   - Participants listed in MANUAL_EXCLUDE are excluded from group mean/median/SD/SEM calculations
   - They are still included in per-participant tables and shown in plots with distinct styling
6. Ratios (per participant x ROI):
   - ratio_Z   = sum_Z_A   / sum_Z_B
   - ratio_SNR = sum_SNR_A / sum_SNR_B
   - ratio_BCA = sum_BCA_A / sum_BCA_B   (signed; negatives allowed)
   - If denominator is 0 or NaN -> ratio = NaN
7. Outputs:
   - Excel workbook with participant sums, ratios, and group summaries (USED vs ALL)
   - Log file with parameters and key counts
   - Plots (PDF + PNG):
       RAW:   SUM_Z, SUM_SNR, SUM_BCA (A vs B), ROI-colored, manual-excluded shown distinct
       RATIO: ratio_Z, ratio_SNR, ratio_BCA (A/B), ROI-colored, manual-excluded shown distinct,
              reference line at y=1.0 for easy interpretation
8. Completion dialog (PySide6) offering to open the output folder (Windows).

This is a personal script intended to run on a single PC; hard-coded paths are OK.
"""

from __future__ import annotations

import os
import math
import re
import datetime as dt
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Set

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle

from PySide6.QtWidgets import QApplication, QMessageBox

from openpyxl.utils import get_column_letter


# =================================================================
# 1. USER INPUTS (EDIT THESE)
# =================================================================

INPUT_DIR_A = r"C:\Users\zackm\OneDrive - Mississippi State University\NERD\2 - Results\1 - FPVS Toolbox Projects\Semantic Categories\1 - Excel Data Files\Green Fruit vs Green Veg"
CONDITION_LABEL_A = "Semantic Response"

INPUT_DIR_B = r"C:\Users\zackm\OneDrive - Mississippi State University\NERD\2 - Results\1 - FPVS Toolbox Projects\Semantic Categories\1 - Excel Data Files\Green Fruit vs Red Fruit"
CONDITION_LABEL_B = "Color Response"

OUTPUT_DIR = r"C:\Users\zackm\OneDrive - Mississippi State University\NERD\2 - Results\1 - FPVS Toolbox Projects\Semantic Categories\Ratio Summaries"
RUN_LABEL = "R21_Preliminary_Data_Run"

ROI_DEFS = {
    "Occipital": ["O1", "O2", "Oz"],
    "LOT": ["P7", "P9", "PO7", "PO3", "O1"],
}

# Oddball harmonic summation definition
ODDBALL_BASE_HZ = 1.2
SUM_UP_TO_HZ = 16.8
EXCLUDED_FREQS_HZ = {6.0, 12.0, 18.0, 24.0}

PALETTE_CHOICE = "vibrant"  # "vibrant", "muted", "colorblind_safe"
PNG_DPI = 300

# Manual Kill List (excluded from group summaries, but still shown/recorded)
"""
Use the list below to manually exclude outliers from group summary calculations.
Format must be "P##" (e.g., "P17").
"""
MANUAL_EXCLUDE = ["P17", "P20"]

# Plot denotation for manual exclusions
MANUAL_EXCLUDED_POINT_COLOR = "#4D4D4D"  # gray
MANUAL_EXCLUDED_POINT_MARKER = "x"

# Stable y-limits (set per metric; set to None for auto)
# RAW SUM plots
YLIM_RAW_SUM_Z: Optional[Tuple[float, float]] = None
YLIM_RAW_SUM_SNR: Optional[Tuple[float, float]] = None
YLIM_RAW_SUM_BCA: Optional[Tuple[float, float]] = None

# RATIO plots
YLIM_RATIO_Z: Optional[Tuple[float, float]] = None
YLIM_RATIO_SNR: Optional[Tuple[float, float]] = None
YLIM_RATIO_BCA: Optional[Tuple[float, float]] = None


# =================================================================
# 2. INTERNAL CONFIGURATION
# =================================================================

SHEET_SNR = "SNR"
SHEET_Z = "Z Score"
SHEET_BCA = "BCA (uV)"
ELECTRODE_COL = "Electrode"

PALETTES = {
    "vibrant": {"Occipital": "#2E86FF", "LOT": "#FF6B2E", "Default": "#7F8C8D"},
    "muted": {"Occipital": "#4C78A8", "LOT": "#F58518", "Default": "#95A5A6"},
    "colorblind_safe": {"Occipital": "#0072B2", "LOT": "#D55E00", "Default": "#CC79A7"},
}

PID_RE = re.compile(r"(P)(\d+)", re.IGNORECASE)

# CV not used in this simplified output, but keep EPS for safe checks
EPS = 1e-12

# Excel QoL formatting
EXCEL_COL_PADDING_CHARS = 2
EXCEL_MIN_COL_WIDTH = 8
EXCEL_MAX_COL_WIDTH = 70


# -----------------------------
# Utility / Parsing
# -----------------------------

def parse_participant_id(filename: str) -> Tuple[str, int]:
    m = PID_RE.search(filename)
    if not m:
        raise ValueError(f"Could not parse participant id from: {filename}")
    return f"P{m.group(2)}", int(m.group(2))


def harmonic_col_to_hz(col: str) -> float:
    s = str(col).strip()
    s = s.replace("HZ", "Hz").replace("hz", "Hz")
    s = s.replace("_Hz", "").replace("Hz", "")
    return float(s)


def _hz_key(hz: float) -> float:
    # stable rounding for dictionary keys
    return float(round(float(hz), 6))


def fmt_hz_list(hz_list: List[float]) -> str:
    if not hz_list:
        return ""
    return ", ".join([f"{float(hz):g}" for hz in hz_list])


def fmt_float(x: Any, ndigits: int = 4) -> str:
    try:
        v = float(x)
    except Exception:
        return "nan"
    return f"{v:.{ndigits}g}" if np.isfinite(v) else "nan"


def safe_sum(x: np.ndarray) -> float:
    n = int(np.sum(~np.isnan(x)))
    return float(np.nansum(x)) if n > 0 else float("nan")


def safe_mean(x: np.ndarray) -> float:
    n = int(np.sum(~np.isnan(x)))
    return float(np.nanmean(x)) if n > 0 else float("nan")


def safe_median(x: np.ndarray) -> float:
    n = int(np.sum(~np.isnan(x)))
    return float(np.nanmedian(x)) if n > 0 else float("nan")


def safe_sd(x: np.ndarray) -> float:
    n = int(np.sum(~np.isnan(x)))
    return float(np.nanstd(x, ddof=1)) if n >= 2 else float("nan")


def safe_sem(x: np.ndarray) -> float:
    n = int(np.sum(~np.isnan(x)))
    return float(np.nanstd(x, ddof=1) / math.sqrt(n)) if n >= 2 else float("nan")


def validate_manual_exclude(pids: List[str]) -> None:
    for pid in pids:
        if not isinstance(pid, str) or not re.fullmatch(r"P\d+", pid.strip()):
            raise ValueError(
                f"MANUAL_EXCLUDE contains invalid PID '{pid}'. "
                f"Require format like 'P17'."
            )


def expected_oddball_harmonics(
    oddball_base_hz: float,
    up_to_hz: float,
    excluded_hz: Set[float],
    tol: float = 1e-9,
) -> List[float]:
    """
    Returns [oddball_base_hz * k] for k>=1 up to <= up_to_hz, excluding any hz in excluded_hz.
    Keys are rounded for stable matching.
    """
    if oddball_base_hz <= 0:
        raise ValueError("ODDBALL_BASE_HZ must be > 0")
    n = int(math.floor((up_to_hz / oddball_base_hz) + tol))
    out: List[float] = []
    for k in range(1, n + 1):
        hz = oddball_base_hz * k
        if hz > up_to_hz + tol:
            continue
        if any(abs(hz - ex) <= 1e-6 for ex in excluded_hz):
            continue
        out.append(_hz_key(hz))
    return out


def build_hz_to_col_map(harm_cols: List[str]) -> Dict[float, str]:
    out: Dict[float, str] = {}
    for c in harm_cols:
        try:
            hz = harmonic_col_to_hz(c)
        except Exception:
            continue
        out[_hz_key(hz)] = c
    return out


def _deterministic_hex_color(name: str) -> str:
    """
    Deterministic color generator from a string, mapped into Matplotlib's tab20 palette
    (20 distinct, colorblind-ish friendly hues). This ensures new ROIs get non-gray colors
    without manual palette edits.
    """
    # Hash -> stable int -> index in [0, 19]
    h = hashlib.md5(name.encode("utf-8")).hexdigest()
    idx = int(h[:8], 16) % 20
    cmap = plt.get_cmap("tab20")
    r, g, b, _ = cmap(idx)
    return "#{:02X}{:02X}{:02X}".format(int(r * 255), int(g * 255), int(b * 255))


def get_roi_palette(rois: List[str], palette_choice: str) -> Dict[str, str]:
    """
    Returns a ROI->color mapping.
    - Starts with the chosen base palette for known ROIs.
    - Adds deterministic colors for any ROI not explicitly defined in that palette.
    """
    base = dict(PALETTES.get(palette_choice, PALETTES["vibrant"]))
    out: Dict[str, str] = {}
    for r in rois:
        if r in base:
            out[r] = base[r]
        else:
            out[r] = _deterministic_hex_color(r)
    out["Default"] = base.get("Default", "#7F8C8D")
    return out


# -----------------------------
# Core Computation
# -----------------------------

def compute_roi_harmonic_means(
    df: pd.DataFrame,
    roi_electrodes: List[str],
    harmonic_cols: List[str],
) -> Dict[float, float]:
    roi_df = df[df[ELECTRODE_COL].isin(roi_electrodes)].copy()
    if roi_df.empty:
        raise ValueError(f"Target electrodes {roi_electrodes} not found in Excel data.")
    out: Dict[float, float] = {}
    for col in harmonic_cols:
        hz = _hz_key(harmonic_col_to_hz(col))
        vals = pd.to_numeric(roi_df[col], errors="raise").to_numpy(dtype=float)
        out[hz] = safe_mean(vals)
    return out


def read_participant_file(xlsx_path: Path) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, List[str]]:
    df_snr = pd.read_excel(xlsx_path, sheet_name=SHEET_SNR)
    df_z = pd.read_excel(xlsx_path, sheet_name=SHEET_Z)
    df_bca = pd.read_excel(xlsx_path, sheet_name=SHEET_BCA)
    harm_cols = [c for c in df_z.columns if c != ELECTRODE_COL]
    return df_snr, df_z, df_bca, harm_cols


def summarize_participant_file_sums(
    xlsx_path: Path,
    condition_label: str,
    expected_hz_keys: List[float],
) -> List[Dict[str, Any]]:
    pid, pid_num = parse_participant_id(xlsx_path.name)
    df_snr, df_z, df_bca, harm_cols = read_participant_file(xlsx_path)

    # Map expected frequencies -> actual column names (from Z sheet columns)
    hz_to_col = build_hz_to_col_map(harm_cols)

    missing = [hz for hz in expected_hz_keys if hz not in hz_to_col]
    if missing:
        raise ValueError(
            f"File '{xlsx_path.name}' is missing expected harmonic columns (Hz): "
            f"[{fmt_hz_list(missing)}]."
        )

    cols_in_order = [hz_to_col[hz] for hz in expected_hz_keys]

    rows: List[Dict[str, Any]] = []
    for roi_name, electrodes in ROI_DEFS.items():
        z_by_hz = compute_roi_harmonic_means(df_z, electrodes, cols_in_order)
        snr_by_hz = compute_roi_harmonic_means(df_snr, electrodes, cols_in_order)
        bca_by_hz = compute_roi_harmonic_means(df_bca, electrodes, cols_in_order)

        sum_z = safe_sum(np.array([z_by_hz[hz] for hz in expected_hz_keys], dtype=float))
        sum_snr = safe_sum(np.array([snr_by_hz[hz] for hz in expected_hz_keys], dtype=float))
        sum_bca = safe_sum(np.array([bca_by_hz[hz] for hz in expected_hz_keys], dtype=float))

        rows.append({
            "condition_label": condition_label,
            "participant_id": pid,
            "participant_num": pid_num,
            "ROI": roi_name,
            "n_harmonics_summed": int(len(expected_hz_keys)),
            "summed_harmonics_hz": fmt_hz_list(expected_hz_keys),
            "sum_Z": sum_z,
            "sum_SNR": sum_snr,
            "sum_BCA_uV": sum_bca,
            "source_file": str(xlsx_path),
        })

    return rows


def compute_ratio_rows_from_sums(
    df_part_all: pd.DataFrame,
    cond_a: str,
    cond_b: str,
) -> pd.DataFrame:
    """
    Builds per-participant x ROI ratios using participant sum rows from both conditions.

    Required df_part_all columns:
      participant_id, ROI, condition_label, sum_Z, sum_SNR, sum_BCA_uV
    """
    needed = {"participant_id", "ROI", "condition_label", "sum_Z", "sum_SNR", "sum_BCA_uV"}
    missing = needed - set(df_part_all.columns)
    if missing:
        raise ValueError(f"df_part_all missing columns: {sorted(list(missing))}")

    a = df_part_all[df_part_all["condition_label"] == cond_a].copy()
    b = df_part_all[df_part_all["condition_label"] == cond_b].copy()

    key_cols = ["participant_id", "ROI"]
    a = a[key_cols + ["sum_Z", "sum_SNR", "sum_BCA_uV"]].rename(columns={
        "sum_Z": "sum_Z_A",
        "sum_SNR": "sum_SNR_A",
        "sum_BCA_uV": "sum_BCA_A_uV",
    })
    b = b[key_cols + ["sum_Z", "sum_SNR", "sum_BCA_uV"]].rename(columns={
        "sum_Z": "sum_Z_B",
        "sum_SNR": "sum_SNR_B",
        "sum_BCA_uV": "sum_BCA_B_uV",
    })

    merged = pd.merge(a, b, on=key_cols, how="inner")

    def _safe_ratio(num: float, den: float) -> float:
        if not (np.isfinite(num) and np.isfinite(den)):
            return float("nan")
        if abs(den) <= EPS:
            return float("nan")
        return float(num / den)

    merged["ratio_Z"] = [
        _safe_ratio(n, d) for n, d in zip(merged["sum_Z_A"].to_numpy(), merged["sum_Z_B"].to_numpy())
    ]
    merged["ratio_SNR"] = [
        _safe_ratio(n, d) for n, d in zip(merged["sum_SNR_A"].to_numpy(), merged["sum_SNR_B"].to_numpy())
    ]
    merged["ratio_BCA"] = [
        _safe_ratio(n, d) for n, d in zip(merged["sum_BCA_A_uV"].to_numpy(), merged["sum_BCA_B_uV"].to_numpy())
    ]

    # Add notes for NaNs due to denom 0 / non-finite
    notes: List[str] = []
    for z_a, z_b, s_a, s_b, b_a, b_b in zip(
        merged["sum_Z_A"].to_numpy(),
        merged["sum_Z_B"].to_numpy(),
        merged["sum_SNR_A"].to_numpy(),
        merged["sum_SNR_B"].to_numpy(),
        merged["sum_BCA_A_uV"].to_numpy(),
        merged["sum_BCA_B_uV"].to_numpy(),
    ):
        row_notes: List[str] = []
        if not (np.isfinite(z_a) and np.isfinite(z_b)) or abs(z_b) <= EPS:
            row_notes.append("bad_ratio_Z")
        if not (np.isfinite(s_a) and np.isfinite(s_b)) or abs(s_b) <= EPS:
            row_notes.append("bad_ratio_SNR")
        if not (np.isfinite(b_a) and np.isfinite(b_b)) or abs(b_b) <= EPS:
            row_notes.append("bad_ratio_BCA")
        notes.append(";".join(row_notes) if row_notes else "")
    merged["ratio_notes"] = notes

    merged = merged.sort_values(["ROI", "participant_id"], kind="stable").reset_index(drop=True)
    return merged


def group_summary_sums(
    df_part: pd.DataFrame,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for (cond, roi), sub in df_part.groupby(["condition_label", "ROI"], sort=True):
        z = pd.to_numeric(sub["sum_Z"], errors="coerce").to_numpy(dtype=float)
        s = pd.to_numeric(sub["sum_SNR"], errors="coerce").to_numpy(dtype=float)
        b = pd.to_numeric(sub["sum_BCA_uV"], errors="coerce").to_numpy(dtype=float)
        rows.append({
            "condition_label": cond,
            "ROI": roi,
            "n_used": int(len(sub)),
            "mean_sum_Z": safe_mean(z),
            "median_sum_Z": safe_median(z),
            "sd_sum_Z": safe_sd(z),
            "sem_sum_Z": safe_sem(z),
            "mean_sum_SNR": safe_mean(s),
            "median_sum_SNR": safe_median(s),
            "sd_sum_SNR": safe_sd(s),
            "sem_sum_SNR": safe_sem(s),
            "mean_sum_BCA_uV": safe_mean(b),
            "median_sum_BCA_uV": safe_median(b),
            "sd_sum_BCA_uV": safe_sd(b),
            "sem_sum_BCA_uV": safe_sem(b),
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["condition_label", "ROI"], kind="stable").reset_index(drop=True)
    return out


def group_summary_ratios(
    df_ratio: pd.DataFrame,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for roi, sub in df_ratio.groupby("ROI", sort=True):
        rz = pd.to_numeric(sub["ratio_Z"], errors="coerce").to_numpy(dtype=float)
        rs = pd.to_numeric(sub["ratio_SNR"], errors="coerce").to_numpy(dtype=float)
        rb = pd.to_numeric(sub["ratio_BCA"], errors="coerce").to_numpy(dtype=float)
        rows.append({
            "ROI": roi,
            "n_used": int(len(sub)),
            "mean_ratio_Z": safe_mean(rz),
            "median_ratio_Z": safe_median(rz),
            "sd_ratio_Z": safe_sd(rz),
            "sem_ratio_Z": safe_sem(rz),
            "mean_ratio_SNR": safe_mean(rs),
            "median_ratio_SNR": safe_median(rs),
            "sd_ratio_SNR": safe_sd(rs),
            "sem_ratio_SNR": safe_sem(rs),
            "mean_ratio_BCA": safe_mean(rb),
            "median_ratio_BCA": safe_median(rb),
            "sd_ratio_BCA": safe_sd(rb),
            "sem_ratio_BCA": safe_sem(rb),
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["ROI"], kind="stable").reset_index(drop=True)
    return out


# -----------------------------
# Plotting
# -----------------------------

@dataclass(frozen=True)
class PlotPanel:
    val_col: str
    mean_col: str
    sem_col: str
    ylabel: str
    hline_y: Optional[float] = None
    ylim: Optional[Tuple[float, float]] = None
    title: Optional[str] = None



def _ordered_conditions(df: pd.DataFrame) -> List[str]:
    present = df["condition_label"].dropna().tolist()
    uniq: List[str] = []
    for x in present:
        if x not in uniq:
            uniq.append(x)

    ordered: List[str] = []
    for pref in [CONDITION_LABEL_A, CONDITION_LABEL_B]:
        if pref in uniq:
            ordered.append(pref)
    for x in uniq:
        if x not in ordered:
            ordered.append(x)
    return ordered


def make_raincloud_figure(
    df_part_all: pd.DataFrame,
    df_group_used: pd.DataFrame,
    panel: PlotPanel,
    out_path_no_ext: Path,
    excluded_col: str = "is_manual_excluded",
):
    """
    Raincloud plot:
      - Violin/box/scatter computed from INCLUDED participants only (excluded_col == False)
      - Manual-excluded participants are still shown as distinct scatter points
      - Group mean/SEM from df_group_used (included only)
    """
    rois = list(ROI_DEFS.keys())
    palette = get_roi_palette(rois, PALETTE_CHOICE)
    conds = _ordered_conditions(df_part_all)

    seed = abs(hash(RUN_LABEL)) % (2**32)
    rng = np.random.default_rng(seed)

    fig, ax = plt.subplots(1, 1, figsize=(10, 6))
    plt.subplots_adjust(bottom=0.18)

    n_rois = max(len(rois), 1)
    cluster_width = 0.70
    step = cluster_width / n_rois
    roi_offsets = (np.arange(n_rois) - (n_rois - 1) / 2.0) * step
    violin_width = min(0.32, step * 0.85)

    for i, cond in enumerate(conds):
        for j, roi in enumerate(rois):
            pos = float(i + roi_offsets[j])
            roi_color = palette.get(roi, palette["Default"])

            sub = df_part_all[
                (df_part_all["condition_label"] == cond) & (df_part_all["ROI"] == roi)
            ].copy()

            if sub.empty:
                continue

            sub_in = sub[sub[excluded_col] == False]  # noqa: E712
            sub_ex = sub[sub[excluded_col] == True]   # noqa: E712

            data_in = pd.to_numeric(sub_in[panel.val_col], errors="coerce").dropna().to_numpy(dtype=float)
            data_ex = pd.to_numeric(sub_ex[panel.val_col], errors="coerce").dropna().to_numpy(dtype=float)

            # Included distribution: violin + points + box
            if data_in.size > 0:
                v = ax.violinplot(data_in, positions=[pos], showextrema=False, widths=violin_width)
                for pc in v["bodies"]:
                    pc.set_facecolor(roi_color)
                    pc.set_edgecolor("none")
                    pc.set_alpha(0.30)
                    # Half-violin clip (left side)
                    clip_rect = Rectangle((-1e9, -1e9), 1e9 + pos, 2e9, transform=ax.transData)
                    pc.set_clip_path(clip_rect)

                jitter = (rng.random(len(data_in)) - 0.5) * (step * 0.30)
                x_pts = np.full(len(data_in), pos + (violin_width * 0.22)) + jitter
                ax.scatter(x_pts, data_in, color=roi_color, alpha=0.60, s=20, zorder=3)

                ax.boxplot(
                    data_in,
                    positions=[pos],
                    widths=max(0.06, violin_width * 0.22),
                    showfliers=False,
                    patch_artist=True,
                    boxprops=dict(facecolor="white", edgecolor=roi_color, linewidth=1.5),
                    whiskerprops=dict(color=roi_color, linewidth=1.5),
                    capprops=dict(color=roi_color, linewidth=1.5),
                    medianprops=dict(color="black", linewidth=2),
                )

                g = df_group_used[(df_group_used["condition_label"] == cond) & (df_group_used["ROI"] == roi)]
                if not g.empty:
                    m = float(g.iloc[0][panel.mean_col])
                    se = float(g.iloc[0][panel.sem_col])
                    if np.isfinite(m):
                        if np.isfinite(se) and se > 0:
                            ax.errorbar(
                                [pos],
                                [m],
                                yerr=[se],
                                fmt="o",
                                color=roi_color,
                                capsize=4,
                                elinewidth=2,
                                markersize=7,
                                zorder=6,
                            )
                        else:
                            ax.plot([pos], [m], marker="o", color=roi_color, markersize=7, zorder=6)

            # Manual-excluded points (shown, but not included in violin/box/means)
            if data_ex.size > 0:
                jitter_ex = (rng.random(len(data_ex)) - 0.5) * (step * 0.30)
                x_ex = np.full(len(data_ex), pos + (violin_width * 0.22)) + jitter_ex
                ax.scatter(
                    x_ex,
                    data_ex,
                    color=MANUAL_EXCLUDED_POINT_COLOR,
                    marker=MANUAL_EXCLUDED_POINT_MARKER,
                    alpha=0.85,
                    s=35,
                    zorder=7,
                )

    if panel.hline_y is not None:
        ax.axhline(panel.hline_y, color="red", linestyle=":", linewidth=2, alpha=0.8)

    if panel.ylim is not None:
        ax.set_ylim(panel.ylim[0], panel.ylim[1])

    ax.set_ylabel(panel.ylabel, fontsize=12, fontweight="bold")
    ax.grid(axis="y", linestyle="--", alpha=0.4)

    ax.set_xticks(np.arange(len(conds)))
    ax.set_xticklabels(conds, fontsize=11, fontweight="bold")

    roi_handles = [
        plt.Line2D([0], [0], marker="o", color="w",
                   markerfacecolor=palette.get(r, "#777"), markersize=10, label=r)
        for r in rois
    ]
    excl_handle = plt.Line2D(
        [0], [0],
        marker=MANUAL_EXCLUDED_POINT_MARKER,
        color=MANUAL_EXCLUDED_POINT_COLOR,
        linestyle="None",
        markersize=10,
        label="Manual excluded",
    )
    ax.legend(handles=roi_handles + [excl_handle], loc="upper right", frameon=True)

    fig.savefig(out_path_no_ext.with_suffix(".pdf"), bbox_inches="tight")
    fig.savefig(out_path_no_ext.with_suffix(".png"), dpi=PNG_DPI, bbox_inches="tight")
    plt.close(fig)


def make_raincloud_figure_roi_x(
    df_part_all: pd.DataFrame,
    df_group_used: pd.DataFrame,
    panel: PlotPanel,
    out_path_no_ext: Path,
    excluded_col: str = "is_manual_excluded",
    xlabel: str = "ROI",
):
    """
    Ratio plots with ROI on the x-axis:
      - One distribution per ROI (violin + scatter + box), computed from INCLUDED participants only
      - Manual-excluded participants are still shown as distinct scatter points
      - Group mean/SEM from df_group_used (included only)
    """
    rois = list(ROI_DEFS.keys())
    palette = get_roi_palette(rois, PALETTE_CHOICE)

    seed = abs(hash(RUN_LABEL)) % (2**32)
    rng = np.random.default_rng(seed)

    fig, ax = plt.subplots(1, 1, figsize=(10, 6))
    plt.subplots_adjust(bottom=0.25, left=0.18)


    # Per-ROI spacing settings (no condition clusters here)
    step = 1.0
    violin_width = 0.70

    for i, roi in enumerate(rois):
        pos = float(i)
        roi_color = palette.get(roi, palette["Default"])

        sub = df_part_all[df_part_all["ROI"] == roi].copy()
        if sub.empty:
            continue

        sub_in = sub[sub[excluded_col] == False]  # noqa: E712
        sub_ex = sub[sub[excluded_col] == True]   # noqa: E712

        data_in = pd.to_numeric(sub_in[panel.val_col], errors="coerce").dropna().to_numpy(dtype=float)
        data_ex = pd.to_numeric(sub_ex[panel.val_col], errors="coerce").dropna().to_numpy(dtype=float)

        if data_in.size > 0:
            v = ax.violinplot(data_in, positions=[pos], showextrema=False, widths=violin_width)
            for pc in v["bodies"]:
                pc.set_facecolor(roi_color)
                pc.set_edgecolor("none")
                pc.set_alpha(0.30)
                # Half-violin clip (left side)
                clip_rect = Rectangle((-1e9, -1e9), 1e9 + pos, 2e9, transform=ax.transData)
                pc.set_clip_path(clip_rect)

            jitter = (rng.random(len(data_in)) - 0.5) * 0.18
            x_pts = np.full(len(data_in), pos + (violin_width * 0.22)) + jitter
            ax.scatter(x_pts, data_in, color=roi_color, alpha=0.60, s=20, zorder=3)

            ax.boxplot(
                data_in,
                positions=[pos],
                widths=0.18,
                showfliers=False,
                patch_artist=True,
                boxprops=dict(facecolor="white", edgecolor=roi_color, linewidth=1.5),
                whiskerprops=dict(color=roi_color, linewidth=1.5),
                capprops=dict(color=roi_color, linewidth=1.5),
                medianprops=dict(color="black", linewidth=2),
            )

            g = df_group_used[df_group_used["ROI"] == roi]
            if not g.empty:
                m = float(g.iloc[0][panel.mean_col])
                se = float(g.iloc[0][panel.sem_col])
                if np.isfinite(m):
                    if np.isfinite(se) and se > 0:
                        ax.errorbar(
                            [pos],
                            [m],
                            yerr=[se],
                            fmt="o",
                            color=roi_color,
                            capsize=4,
                            elinewidth=2,
                            markersize=7,
                            zorder=6,
                        )
                    else:
                        ax.plot([pos], [m], marker="o", color=roi_color, markersize=7, zorder=6)

        if data_ex.size > 0:
            jitter_ex = (rng.random(len(data_ex)) - 0.5) * 0.18
            x_ex = np.full(len(data_ex), pos + (violin_width * 0.22)) + jitter_ex
            ax.scatter(
                x_ex,
                data_ex,
                color=MANUAL_EXCLUDED_POINT_COLOR,
                marker=MANUAL_EXCLUDED_POINT_MARKER,
                alpha=0.85,
                s=35,
                zorder=7,
            )

    if panel.hline_y is not None:
        ax.axhline(panel.hline_y, color="red", linestyle=":", linewidth=2, alpha=0.8)

    if panel.ylim is not None:
        ax.set_ylim(panel.ylim[0], panel.ylim[1])

    # Ratio plots: keep y-axis label short so it never clips
    ax.set_ylabel(panel.ylabel, fontsize=12, fontweight="bold")  # should be "Ratio"
    ax.set_xlabel(xlabel, fontsize=12, fontweight="bold")

    if panel.title:
        ax.set_title(panel.title, fontsize=13, fontweight="bold", pad=12)

    ax.grid(axis="y", linestyle="--", alpha=0.4)

    ax.set_xticks(np.arange(len(rois)))
    ax.set_xticklabels(rois, rotation=25, ha="right", fontsize=11, fontweight="bold")

    roi_handles = [
        plt.Line2D([0], [0], marker="o", color="w",
                   markerfacecolor=palette.get(r, "#777"), markersize=10, label=r)
        for r in rois
    ]
    excl_handle = plt.Line2D(
        [0], [0],
        marker=MANUAL_EXCLUDED_POINT_MARKER,
        color=MANUAL_EXCLUDED_POINT_COLOR,
        linestyle="None",
        markersize=10,
        label="Manual excluded",
    )
    ax.legend(handles=roi_handles + [excl_handle], loc="upper right", frameon=True)

    fig.savefig(out_path_no_ext.with_suffix(".pdf"), bbox_inches="tight")
    fig.savefig(out_path_no_ext.with_suffix(".png"), dpi=PNG_DPI, bbox_inches="tight")
    plt.close(fig)


# -----------------------------
# QOL: GUI Completion Dialog (Windows only)
# -----------------------------

def show_completion_dialog(output_path: Path):
    _app = QApplication.instance() or QApplication([])

    msg = QMessageBox()
    msg.setWindowTitle("Processing Complete")
    msg.setText("Aggregation finished successfully.")
    msg.setInformativeText("Would you like to open the output folder?")
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    msg.setIcon(QMessageBox.Information)

    if msg.exec() == QMessageBox.Yes:
        os.startfile(str(output_path))


# -----------------------------
# Excel QoL formatting
# -----------------------------

def _apply_excel_qol(writer: pd.ExcelWriter) -> None:
    """
    QoL:
      1) Auto-fit column widths with small padding.
      2) Apply auto-filters to every column on every sheet.
    """
    wb = writer.book
    for sheet_name, ws in writer.sheets.items():
        # Apply filters (header row assumed at row 1)
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        # Auto width with padding
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)

            width = max(
                EXCEL_MIN_COL_WIDTH,
                min(EXCEL_MAX_COL_WIDTH, max_len + EXCEL_COL_PADDING_CHARS),
            )
            ws.column_dimensions[col_letter].width = width

    # wb is saved when the ExcelWriter context exits
    _ = wb


# -----------------------------
# Main Execution
# -----------------------------

def main():
    validate_manual_exclude(MANUAL_EXCLUDE)

    out_dir = Path(OUTPUT_DIR).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    log_lines: List[str] = []

    def log(msg: str) -> None:
        print(msg)
        log_lines.append(msg)

    expected_hz = expected_oddball_harmonics(
        oddball_base_hz=ODDBALL_BASE_HZ,
        up_to_hz=SUM_UP_TO_HZ,
        excluded_hz=EXCLUDED_FREQS_HZ,
    )

    log("=" * 110)
    log(f"RUN_LABEL: {RUN_LABEL}")
    log(f"Timestamp: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"Output Dir:  {str(out_dir)}")
    log("-" * 110)
    log("PARAMETERS")
    log(f"  CONDITION A: {CONDITION_LABEL_A}")
    log(f"  CONDITION B: {CONDITION_LABEL_B}")
    log(f"  ODDBALL_BASE_HZ: {ODDBALL_BASE_HZ}")
    log(f"  SUM_UP_TO_HZ: {SUM_UP_TO_HZ}")
    log(f"  EXCLUDED_FREQS_HZ: {sorted(list(EXCLUDED_FREQS_HZ))}")
    log(f"  INCLUDED_ODDBALL_HARMONICS_HZ (n={len(expected_hz)}): [{fmt_hz_list(expected_hz)}]")
    log(f"  ROIs: {list(ROI_DEFS.keys())}")
    log(f"  MANUAL_EXCLUDE: {MANUAL_EXCLUDE}")
    log("=" * 110)

    def index_folder(path: str, label: str) -> Tuple[List[Path], Dict[str, Path]]:
        xlsx_files = sorted(Path(path).expanduser().glob("*.xlsx"))
        log(f"[{label}] Found {len(xlsx_files)} .xlsx files.")
        pid_to_path: Dict[str, Path] = {}
        for f in xlsx_files:
            pid, _ = parse_participant_id(f.name)
            pid_to_path[pid] = f
        return xlsx_files, pid_to_path

    _, pid_map_a = index_folder(INPUT_DIR_A, CONDITION_LABEL_A)
    _, pid_map_b = index_folder(INPUT_DIR_B, CONDITION_LABEL_B)

    pids_a = sorted(list(pid_map_a.keys()), key=lambda s: int(s[1:]) if s[1:].isdigit() else 999999)
    pids_b = sorted(list(pid_map_b.keys()), key=lambda s: int(s[1:]) if s[1:].isdigit() else 999999)
    pids_paired = sorted(list(set(pids_a).intersection(set(pids_b))),
                         key=lambda s: int(s[1:]) if s[1:].isdigit() else 999999)

    log("-" * 110)
    log(f"Participants in {CONDITION_LABEL_A}: {len(pids_a)}")
    log(f"Participants in {CONDITION_LABEL_B}: {len(pids_b)}")
    log(f"Participants paired (present in BOTH folders): {len(pids_paired)}")
    log("-" * 110)

    if not pids_paired:
        raise RuntimeError("No paired participants found between the two folders.")

    # Manual exclusion report
    manual_set = set(MANUAL_EXCLUDE)
    manual_in_paired = sorted([p for p in pids_paired if p in manual_set],
                              key=lambda s: int(s[1:]) if s[1:].isdigit() else 999999)
    manual_not_found = sorted([p for p in MANUAL_EXCLUDE if p not in set(pids_paired)],
                              key=lambda s: int(s[1:]) if s[1:].isdigit() else 999999)

    log("MANUAL EXCLUSION REPORT")
    log(f"  Manual exclude list: {MANUAL_EXCLUDE}")
    log(f"  Found among paired:  {manual_in_paired}")
    log(f"  Not found in paired: {manual_not_found}")
    log("-" * 110)

    # Participant-level sums (ALL; including manual excluded)
    part_rows: List[Dict[str, Any]] = []
    for pid in pids_paired:
        part_rows.extend(summarize_participant_file_sums(pid_map_a[pid], CONDITION_LABEL_A, expected_hz))
        part_rows.extend(summarize_participant_file_sums(pid_map_b[pid], CONDITION_LABEL_B, expected_hz))

    df_part_all = pd.DataFrame(part_rows)
    df_part_all["is_manual_excluded"] = df_part_all["participant_id"].isin(manual_set)

    # USED subset (excludes manual excluded participants)
    df_part_used = df_part_all[~df_part_all["is_manual_excluded"]].copy()

    log(f"Paired participants total: {len(pids_paired)} | used after MANUAL exclusions: {df_part_used['participant_id'].nunique()}")
    log("-" * 110)

    # Ratios (ALL; including manual excluded)
    df_ratio_all = compute_ratio_rows_from_sums(df_part_all, CONDITION_LABEL_A, CONDITION_LABEL_B)
    df_ratio_all["is_manual_excluded"] = df_ratio_all["participant_id"].isin(manual_set)

    df_ratio_used = df_ratio_all[~df_ratio_all["is_manual_excluded"]].copy()

    # Group summaries (USED only)
    df_group_sums_used = group_summary_sums(df_part_used)
    df_group_ratios_used = group_summary_ratios(df_ratio_used)

    log("GROUP SUMMARY (USED): SUMS by (condition x ROI) [mean / median / SD / SEM]")
    if df_group_sums_used.empty:
        log("  (empty)")
    else:
        log(df_group_sums_used.to_string(index=False))
    log("-" * 110)

    log("GROUP SUMMARY (USED): RATIOS by ROI [mean / median / SD / SEM]")
    if df_group_ratios_used.empty:
        log("  (empty)")
    else:
        log(df_group_ratios_used.to_string(index=False))
    log("-" * 110)

    # Plots (RAW SUMS): A vs B
    df_group_sums_plot = df_group_sums_used.rename(columns={
        "mean_sum_Z": "mean",
        "sem_sum_Z": "sem",
    }).copy()

    make_raincloud_figure(
        df_part_all.rename(columns={"sum_Z": "val"}).copy(),
        df_group_sums_plot,
        PlotPanel(val_col="val", mean_col="mean", sem_col="sem",
                  ylabel="SUM(Z) across oddball harmonics", ylim=YLIM_RAW_SUM_Z),
        out_dir / f"Plot_{RUN_LABEL}_RAW_SUM_Z",
    )

    df_group_sums_plot = df_group_sums_used.rename(columns={
        "mean_sum_SNR": "mean",
        "sem_sum_SNR": "sem",
    }).copy()

    make_raincloud_figure(
        df_part_all.rename(columns={"sum_SNR": "val"}).copy(),
        df_group_sums_plot,
        PlotPanel(val_col="val", mean_col="mean", sem_col="sem",
                  ylabel="SUM(SNR) across oddball harmonics", ylim=YLIM_RAW_SUM_SNR),
        out_dir / f"Plot_{RUN_LABEL}_RAW_SUM_SNR",
    )

    df_group_sums_plot = df_group_sums_used.rename(columns={
        "mean_sum_BCA_uV": "mean",
        "sem_sum_BCA_uV": "sem",
    }).copy()

    make_raincloud_figure(
        df_part_all.rename(columns={"sum_BCA_uV": "val"}).copy(),
        df_group_sums_plot,
        PlotPanel(val_col="val", mean_col="mean", sem_col="sem",
                  ylabel="SUM(BCA) (µV) across oddball harmonics", ylim=YLIM_RAW_SUM_BCA),
        out_dir / f"Plot_{RUN_LABEL}_RAW_SUM_BCA",
    )

    # Plots (RATIOS): ROI on x-axis, reference line at 1.0
    ratio_label = f"{CONDITION_LABEL_A} / {CONDITION_LABEL_B}"

    df_ratio_plot = df_ratio_all.copy()
    df_group_ratio_plot = df_group_ratios_used.copy()

    make_raincloud_figure_roi_x(
        df_ratio_plot.rename(columns={"ratio_Z": "val"}).copy(),
        df_group_ratio_plot.rename(columns={"mean_ratio_Z": "mean", "sem_ratio_Z": "sem"}).copy(),
        PlotPanel(
            val_col="val",
            mean_col="mean",
            sem_col="sem",
            ylabel="Ratio",
            hline_y=1.0,
            ylim=YLIM_RATIO_Z,
            title="High-Level:Low-Level Ratio using Z-Scores",
        ),

        out_dir / f"Plot_{RUN_LABEL}_RATIO_Z",
        xlabel="ROI",
    )

    make_raincloud_figure_roi_x(
        df_ratio_plot.rename(columns={"ratio_SNR": "val"}).copy(),
        df_group_ratio_plot.rename(columns={"mean_ratio_SNR": "mean", "sem_ratio_SNR": "sem"}).copy(),
        PlotPanel(
            val_col="val",
            mean_col="mean",
            sem_col="sem",
            ylabel="Ratio",
            hline_y=1.0,
            ylim=YLIM_RATIO_SNR,
            title="High-Level:Low-Level Ratio using SNR",
        ),

        out_dir / f"Plot_{RUN_LABEL}_RATIO_SNR",
        xlabel="ROI",
    )

    make_raincloud_figure_roi_x(
        df_ratio_plot.rename(columns={"ratio_BCA": "val"}).copy(),
        df_group_ratio_plot.rename(columns={"mean_ratio_BCA": "mean", "sem_ratio_BCA": "sem"}).copy(),
        PlotPanel(
            val_col="val",
            mean_col="mean",
            sem_col="sem",
            ylabel="Ratio",
            hline_y=1.0,
            ylim=YLIM_RATIO_BCA,
            title="High-Level:Low-Level Ratio using Summed BCA",
        ),

        out_dir / f"Plot_{RUN_LABEL}_RATIO_BCA",
        xlabel="ROI",
    )

    # Excel + log outputs
    out_xlsx = out_dir / f"Metrics_{RUN_LABEL}.xlsx"

    params_df = pd.DataFrame([
        {"key": "RUN_LABEL", "value": RUN_LABEL},
        {"key": "CONDITION_LABEL_A", "value": CONDITION_LABEL_A},
        {"key": "CONDITION_LABEL_B", "value": CONDITION_LABEL_B},
        {"key": "ODDBALL_BASE_HZ", "value": ODDBALL_BASE_HZ},
        {"key": "SUM_UP_TO_HZ", "value": SUM_UP_TO_HZ},
        {"key": "EXCLUDED_FREQS_HZ", "value": str(sorted(list(EXCLUDED_FREQS_HZ)))},
        {"key": "INCLUDED_ODDBALL_HARMONICS_HZ", "value": fmt_hz_list(expected_hz)},
        {"key": "N_INCLUDED_HARMONICS", "value": len(expected_hz)},
        {"key": "MANUAL_EXCLUDE", "value": str(MANUAL_EXCLUDE)},
        {"key": "MANUAL_FOUND_IN_PAIRED", "value": str(manual_in_paired)},
        {"key": "MANUAL_NOT_FOUND_IN_PAIRED", "value": str(manual_not_found)},
    ])

    manual_excl_df = pd.DataFrame([
        {"participant_id": pid, "found_in_paired": (pid in set(pids_paired))}
        for pid in MANUAL_EXCLUDE
    ])

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        params_df.to_excel(writer, sheet_name="Parameters", index=False)
        manual_excl_df.to_excel(writer, sheet_name="Manual_Exclusions", index=False)

        df_part_all.to_excel(writer, sheet_name="Participant_Sums_ALL", index=False)
        df_part_used.to_excel(writer, sheet_name="Participant_Sums_USED", index=False)

        df_ratio_all.to_excel(writer, sheet_name="Ratios_ALL", index=False)
        df_ratio_used.to_excel(writer, sheet_name="Ratios_USED", index=False)

        df_group_sums_used.to_excel(writer, sheet_name="Group_Sums_USED", index=False)
        df_group_ratios_used.to_excel(writer, sheet_name="Group_Ratios_USED", index=False)

        _apply_excel_qol(writer)

    out_log = out_dir / f"Log_{RUN_LABEL}.txt"
    out_log.write_text("\n".join(log_lines) + "\n", encoding="utf-8")

    log("=" * 110)
    log(f"Success! Outputs saved to: {out_dir}")
    log(f"Excel: {out_xlsx.name}")
    log(f"Log:   {out_log.name}")
    log("=" * 110)

    show_completion_dialog(out_dir)


if __name__ == "__main__":
    main()
