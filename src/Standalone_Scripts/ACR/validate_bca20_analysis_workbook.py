"""Read-only structural and numerical QA for the ACR BCA20 expert workbook."""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
import json
from pathlib import Path
import re
from typing import Any, Sequence

import numpy as np
import pandas as pd
from openpyxl import load_workbook

if __package__ in {None, ""}:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from Standalone_Scripts.ACR.bca20_common import (  # noqa: E402
        audit_configured_roi_input,
        read_configured_roi_input,
        sha256_file,
        write_json,
    )
else:
    from .bca20_common import (
        audit_configured_roi_input,
        read_configured_roi_input,
        sha256_file,
        write_json,
    )


EXPECTED_SHEETS = (
    "Read_Me",
    "ROI_Long",
    "Normalization",
    "Raw_Wide",
    "RMS_Wide",
    "SignedMean_Wide",
    "Participants",
    "Condition_Coverage",
    "Cell_Coverage",
    "ROI_Definitions",
    "Harmonics",
    "Exclusions",
)
ROIS = ("LOT", "ROT", "O", "Frontal", "PO", "CP")
ROI_DISPLAY_NAMES = {
    "LOT": "LOT",
    "ROT": "ROT",
    "O": "Occipital",
    "Frontal": "Frontal",
    "PO": "Parieto-Occipital",
    "CP": "Centro-Parietal",
}
REMOVED_PUBLIC_HEADERS = {
    "participant number",
    "group id",
    "group label",
    "cohort id",
    "source workbook relative path",
    "source workbook sha256",
    "source workbook size bytes",
    "roi role",
}
EXPECTED_ROI_LONG_HEADERS = (
    "PID",
    "Group",
    "Condition",
    "ROI",
    "ROI Electrodes",
    "Raw Summed BCA",
    "RMS Normalized BCA",
    "Signed Mean Normalized BCA",
)
EXPECTED_NORMALIZATION_HEADERS = (
    "PID",
    "Group",
    "Condition",
    "Whole-Scalp RMS BCA Denominator",
    "Whole-Scalp Signed Mean BCA Denominator",
    "Signed Mean Stability Q",
    "Signed Mean Stable (Q >= 0.05)",
)
NORMALIZATION_REFERENCE_HEADERS = frozenset(EXPECTED_NORMALIZATION_HEADERS[3:])


def _normalized_header(value: object) -> str:
    return re.sub(r"[_\s]+", " ", str(value).strip().casefold())


def _wide_matches(
    workbook: Path,
    long_data: pd.DataFrame,
    *,
    sheet_name: str,
    value_column: str,
    metric_label: str,
) -> bool:
    wide = pd.read_excel(workbook, sheet_name=sheet_name)
    expected = long_data.pivot(index=["subject", "condition"], columns="roi", values=value_column)
    actual = wide.set_index(["PID", "Condition"])
    if set(expected.index) != set(actual.index):
        return False
    actual = actual.reindex(expected.index)
    for roi in ROIS:
        column = f"{ROI_DISPLAY_NAMES[roi]} {metric_label}"
        if column not in actual:
            return False
        if not np.isclose(
            expected[roi].to_numpy(dtype=float),
            actual[column].to_numpy(dtype=float),
            atol=1e-12,
            rtol=1e-10,
            equal_nan=True,
        ).all():
            return False
    return True


def _normalization_matches(workbook: Path, long_data: pd.DataFrame) -> bool:
    """Reconcile one normalization-reference row per observed PID-condition."""

    reference_columns = (
        "subject",
        "group",
        "condition",
        "global_mean",
        "global_rms",
        "mean_abs_over_rms",
    )
    if not set(reference_columns).issubset(long_data.columns):
        return False

    references = long_data.loc[:, reference_columns].copy()
    grouped = references.groupby(["subject", "condition"], observed=True)
    if not all(
        (grouped[column].nunique(dropna=False) == 1).all()
        for column in ("group", "global_mean", "global_rms", "mean_abs_over_rms")
    ):
        return False
    expected = references.drop_duplicates(["subject", "condition"]).set_index(
        ["subject", "condition"]
    )

    actual = pd.read_excel(workbook, sheet_name="Normalization")
    if tuple(str(column).strip() for column in actual.columns) != EXPECTED_NORMALIZATION_HEADERS:
        return False
    actual = actual.rename(
        columns={
            "PID": "subject",
            "Group": "group",
            "Condition": "condition",
            "Whole-Scalp RMS BCA Denominator": "global_rms",
            "Whole-Scalp Signed Mean BCA Denominator": "global_mean",
            "Signed Mean Stability Q": "mean_abs_over_rms",
            "Signed Mean Stable (Q >= 0.05)": "stable",
        }
    )
    if actual.duplicated(["subject", "condition"]).any():
        return False
    actual = actual.set_index(["subject", "condition"])
    if set(expected.index) != set(actual.index):
        return False
    actual = actual.reindex(expected.index)
    actual_groups = (
        actual["group"]
        .astype(str)
        .str.strip()
        .str.casefold()
        .str.replace("-", "_", regex=False)
    )
    if not expected["group"].astype(str).equals(actual_groups):
        return False
    for column in ("global_mean", "global_rms", "mean_abs_over_rms"):
        if not np.isclose(
            expected[column].to_numpy(dtype=float),
            actual[column].to_numpy(dtype=float),
            atol=1e-12,
            rtol=1e-10,
            equal_nan=True,
        ).all():
            return False
    expected_stable = np.where(expected["mean_abs_over_rms"] >= 0.05, "yes", "no")
    actual_stable = actual["stable"].astype(str).str.strip().str.casefold().to_numpy()
    return bool(np.array_equal(expected_stable, actual_stable))


def validate_workbook(path: str | Path) -> dict[str, Any]:
    workbook_path = Path(path).expanduser().resolve(strict=True)
    long_data, _ = read_configured_roi_input(workbook_path)
    provenance = audit_configured_roi_input(workbook_path, row_count=len(long_data))
    duplicate_keys = int(long_data.duplicated(["subject", "condition", "roi"]).sum())
    cells = long_data[["subject", "condition"]].drop_duplicates()
    rois_per_cell = long_data.groupby(["subject", "condition"], observed=True)["roi"].nunique()

    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    sheet_names_match = tuple(workbook.sheetnames) == EXPECTED_SHEETS
    formula_cells = 0
    error_cells = 0
    table_counts: dict[str, int] = {}
    public_headers: set[str] = set()
    headers_by_sheet: dict[str, tuple[str, ...]] = {}
    public_group_values: set[str] = set()
    for worksheet in workbook.worksheets:
        table_counts[worksheet.title] = len(worksheet.tables)
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in worksheet[1]
        ]
        headers_by_sheet[worksheet.title] = tuple(headers)
        public_headers.update(header for header in headers if header)
        normalized_headers = [_normalized_header(header) for header in headers]
        if "group" in normalized_headers:
            group_column = normalized_headers.index("group") + 1
            public_group_values.update(
                str(worksheet.cell(row=row, column=group_column).value).strip()
                for row in range(2, worksheet.max_row + 1)
                if worksheet.cell(row=row, column=group_column).value is not None
            )
        for row in worksheet.iter_rows():
            for cell in row:
                formula_cells += int(cell.data_type == "f")
                error_cells += int(cell.data_type == "e")
    workbook.close()

    group_counts = (
        long_data[["subject", "group"]]
        .drop_duplicates()
        .groupby("group", observed=True)["subject"]
        .nunique()
        .astype(int)
        .to_dict()
    )
    checks = {
        "sheet_names_and_order": sheet_names_match,
        "one_table_per_sheet": all(count == 1 for count in table_counts.values()),
        "no_formula_cells": formula_cells == 0,
        "no_error_cells": error_cells == 0,
        "no_duplicate_long_keys": duplicate_keys == 0,
        "six_rois_per_observed_cell": bool((rois_per_cell == len(ROIS)).all()),
        "removed_columns_absent": not REMOVED_PUBLIC_HEADERS.intersection(
            {_normalized_header(header) for header in public_headers}
        ),
        "no_underscore_column_headers": all(
            "_" not in header for header in public_headers
        ),
        "roi_long_headers_match_clean_contract": headers_by_sheet.get("ROI_Long")
        == EXPECTED_ROI_LONG_HEADERS,
        "normalization_headers_match_clean_contract": headers_by_sheet.get(
            "Normalization"
        )
        == EXPECTED_NORMALIZATION_HEADERS,
        "reference_and_diagnostic_columns_absent_from_roi_long": not (
            NORMALIZATION_REFERENCE_HEADERS
            & set(headers_by_sheet.get("ROI_Long", ()))
        ),
        "signed_mean_wide_diagnostics_absent": not (
            {"Signed Mean Stability Q", "Signed Mean Stable (Q >= 0.05)"}
            & set(headers_by_sheet.get("SignedMean_Wide", ()))
        ),
        "normalization_references_reconcile": _normalization_matches(
            workbook_path,
            long_data,
        ),
        "public_group_values_are_canonical": public_group_values
        == {"anxious", "non-anxious"},
        "raw_wide_reconciles": _wide_matches(
            workbook_path,
            long_data,
            sheet_name="Raw_Wide",
            value_column="raw",
            metric_label="Raw Summed BCA",
        ),
        "rms_wide_reconciles": _wide_matches(
            workbook_path,
            long_data,
            sheet_name="RMS_Wide",
            value_column="rms_norm",
            metric_label="RMS Normalized BCA",
        ),
        "signed_mean_wide_reconciles": _wide_matches(
            workbook_path,
            long_data,
            sheet_name="SignedMean_Wide",
            value_column="mean_norm",
            metric_label="Signed Mean Normalized BCA",
        ),
    }
    return {
        "schema_version": 1,
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "workbook": {
            "path": str(workbook_path),
            "sha256": sha256_file(workbook_path),
            "size_bytes": int(workbook_path.stat().st_size),
        },
        "passed": bool(all(checks.values())),
        "checks": checks,
        "counts": {
            "sheets": len(workbook.sheetnames),
            "roi_long_rows": int(len(long_data)),
            "participant_condition_cells": int(len(cells)),
            "participants": int(long_data["subject"].nunique()),
            "groups": group_counts,
            "public_group_values": sorted(public_group_values),
            "conditions": int(long_data["condition"].nunique()),
            "rois": int(long_data["roi"].nunique()),
            "formula_cells": formula_cells,
            "error_cells": error_cells,
        },
        "tables_per_sheet": table_counts,
        "provenance_audit": provenance,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    return parser


def main(argv: Sequence[str] | None = None) -> None:
    args = build_parser().parse_args(argv)
    result = validate_workbook(args.workbook)
    if args.output is not None:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        write_json(args.output, result)
    print(json.dumps({"ok": result["passed"], "workbook": result["workbook"]["path"]}))
    if not result["passed"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
