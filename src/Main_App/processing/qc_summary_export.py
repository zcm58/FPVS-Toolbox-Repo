"""Participant-level processing QC summary workbook export."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from Main_App.processing.processing_ledger import (
    MISSING_EXPECTED_OUTPUTS_WARNING,
    ProcessingInputState,
    ProcessingPlan,
    load_ledger,
)

QC_SUMMARY_FILENAME = "Processing_QC_Summary.xlsx"
QC_SUMMARY_SHEET = "Participant QC"
QUALITY_CHECK_FOLDER = "Quality Check"
QC_SUMMARY_HEADERS = (
    "PID",
    "Manually Removed Electrodes",
    "Auto-Detected Removed Electrodes (Low SD)",
    "Flagged Removed-Electrode Candidates (High Amplitude)",
    "Flagged Removed-Electrode Candidates (Spatial Consistency)",
    "Kurtosis-Rejected Electrodes",
    "Electrodes Interpolated",
    "Total Number of Electrodes removed/rejected",
    "Raw QC Warnings",
    "Missing Conditions",
    "Included in Final Set",
    "Exclusion Reason",
)


def _quality_check_root(project: Any) -> Path:
    return Path(project.project_root) / QUALITY_CHECK_FOLDER


def _string_list(value: Any) -> list[str]:
    if isinstance(value, str):
        text = value.strip()
        return [text] if text else []
    if isinstance(value, Sequence):
        return [str(item) for item in value if str(item).strip()]
    return []


def _int_or_default(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def _result_by_path(results: Sequence[Mapping[str, Any]]) -> dict[Path, Mapping[str, Any]]:
    by_path: dict[Path, Mapping[str, Any]] = {}
    for result in results:
        raw_path_value = result.get("file")
        if raw_path_value:
            by_path[Path(str(raw_path_value)).resolve()] = result
    return by_path


def _channels_from_result(result: Mapping[str, Any] | None) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    interpolated = _string_list(audit.get("interpolated_channels"))
    return interpolated


def _raw_qc_channels_from_result(result: Mapping[str, Any] | None) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    raw_qc = result.get("raw_channel_qc") if isinstance(result.get("raw_channel_qc"), Mapping) else {}
    return (
        _string_list(audit.get("raw_qc_bad_channels"))
        or _string_list(raw_qc.get("bad_channels"))
    )


def _raw_qc_low_variance_channels_from_result(
    result: Mapping[str, Any] | None,
) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    raw_qc = result.get("raw_channel_qc") if isinstance(result.get("raw_channel_qc"), Mapping) else {}
    return (
        _string_list(audit.get("raw_qc_low_variance_channels"))
        or _string_list(raw_qc.get("low_variance_channels"))
    )


def _raw_qc_manual_removed_channels_from_result(
    result: Mapping[str, Any] | None,
) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    raw_qc = result.get("raw_channel_qc") if isinstance(result.get("raw_channel_qc"), Mapping) else {}
    return (
        _string_list(audit.get("raw_qc_manual_removed_channels"))
        or _string_list(raw_qc.get("manual_removed_channels"))
    )


def _raw_qc_spatial_outlier_channels_from_result(
    result: Mapping[str, Any] | None,
) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    raw_qc = result.get("raw_channel_qc") if isinstance(result.get("raw_channel_qc"), Mapping) else {}
    return (
        _string_list(audit.get("raw_qc_spatial_outlier_channels"))
        or _string_list(raw_qc.get("spatial_outlier_channels"))
    )


def _raw_qc_high_amplitude_channels_from_result(
    result: Mapping[str, Any] | None,
) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    raw_qc = result.get("raw_channel_qc") if isinstance(result.get("raw_channel_qc"), Mapping) else {}
    return (
        _string_list(audit.get("raw_qc_high_amplitude_channels"))
        or _string_list(raw_qc.get("high_amplitude_channels"))
    )


def _raw_qc_warning_rules_from_result(result: Mapping[str, Any] | None) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    raw_qc = result.get("raw_channel_qc") if isinstance(result.get("raw_channel_qc"), Mapping) else {}
    return (
        _string_list(audit.get("raw_qc_warning_rules"))
        or _string_list(raw_qc.get("warning_rules"))
    )


def _kurtosis_channels_from_result(result: Mapping[str, Any] | None) -> list[str]:
    if not result:
        return []
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    return _string_list(audit.get("kurtosis_bad_channels"))


def _join_channels(channels: Sequence[str]) -> str:
    return ", ".join(channels) if channels else "None"


def _unique_ordered(*groups: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    merged: list[str] = []
    for group in groups:
        for channel in group:
            if channel in seen:
                continue
            seen.add(channel)
            merged.append(channel)
    return merged


def _count_from_result(result: Mapping[str, Any] | None, fallback: int) -> int:
    if not result:
        return fallback
    audit = result.get("audit") if isinstance(result.get("audit"), Mapping) else {}
    raw_qc = result.get("raw_channel_qc") if isinstance(result.get("raw_channel_qc"), Mapping) else {}
    return max(
        _int_or_default(audit.get("n_rejected"), fallback),
        _int_or_default(raw_qc.get("n_bad_channels"), 0),
    )


def _entry_for_pid(ledger: Mapping[str, Any], pid: str) -> Mapping[str, Any]:
    entries = ledger.get("entries")
    if not isinstance(entries, Mapping):
        return {}
    entry = entries.get(pid)
    return entry if isinstance(entry, Mapping) else {}


def _has_missing_condition_warning(entry: Mapping[str, Any]) -> bool:
    return (
        str(entry.get("completion_warning") or "") == MISSING_EXPECTED_OUTPUTS_WARNING
        or str(entry.get("failure_reason") or "") == MISSING_EXPECTED_OUTPUTS_WARNING
        or str(entry.get("condition_completeness") or "").casefold() == "partial"
    )


def _exclusion_reason(
    entry: Mapping[str, Any],
    result: Mapping[str, Any] | None,
) -> str:
    if result and str(result.get("status") or "").casefold() == "excluded":
        return str(
            result.get("message")
            or result.get("reason")
            or "Raw file was excluded from processing."
        )
    status = str(entry.get("status") or "").casefold()
    if status == "excluded":
        return str(
            entry.get("exclusion_message")
            or entry.get("exclusion_reason")
            or "Raw file was excluded from processing."
        )
    if status == "failed":
        return str(entry.get("failure_message") or entry.get("failure_reason") or "")
    return ""


def _is_legacy_partial_condition_entry(
    state: ProcessingInputState,
    entry: Mapping[str, Any],
) -> bool:
    status = str(entry.get("status") or "").strip().casefold()
    return (
        status == "failed"
        and any(path.exists() for path in state.expected_outputs)
        and any(not path.exists() for path in state.expected_outputs)
    )


def _resolved_path_strings(values: Sequence[Any]) -> set[str]:
    resolved: set[str] = set()
    for value in values:
        try:
            resolved.add(str(Path(str(value)).resolve()))
        except (OSError, TypeError, ValueError):
            continue
    return resolved


def _missing_condition_labels(
    plan: ProcessingPlan,
    state: ProcessingInputState,
    entry: Mapping[str, Any],
) -> list[str]:
    labels = _string_list(entry.get("missing_condition_labels"))
    if labels:
        return labels

    missing_outputs = _string_list(entry.get("missing_outputs"))
    if not missing_outputs and (
        _has_missing_condition_warning(entry)
        or _is_legacy_partial_condition_entry(state, entry)
    ):
        missing_outputs = [str(path) for path in state.expected_outputs if not path.exists()]
    missing = _resolved_path_strings(missing_outputs)
    if not missing:
        return []

    derived: list[str] = []
    for index, output_path in enumerate(state.expected_outputs):
        if str(output_path.resolve()) not in missing:
            continue
        if index < len(plan.condition_labels):
            derived.append(str(plan.condition_labels[index]))
        else:
            derived.append(output_path.parent.name)
    return derived


def _cache_qc_for_entry(project: Any, entry: Mapping[str, Any]) -> Mapping[str, Any]:
    raw_file = entry.get("raw_file")
    if not raw_file:
        return {}
    try:
        raw_path = Path(str(raw_file)).resolve()
        expected_size = int(entry.get("raw_size"))
        expected_mtime_ns = int(entry.get("raw_mtime_ns"))
    except (OSError, TypeError, ValueError):
        return {}

    cache_dir = Path(project.project_root) / ".fpvs_cache" / "preprocessed"
    if not cache_dir.exists():
        return {}

    matches: list[tuple[float, Mapping[str, Any]]] = []
    for meta_path in cache_dir.glob("*.json"):
        try:
            metadata = json.loads(meta_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(metadata, Mapping):
            continue
        payload = metadata.get("payload")
        if not isinstance(payload, Mapping):
            continue
        source_path = payload.get("source_path")
        try:
            same_source = Path(str(source_path)).resolve() == raw_path
        except (OSError, TypeError, ValueError):
            same_source = False
        if not same_source:
            continue
        try:
            source_size = int(payload.get("source_size") or -1)
            source_mtime_ns = int(payload.get("source_mtime_ns") or -1)
        except (TypeError, ValueError):
            continue
        if source_size != expected_size or source_mtime_ns != expected_mtime_ns:
            continue
        try:
            timestamp = meta_path.stat().st_mtime
        except OSError:
            timestamp = 0.0
        matches.append((timestamp, metadata))

    if not matches:
        return {}
    return max(matches, key=lambda item: item[0])[1]


def build_processing_qc_rows(
    project: Any,
    plan: ProcessingPlan,
    results: Sequence[Mapping[str, Any]],
) -> list[dict[str, object]]:
    """Build participant rows from the final ledger plus current run details."""

    ledger = load_ledger(Path(project.project_root))
    results_by_path = _result_by_path(results)
    rows: list[dict[str, object]] = []
    for state in plan.states:
        entry = _entry_for_pid(ledger, state.participant_id)
        cache_entry = _cache_qc_for_entry(project, entry) if entry else {}
        result = results_by_path.get(state.info.path.resolve())
        raw_qc_low_variance_channels = _raw_qc_low_variance_channels_from_result(
            result
        ) or _string_list(
            entry.get("raw_qc_low_variance_channels")
        ) or _string_list(
            cache_entry.get("raw_qc_low_variance_channels")
        )
        raw_qc_manual_removed_channels = _raw_qc_manual_removed_channels_from_result(
            result
        ) or _string_list(
            entry.get("raw_qc_manual_removed_channels")
        ) or _string_list(
            cache_entry.get("raw_qc_manual_removed_channels")
        )
        raw_qc_high_amplitude_channels = _raw_qc_high_amplitude_channels_from_result(
            result
        ) or _string_list(
            entry.get("raw_qc_high_amplitude_channels")
        ) or _string_list(
            cache_entry.get("raw_qc_high_amplitude_channels")
        )
        raw_qc_spatial_outlier_channels = _raw_qc_spatial_outlier_channels_from_result(
            result
        ) or _string_list(
            entry.get("raw_qc_spatial_outlier_channels")
        ) or _string_list(
            cache_entry.get("raw_qc_spatial_outlier_channels")
        )
        raw_qc_warning_rules = _raw_qc_warning_rules_from_result(result) or _string_list(
            entry.get("raw_qc_warning_rules")
        ) or _string_list(
            cache_entry.get("raw_qc_warning_rules")
        )
        raw_qc_channels = _raw_qc_channels_from_result(result) or _string_list(
            entry.get("raw_qc_bad_channels")
        ) or _string_list(
            cache_entry.get("raw_qc_bad_channels")
        )
        if not raw_qc_channels:
            raw_qc_channels = _unique_ordered(
                raw_qc_manual_removed_channels,
                raw_qc_low_variance_channels,
                raw_qc_high_amplitude_channels,
                raw_qc_spatial_outlier_channels,
            )
        if (
            raw_qc_channels
            and not raw_qc_low_variance_channels
            and not raw_qc_high_amplitude_channels
            and not raw_qc_spatial_outlier_channels
        ):
            manual_lookup = {
                channel.casefold() for channel in raw_qc_manual_removed_channels
            }
            raw_qc_low_variance_channels = [
                channel
                for channel in raw_qc_channels
                if channel.casefold() not in manual_lookup
            ]
        kurtosis_channels = _kurtosis_channels_from_result(result) or _string_list(
            entry.get("kurtosis_bad_channels")
        ) or _string_list(
            cache_entry.get("kurtosis_bad_channels")
        )
        interpolated_channels = _channels_from_result(result) or _string_list(
            entry.get("interpolated_channels")
        ) or _string_list(
            cache_entry.get("interpolated_channels")
        )
        status = str(entry.get("status") or "").strip().casefold()
        has_missing_conditions = _has_missing_condition_warning(
            entry
        ) or _is_legacy_partial_condition_entry(state, entry)
        missing_condition_labels = _missing_condition_labels(plan, state, entry)
        if status == "completed" and not interpolated_channels:
            interpolated_channels = _unique_ordered(raw_qc_channels, kurtosis_channels)
        fallback_count = max(
            _int_or_default(entry.get("n_rejected"), 0),
            _int_or_default(cache_entry.get("n_rejected"), 0),
            len(
                _unique_ordered(
                    raw_qc_channels,
                    raw_qc_manual_removed_channels,
                    kurtosis_channels,
                    interpolated_channels,
                )
            ),
        )
        count = _count_from_result(
            result,
            fallback_count,
        )
        included = status == "completed" or has_missing_conditions
        if included and has_missing_conditions:
            included_text = "Included (partial conditions)"
        else:
            included_text = "Included" if included else "Excluded"
        rows.append(
            {
                "PID": state.participant_id,
                "Manually Removed Electrodes": _join_channels(
                    raw_qc_manual_removed_channels
                ),
                "Auto-Detected Removed Electrodes (Low SD)": _join_channels(
                    raw_qc_low_variance_channels
                ),
                "Flagged Removed-Electrode Candidates (High Amplitude)": _join_channels(
                    raw_qc_high_amplitude_channels
                ),
                "Flagged Removed-Electrode Candidates (Spatial Consistency)": _join_channels(
                    raw_qc_spatial_outlier_channels
                ),
                "Kurtosis-Rejected Electrodes": _join_channels(kurtosis_channels),
                "Electrodes Interpolated": _join_channels(interpolated_channels),
                "Total Number of Electrodes removed/rejected": count,
                "Raw QC Warnings": _join_channels(raw_qc_warning_rules),
                "Missing Conditions": _join_channels(missing_condition_labels),
                "Included in Final Set": included_text,
                "Exclusion Reason": _exclusion_reason(entry, result),
            }
        )
    return rows


def export_processing_qc_summary(
    project: Any,
    plan: ProcessingPlan,
    results: Sequence[Mapping[str, Any]],
) -> Path:
    """Write the participant QC summary workbook under the project Quality Check folder."""

    rows = build_processing_qc_rows(project, plan, results)
    target = _quality_check_root(project).resolve() / QC_SUMMARY_FILENAME
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = QC_SUMMARY_SHEET
    worksheet.append(list(QC_SUMMARY_HEADERS))
    for row in rows:
        worksheet.append([row[header] for header in QC_SUMMARY_HEADERS])

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 80)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(target)
    return target


__all__ = [
    "QC_SUMMARY_FILENAME",
    "QC_SUMMARY_HEADERS",
    "QC_SUMMARY_SHEET",
    "QUALITY_CHECK_FOLDER",
    "build_processing_qc_rows",
    "export_processing_qc_summary",
]
