"""This module handles the data quality check that occurs prior to data processing. ."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QEventLoop, QObject, QThread, Signal, Slot, Qt
from PySide6.QtWidgets import (
    QAbstractItemView,
    QHeaderView,
    QMessageBox,
    QSizePolicy,
    QTableWidgetItem,
)

from Main_App.gui.components import make_action_button
from Main_App.io.load_utils import format_bdf_recording_not_started_message
from Main_App.processing.qc_summary_export import QUALITY_CHECK_FOLDER
from Main_App.processing.preflight_qc import (
    HeaderOnlyPreflight,
    PreflightQcFileResult,
    PreflightQcScan,
    scan_preprocessing_qc,
    scan_recording_not_started_files,
)
from Main_App.processing.removed_electrode_detection import (
    REMOVED_ELECTRODE_DETECTION_MODE_MANUAL,
    normalize_manual_removed_electrodes_map,
    parse_electrode_list,
)
from Main_App.projects.preprocessing_settings import (
    normalize_manual_excluded_participants,
)

logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())

_DATA_QUALITY_CHECK_TITLE = "Data Quality Check"
_DATA_QUALITY_SCAN_WAIT_MESSAGE = (
    "FPVS Toolbox is currently checking your data for anything strange. This will only take a moment."
)
_DATA_QUALITY_REVIEW_FLAGS_FILENAME = "Data_Quality_Check_Review_Flags.xlsx"
_DATA_QUALITY_STEP_TOTAL = 5


class _PreflightQcWorker(QObject):
    progress = Signal(str, int, int)
    finished = Signal(object)
    failed = Signal(str)

    def __init__(
        self,
        raw_file_infos: Sequence[Any],
        settings: dict[str, Any],
        skip_paths: Sequence[Path],
        max_workers: int,
    ) -> None:
        super().__init__()
        self._raw_file_infos = list(raw_file_infos)
        self._settings = dict(settings)
        self._skip_paths = [Path(path) for path in skip_paths]
        self._max_workers = max(1, int(max_workers))
        self._cancelled = False

    def cancel(self) -> None:
        self._cancelled = True

    def run(self) -> None:
        try:
            scan = scan_preprocessing_qc(
                self._raw_file_infos,
                self._settings,
                skip_paths=self._skip_paths,
                max_workers=self._max_workers,
                progress=self.progress.emit,
                should_cancel=lambda: self._cancelled,
            )
        except Exception as exc:  # pragma: no cover - defensive signal bridge
            logger.exception("Preprocessing QC scan failed.")
            self.failed.emit(str(exc))
            return
        self.finished.emit(scan)


def _participant_sort_key(value: str) -> tuple[str, int, str]:
    prefix = "".join(ch for ch in value if not ch.isdigit()).casefold()
    digits = "".join(ch for ch in value if ch.isdigit())
    number = int(digits) if digits else -1
    return prefix, number, value.casefold()


def _merge_removed_maps(
    existing: dict[str, list[str]],
    suggestions: dict[str, list[str]],
) -> dict[str, list[str]]:
    merged = normalize_manual_removed_electrodes_map(existing)
    for pid, electrodes in suggestions.items():
        current = merged.get(pid, [])
        seen = {electrode.casefold() for electrode in current}
        for electrode in electrodes:
            if electrode.casefold() in seen:
                continue
            current.append(electrode)
            seen.add(electrode.casefold())
        merged[pid] = current
    return dict(sorted(merged.items(), key=lambda item: _participant_sort_key(item[0])))


def _casefold_electrode_lookup(
    values: dict[str, list[str]],
    participant_id: str,
) -> list[str]:
    if participant_id in values:
        return list(values[participant_id])
    key = participant_id.casefold()
    for candidate, electrodes in values.items():
        if candidate.casefold() == key:
            return list(electrodes)
    return []


def _path_strings(items: Sequence[HeaderOnlyPreflight]) -> list[str]:
    return [str(item.path.resolve()) for item in items]


def _path_key(path: Path) -> str:
    try:
        return str(path.resolve()).casefold()
    except (OSError, RuntimeError, ValueError):
        return str(path).casefold()


def _scan_progress_text(message: str | None = None) -> str:
    detail = (message or "").strip()
    if detail:
        return f"{_DATA_QUALITY_SCAN_WAIT_MESSAGE}\n\n{detail}"
    return _DATA_QUALITY_SCAN_WAIT_MESSAGE


def _show_data_quality_notice(
    host: Any,
    heading: str,
    message: str,
    *,
    details: str = "",
) -> None:
    box = QMessageBox(host)
    box.setIcon(QMessageBox.Information)
    box.setWindowTitle(_DATA_QUALITY_CHECK_TITLE)
    box.setText(heading)
    box.setInformativeText(message)
    if details:
        box.setDetailedText(details)
    box.setStandardButtons(QMessageBox.Ok)
    box.exec()


def _set_label(host: Any, attr_name: str, text: str) -> None:
    label = getattr(host, attr_name, None)
    if label is not None:
        label.setText(text)


def _set_progress(host: Any, completed: int, total: int) -> None:
    bar = getattr(host, "progress_bar", None)
    if bar is None:
        return
    pct = 0 if total <= 0 else round(max(0, completed) / max(1, total) * 100)
    bar.setRange(0, 100)
    bar.setValue(max(0, min(100, pct)))
    bar.setFormat("%p%")


def _set_spinner_running(host: Any, running: bool) -> None:
    spinner = getattr(host, "processing_spinner", None)
    if spinner is None:
        return
    if running:
        spinner.start()
    else:
        spinner.stop()
        spinner.update()


def _set_card_title(host: Any, attr_name: str, title: str) -> None:
    card = getattr(host, attr_name, None)
    if card is None:
        return
    try:
        card.header.title_label.setText(title)
    except RuntimeError:
        pass


def _set_review_visible(host: Any, visible: bool, *, title: str = "Review") -> None:
    card = getattr(host, "processing_files_card", None)
    if card is not None:
        card.setVisible(visible)
    if visible:
        _set_card_title(host, "processing_files_card", title)


def _set_progress_visible(host: Any, visible: bool) -> None:
    bar = getattr(host, "progress_bar", None)
    if bar is not None:
        bar.setVisible(visible)
    heading = getattr(host, "processing_progress_heading_label", None)
    if heading is not None:
        heading.setVisible(visible)


def _set_data_quality_sections(
    host: Any,
    *,
    summary_heading: str,
    live_heading: str,
    checklist: Sequence[str],
) -> None:
    _set_label(host, "processing_summary_heading_label", summary_heading)
    _set_label(host, "processing_live_heading_label", live_heading)
    _set_label(host, "processing_progress_heading_label", "Progress")
    _set_label(host, "processing_checklist_heading_label", "Checks in this step")
    checklist_panel = getattr(host, "processing_checklist_panel", None)
    if checklist_panel is not None:
        checklist_panel.setVisible(bool(checklist))
    _set_label(
        host,
        "processing_checklist_label",
        "\n".join(f"- {item}" for item in checklist),
    )


def _begin_preflight_page(
    host: Any,
    *,
    step: int,
    title: str,
    message: str,
    busy: bool,
    review_visible: bool,
    review_title: str = "Review",
    progress_visible: bool = True,
    checklist: Sequence[str] = (),
) -> None:
    if hasattr(host, "_busy_start"):
        host._busy_start()
    if hasattr(host, "_set_controls_enabled"):
        host._set_controls_enabled(False)
    _set_label(host, "processing_title_label", _DATA_QUALITY_CHECK_TITLE)
    _set_data_quality_sections(
        host,
        summary_heading="What to do now",
        live_heading="Live status",
        checklist=checklist,
    )
    step_label = getattr(host, "processing_step_label", None)
    if step_label is not None:
        step_label.setText(f"Step {step} of {_DATA_QUALITY_STEP_TOTAL}: {title}")
        step_label.setVisible(True)
    _set_label(host, "processing_message_label", message)
    _set_label(host, "processing_summary_label", "Preparing data quality checks...")
    _set_label(host, "processing_current_file_label", "Latest file: Not started")
    _set_card_title(host, "processing_status_card", title)
    _set_review_visible(host, review_visible, title=review_title)
    _set_progress_visible(host, progress_visible)
    _set_spinner_running(host, busy)
    _set_progress(host, 0, 1)
    button = getattr(host, "btn_start", None)
    if button is not None:
        button.hide()


def _clear_preflight_actions(host: Any) -> None:
    layout = getattr(host, "processing_action_layout", None)
    slot = getattr(host, "processing_action_slot", None)
    for button in list(getattr(host, "_preflight_qc_action_buttons", []) or []):
        try:
            if layout is not None and layout.indexOf(button) >= 0:
                layout.removeWidget(button)
            button.deleteLater()
        except RuntimeError:
            continue
    host._preflight_qc_action_buttons = []
    if slot is not None:
        slot.setMinimumWidth(0)
        slot.setMaximumWidth(16777215)
        slot.setVisible(False)


def _install_preflight_actions(
    host: Any,
    actions: Sequence[tuple[str, str, str]],
    callback: Any,
) -> None:
    _clear_preflight_actions(host)
    layout = getattr(host, "processing_action_layout", None)
    slot = getattr(host, "processing_action_slot", None)
    if layout is None or slot is None:
        return
    buttons = []
    for label, choice, variant in actions:
        button = make_action_button(label, variant=variant, parent=slot)
        button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        button.clicked.connect(lambda _checked=False, value=choice: callback(value))
        layout.addWidget(button, 0, Qt.AlignRight)
        buttons.append(button)
    if buttons:
        width = max(button.sizeHint().width() for button in buttons)
        width = max(176, min(width, 260))
        for button in buttons:
            button.setMinimumWidth(width)
        slot.setMinimumWidth(width)
        slot.setMaximumWidth(width)
        slot.setVisible(True)
    host._preflight_qc_action_buttons = buttons


def _set_preflight_table(
    host: Any,
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    *,
    editable_last_column: bool = False,
) -> None:
    table = getattr(host, "processing_files_table", None)
    if table is None:
        return
    table.clearContents()
    table.setWordWrap(True)
    table.setColumnCount(len(headers))
    table.setHorizontalHeaderLabels(list(headers))
    header = table.horizontalHeader()
    for column_index in range(len(headers)):
        mode = (
            QHeaderView.Stretch
            if column_index == len(headers) - 1
            else QHeaderView.ResizeToContents
        )
        header.setSectionResizeMode(column_index, mode)
    table.setRowCount(len(rows))
    table.setEditTriggers(
        QAbstractItemView.AllEditTriggers
        if editable_last_column
        else QAbstractItemView.NoEditTriggers
    )
    table.setSelectionMode(
        QAbstractItemView.SingleSelection
        if editable_last_column
        else QAbstractItemView.NoSelection
    )
    for row_index, row_values in enumerate(rows):
        for column_index, value in enumerate(row_values):
            item = QTableWidgetItem(str(value))
            if not editable_last_column or column_index != len(headers) - 1:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if column_index == 0:
                item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row_index, column_index, item)
    table.resizeRowsToContents()
    table.scrollToTop()


def _await_preflight_choice(
    host: Any,
    actions: Sequence[tuple[str, str, str]],
) -> str:
    if not actions:
        return ""
    if (
        getattr(host, "processing_action_layout", None) is None
        or getattr(host, "processing_action_slot", None) is None
    ):
        return actions[0][1]

    loop = QEventLoop(host)
    result = {"choice": ""}

    def _choose(choice: str) -> None:
        result["choice"] = choice
        loop.quit()

    _set_spinner_running(host, False)
    _install_preflight_actions(host, actions, _choose)
    loop.exec()
    _clear_preflight_actions(host)
    return result["choice"]


def _file_name_from_progress(message: str) -> str:
    for prefix in ("Scanning ", "Finished "):
        if message.startswith(prefix):
            return message[len(prefix) :].strip().casefold()
    return ""


def _update_scan_row(host: Any, message: str) -> None:
    file_name = _file_name_from_progress(message)
    if not file_name:
        return
    rows = getattr(host, "_preflight_qc_file_rows", {}) or {}
    row = rows.get(file_name)
    table = getattr(host, "processing_files_table", None)
    if table is None or row is None:
        return
    status_item = table.item(row, 0)
    if status_item is None:
        status_item = QTableWidgetItem()
        table.setItem(row, 0, status_item)
    status_item.setText("Checked" if message.startswith("Finished ") else "Scanning")
    status_item.setTextAlignment(Qt.AlignCenter)
    table.scrollToItem(status_item)


class _PreflightQcEmbeddedBridge(QObject):
    """Keep preflight worker signals marshalled through the GUI thread."""

    def __init__(
        self,
        host: Any,
        thread: QThread,
        result_holder: dict[str, Any],
        loop: QEventLoop,
    ) -> None:
        super().__init__(host)
        self._host = host
        self._thread = thread
        self._result_holder = result_holder
        self._loop = loop

    @Slot(str, int, int)
    def on_progress(self, message: str, completed: int, total: int) -> None:
        _set_progress(self._host, completed, total)
        _set_label(self._host, "processing_summary_label", _DATA_QUALITY_SCAN_WAIT_MESSAGE)
        _set_label(self._host, "processing_current_file_label", message)
        _update_scan_row(self._host, message)

    @Slot(object)
    def on_finished(self, scan: object) -> None:
        self._result_holder["scan"] = scan
        _set_progress(self._host, 1, 1)
        _set_label(
            self._host,
            "processing_current_file_label",
            "Finalizing data quality review...",
        )
        self._thread.quit()

    @Slot(str)
    def on_failed(self, message: str) -> None:
        self._result_holder["error"] = message
        self._thread.quit()

    @Slot()
    def on_thread_finished(self) -> None:
        self._loop.quit()


def _confirm_recording_not_started(
    host: Any,
    flagged: Sequence[HeaderOnlyPreflight],
) -> bool:
    if not flagged:
        return True
    names = [item.path.name for item in flagged]
    _show_data_quality_notice(
        host,
        "Some files do not contain recording data.",
        'These files look like BioSemi recordings that were created, but no data '
        'was actually recorded. The most likely reason is that the experiment '
        'administrator forgot to click "Start Recording".',
        details="\n".join(names),
    )
    _begin_preflight_page(
        host,
        step=1,
        title="Check Raw Files",
        message="FPVS Toolbox found files that do not contain recording data.",
        busy=False,
        review_visible=True,
        review_title="Files to Exclude",
        progress_visible=False,
        checklist=(
            "Confirm files that contain no recording samples",
            "Exclude empty files from processing",
            "Keep the original raw files unchanged",
        ),
    )
    _set_label(
        host,
        "processing_summary_label",
        "It appears that the following files were created, but no data actually "
        "exists inside the files.",
    )
    _set_label(
        host,
        "processing_current_file_label",
        'The most likely explanation is that the experiment administrator forgot to push "Start Recording" on BioSemi.',
    )
    _set_preflight_table(
        host,
        ["File", "Recommended action"],
        [(name, "Exclude from processing") for name in names],
    )
    choice = _await_preflight_choice(
        host,
        (
            ("Exclude Files", "exclude", "primary"),
            ("Cancel Processing", "cancel", "secondary"),
        ),
    )
    if choice != "exclude":
        try:
            host.log("Data quality check cancelled at recording-not-started review.")
        except (AttributeError, TypeError, RuntimeError):
            pass
        return False
    try:
        host.log(format_bdf_recording_not_started_message(names), level=logging.WARNING)
    except (AttributeError, TypeError, RuntimeError):
        pass
    return True


def _run_scan_embedded(
    host: Any,
    raw_file_infos: Sequence[Any],
    params: dict[str, Any],
    *,
    skip_paths: Sequence[Path],
) -> PreflightQcScan | None:
    skip_keys = {_path_key(Path(path)) for path in skip_paths}
    remaining = [
        info
        for info in raw_file_infos
        if _path_key(Path(info.path)) not in skip_keys
    ]
    if not remaining:
        return PreflightQcScan(results=())

    _begin_preflight_page(
        host,
        step=2,
        title="Scan Signal Health",
        message=_DATA_QUALITY_SCAN_WAIT_MESSAGE,
        busy=True,
        review_visible=False,
        checklist=(
            "Look for electrodes that appear physically disconnected",
            "Check for participant-level signal failures",
            "Screen for unusual narrowband spectral peaks",
        ),
    )
    host._preflight_qc_file_rows = {}
    try:
        max_workers = max(1, int(getattr(host, "max_workers", 1) or 1))
    except (TypeError, ValueError):
        max_workers = 1
    worker_count = min(max_workers, len(remaining))
    try:
        host.log(
            f"Data quality scan using {worker_count} parallel worker(s).",
            level=logging.INFO,
        )
    except (AttributeError, TypeError, RuntimeError):
        pass
    _set_label(host, "processing_summary_label", _DATA_QUALITY_SCAN_WAIT_MESSAGE)
    _set_label(
        host,
        "processing_current_file_label",
        f"Starting data quality scan with {worker_count} parallel worker(s)...",
    )
    _set_progress(host, 0, len(remaining))

    thread = QThread(host)
    worker = _PreflightQcWorker(
        remaining,
        params,
        skip_paths,
        max_workers=worker_count,
    )
    worker.moveToThread(thread)
    result_holder: dict[str, Any] = {}
    loop = QEventLoop(host)
    bridge = _PreflightQcEmbeddedBridge(host, thread, result_holder, loop)
    host._preflight_qc_thread = thread
    host._preflight_qc_worker = worker
    host._preflight_qc_bridge = bridge

    def _request_cancel(_choice: str) -> None:
        result_holder["cancelled"] = True
        worker.cancel()
        _clear_preflight_actions(host)
        _set_label(
            host,
            "processing_current_file_label",
            "Cancelling after the current file...",
        )

    _install_preflight_actions(
        host,
        (("Cancel Check", "cancel", "secondary"),),
        _request_cancel,
    )
    worker.progress.connect(bridge.on_progress)
    worker.finished.connect(bridge.on_finished)
    worker.failed.connect(bridge.on_failed)
    thread.started.connect(worker.run)
    worker.finished.connect(worker.deleteLater)
    worker.failed.connect(worker.deleteLater)
    thread.finished.connect(bridge.on_thread_finished)
    thread.finished.connect(thread.deleteLater)
    thread.start()

    loop.exec()

    _clear_preflight_actions(host)
    host._preflight_qc_thread = None
    host._preflight_qc_worker = None
    host._preflight_qc_bridge = None

    error = result_holder.get("error")
    if error:
        _set_label(host, "processing_summary_label", "Data quality check could not complete.")
        _set_label(host, "processing_current_file_label", str(error))
        _await_preflight_choice(
            host,
            (("Cancel Processing", "cancel", "primary"),),
        )
        return None

    if result_holder.get("cancelled"):
        return None

    scan = result_holder.get("scan")
    return scan if isinstance(scan, PreflightQcScan) else None


def _review_removed_electrodes(
    host: Any,
    raw_file_infos: Sequence[Any],
    params: dict[str, Any],
    scan: PreflightQcScan,
) -> bool:
    existing = normalize_manual_removed_electrodes_map(
        params.get("manual_removed_electrodes")
    )
    prepopulated = _merge_removed_maps(existing, scan.suggested_removed_electrodes)
    participant_ids = [str(info.subject_id) for info in raw_file_infos]
    prompt = (
        "FPVS Toolbox detected that the following electrodes were physically "
        "removed from the cap prior to the start of each respective experiment. "
        "Please review and confirm this list, and add any electrodes that were "
        "removed if they are not present on this list."
    )
    _show_data_quality_notice(
        host,
        "Review electrodes that may have been removed before recording.",
        "FPVS Toolbox will show a participant-by-participant table next. "
        "Please confirm the list and add any electrodes that were removed but "
        "are missing from the table.",
    )
    _begin_preflight_page(
        host,
        step=3,
        title="Confirm Removed Electrodes",
        message="Review the removed-electrode list before processing begins.",
        busy=False,
        review_visible=True,
        review_title="Removed Electrodes",
        progress_visible=False,
        checklist=(
            "Review electrodes detected as removed before recording",
            "Add any missing removed electrodes",
            "Save the confirmed list into project settings",
        ),
    )
    _set_label(host, "processing_summary_label", prompt)
    _set_label(
        host,
        "processing_current_file_label",
        "Edit the removed-electrode list directly in the table, then save to continue.",
    )
    seen_pids: set[str] = set()
    all_pids: list[str] = []
    for source_pid in (*participant_ids, *tuple(prepopulated)):
        pid = str(source_pid).strip()
        if not pid:
            continue
        key = pid.casefold()
        if key in seen_pids:
            continue
        seen_pids.add(key)
        all_pids.append(pid)
    all_pids.sort(key=_participant_sort_key)
    rows = [
        (pid, ", ".join(_casefold_electrode_lookup(prepopulated, pid)))
        for pid in all_pids
    ]
    _set_preflight_table(
        host,
        ["PID", "Removed electrodes"],
        rows,
        editable_last_column=True,
    )
    choice = _await_preflight_choice(
        host,
        (
            ("Save / Next", "save", "primary"),
            ("Cancel Processing", "cancel", "secondary"),
        ),
    )
    if choice != "save":
        try:
            host.log("Data quality check cancelled at removed-electrode review.")
        except (AttributeError, TypeError, RuntimeError):
            pass
        return False

    updated_map = normalize_manual_removed_electrodes_map(
        _manual_removed_electrodes_from_table(host)
    )
    updated_preproc = dict(getattr(host.currentProject, "preprocessing", {}) or {})
    updated_preproc["removed_electrode_detection_mode"] = (
        REMOVED_ELECTRODE_DETECTION_MODE_MANUAL
    )
    updated_preproc["manual_removed_electrodes"] = updated_map
    try:
        normalized = host.currentProject.update_preprocessing(updated_preproc)
        host.currentProject.save()
    except ValueError as exc:
        QMessageBox.warning(host, "Invalid Manual Removed Electrodes", str(exc))
        return False
    except OSError as exc:
        logger.exception("Failed to save manual removed-electrode settings.")
        QMessageBox.critical(host, "Project Save Error", str(exc))
        return False

    params["manual_removed_electrodes"] = dict(
        normalized.get("manual_removed_electrodes") or {}
    )
    params["removed_electrode_detection_mode"] = normalized.get(
        "removed_electrode_detection_mode"
    )
    params["auto_detect_removed_electrodes"] = bool(
        normalized.get("auto_detect_removed_electrodes")
    )
    host.validated_params = params
    try:
        host.log("Data quality check saved the reviewed removed-electrode list.")
    except (AttributeError, TypeError, RuntimeError):
        pass
    return True


def _manual_removed_electrodes_from_table(host: Any) -> dict[str, list[str]]:
    table = getattr(host, "processing_files_table", None)
    if table is None:
        return {}
    values: dict[str, list[str]] = {}
    for row in range(table.rowCount()):
        pid_item = table.item(row, 0)
        electrodes_item = table.item(row, 1)
        pid = pid_item.text().strip() if pid_item else ""
        if not pid:
            continue
        electrodes_text = electrodes_item.text() if electrodes_item else ""
        values[pid] = parse_electrode_list(electrodes_text)
    return values


def _hard_candidate_reason(result: PreflightQcFileResult) -> str:
    if result.raw_qc_excluded:
        return result.raw_qc_message or "Raw channel-health check failed."
    if result.raw_spectral_widespread:
        return (
            result.raw_spectral_message
            or "Widespread raw spectral artifact detected before preprocessing."
        )
    return "Participant-level data quality exclusion criterion met."


def _confirm_hard_exclusions(
    host: Any,
    params: dict[str, Any],
    scan: PreflightQcScan,
) -> set[str]:
    candidates = scan.hard_exclusion_candidates
    if not candidates:
        return set()
    _show_data_quality_notice(
        host,
        "Review participants that may need to be excluded.",
        "FPVS Toolbox found participant-level data quality problems. The next "
        "screen lists the candidates and lets you add them to the manual "
        "participant exclusion list before processing starts.",
    )
    _begin_preflight_page(
        host,
        step=4,
        title="Confirm Participant Exclusions",
        message="Review participant-level data quality failures before processing begins.",
        busy=False,
        review_visible=True,
        review_title="Participant Exclusions",
        progress_visible=False,
        checklist=(
            "Review participants with hard data quality failures",
            "Add confirmed cases to the participant exclusion list",
            "Leave uncertain cases available if you want to inspect them later",
        ),
    )
    _set_label(
        host,
        "processing_summary_label",
        "FPVS Toolbox found participant-level data quality failures that should "
        "not enter the processed dataset.",
    )
    _set_label(
        host,
        "processing_current_file_label",
        "Review the candidates below. You can add them to the manual participant "
        "exclusion list or continue without changing the list.",
    )
    _set_preflight_table(
        host,
        ["PID", "Reason"],
        [
            (result.participant_id, _hard_candidate_reason(result))
            for result in candidates
        ],
    )
    choice = _await_preflight_choice(
        host,
        (
            ("Add Exclusions", "add", "primary"),
            ("Continue Without Adding", "skip", "secondary"),
        ),
    )
    if choice != "add":
        return set()

    current = normalize_manual_excluded_participants(
        params.get("manual_excluded_participants")
    )
    updated = normalize_manual_excluded_participants(
        [*current, *(result.participant_id for result in candidates)]
    )
    updated_preproc = dict(getattr(host.currentProject, "preprocessing", {}) or {})
    updated_preproc["manual_excluded_participants"] = updated
    try:
        normalized = host.currentProject.update_preprocessing(updated_preproc)
        host.currentProject.save()
    except (OSError, ValueError, RuntimeError) as exc:
        logger.exception("Failed to save participant exclusions.")
        QMessageBox.critical(
            host,
            "Project Save Error",
            f"Could not save participant exclusions: {exc}",
        )
        return set()
    params["manual_excluded_participants"] = list(
        normalized.get("manual_excluded_participants") or []
    )
    host.validated_params = params
    accepted = {result.participant_id.casefold() for result in candidates}
    try:
        host.log(
            "Data quality check added participant exclusion(s): "
            + ", ".join(params["manual_excluded_participants"]),
            level=logging.WARNING,
        )
    except (AttributeError, TypeError, RuntimeError):
        pass
    return accepted


def _quality_check_dir(host: Any) -> Path:
    project = getattr(host, "currentProject", None)
    root = Path(getattr(project, "project_root", "."))
    return root / QUALITY_CHECK_FOLDER


def _style_preflight_review_sheet(worksheet: Any) -> None:
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12),
            80,
        )


def _write_preflight_review_flags(
    host: Any,
    rows: Sequence[tuple[str, str, str]],
) -> Path:
    target = _quality_check_dir(host).resolve() / _DATA_QUALITY_REVIEW_FLAGS_FILENAME
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Review Flags"
    worksheet.append(("PID", "Source File", "Flagged Item"))
    for row in rows:
        worksheet.append(row)
    _style_preflight_review_sheet(worksheet)
    workbook.save(target)
    return target


def _show_suspicious_remainder(
    host: Any,
    scan: PreflightQcScan,
    accepted_hard_exclusions: set[str],
) -> bool:
    rows: list[tuple[str, str, str]] = []
    for result in scan.suspicious_results:
        if result.participant_id.casefold() in accepted_hard_exclusions:
            continue
        fragments: list[str] = []
        if result.load_error:
            fragments.append(f"could not be scanned ({result.load_error})")
        if result.warning_rules:
            fragments.append(
                "raw data warning rule(s): " + ", ".join(result.warning_rules)
            )
        if result.high_amplitude_channels:
            fragments.append(
                "high-amplitude channel(s): "
                + ", ".join(result.high_amplitude_channels)
            )
        if result.spatial_outlier_channels:
            fragments.append(
                "spatially inconsistent channel(s): "
                + ", ".join(result.spatial_outlier_channels)
            )
        if result.raw_spectral_flagged_channels and not result.raw_spectral_widespread:
            fragments.append(
                "localized raw spectral flag(s): "
                + ", ".join(result.raw_spectral_flagged_channels[:8])
            )
        if fragments:
            rows.append((result.participant_id, result.path.name, "; ".join(fragments)))
    if not rows:
        return True

    report_path: Path | None = None
    report_message = ""
    try:
        report_path = _write_preflight_review_flags(host, rows)
        report_message = f"Review flags saved to: {report_path}"
        try:
            host.log(report_message, level=logging.WARNING)
        except (AttributeError, TypeError, RuntimeError):
            pass
    except OSError as exc:
        logger.exception("Failed to save data quality review flags workbook.")
        report_message = f"Could not save review flags workbook: {exc}"

    _show_data_quality_notice(
        host,
        "Some items should be reviewed later.",
        "FPVS Toolbox found data quality items that were not automatically "
        "removed. These items will not stop processing, but you should review "
        "them before relying on the affected results.",
        details=report_message,
    )
    _begin_preflight_page(
        host,
        step=5,
        title="Review Remaining Flags",
        message="FPVS Toolbox found items that were flagged for review but were not removed automatically.",
        busy=False,
        review_visible=True,
        review_title="Review Flags",
        progress_visible=False,
        checklist=(
            "Review items that were flagged but not removed",
            "Use the saved workbook as a later review checklist",
            "Continue processing when you have noted the flagged items",
        ),
    )
    _set_label(
        host,
        "processing_summary_label",
        "FPVS Toolbox found items that were flagged for review but were not removed automatically.",
    )
    _set_label(
        host,
        "processing_current_file_label",
        f"Please make note of these and manually investigate them later. {report_message}",
    )
    _set_preflight_table(host, ["PID", "Source File", "Flagged item"], rows)
    choice = _await_preflight_choice(
        host,
        (
            ("Continue Processing", "continue", "primary"),
            ("Cancel Processing", "cancel", "secondary"),
        ),
    )
    return choice == "continue"


def run_preprocessing_qc_workflow(
    host: Any,
    raw_file_infos: Sequence[Any],
    params: dict[str, Any],
) -> bool:
    """Run the embedded pre-processing QC review phases."""

    if not raw_file_infos:
        return True

    _show_data_quality_notice(
        host,
        "FPVS Toolbox will check your data before processing.",
        "This guided check looks for empty recording files, electrodes that may "
        "have been physically removed before recording, and participant-level "
        "signal problems. You will be asked to confirm anything that could "
        "change what gets processed.",
    )
    _begin_preflight_page(
        host,
        step=1,
        title="Check Raw Files",
        message=_DATA_QUALITY_SCAN_WAIT_MESSAGE,
        busy=True,
        review_visible=False,
        checklist=(
            "Check whether each BDF contains real recording data",
            "Find files created when recording was not started",
            "Prepare the deeper signal-health scan",
        ),
    )
    _set_label(host, "processing_summary_label", "Checking BDF headers...")
    _set_label(
        host,
        "processing_current_file_label",
        "Looking for empty recordings before the deeper data quality scan.",
    )
    header_only = scan_recording_not_started_files(raw_file_infos)
    if header_only and not _confirm_recording_not_started(host, header_only):
        return False
    params["_fpvs_preflight_recording_not_started_files"] = _path_strings(header_only)
    header_only_keys = {_path_key(item.path) for item in header_only}
    active_infos = [
        info
        for info in raw_file_infos
        if _path_key(Path(info.path)) not in header_only_keys
    ]

    scan = _run_scan_embedded(
        host,
        active_infos,
        params,
        skip_paths=[item.path for item in header_only],
    )
    if scan is None or scan.cancelled:
        return False

    if active_infos and not _review_removed_electrodes(host, active_infos, params, scan):
        return False

    accepted_hard_exclusions = _confirm_hard_exclusions(host, params, scan)
    if not _show_suspicious_remainder(host, scan, accepted_hard_exclusions):
        return False
    return True


__all__ = ["run_preprocessing_qc_workflow"]
